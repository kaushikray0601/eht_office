import json
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from eht.data_service import store_calculated_results
from eht.heat_loss_methods import DEFAULT_HEAT_LOSS_METHOD
from eht.models import BOQ, HeatTracingInput, ManagedProject, ProjectData
from eht.sld_payload import build_project_sld_payload


def make_project(project_id='p1'):
    ManagedProject.objects.get_or_create(
        proj_id=project_id,
        defaults={'description': f'Project {project_id}', 'is_active': True},
    )
    return ProjectData.objects.create(
        proj_id=project_id,
        min_amb_t=20.0,
        max_amb_t=45.0,
        startup_t=15.0,
        area_class='SAFE',
        temp_class='T3',
        voltage=230.0,
        max_cb_size=10,
        restrict_cb_current=80.0,
        vendor='CHR',
        spiral_wrap_allowed=True,
        spiral_factor=2.0,
        margin_on_tracer_lengths=10.0,
        voltage_var_factor=0.0,
        res_tol=10.0,
        termination_margin=250.0,
        heat_loss_sf=1.2,
        heat_loss_method=DEFAULT_HEAT_LOSS_METHOD,
        rtd_thrm='TI',
        wind_speed=32.0,
        req_local_isolator='required',
        caution_label_interval=10.0,
        isolator_location='bothSides',
        ckt_ln=30.0,
        loop_ln=12.0,
        allowablevdrop=5.0,
    )


def make_input_line(project_id, line_id):
    return HeatTracingInput.objects.create(
        proj_id=project_id,
        line_id=line_id,
        service_type='EP',
        line_size=2.0,
        line_length=10.0,
        ins_mat_type='Mineral Wool',
        insul_thick=50.0,
        maint_temp=120.0,
        oper_temp=100.0,
        design_temp=140.0,
        status='confirmed',
    )


def make_reporting_snapshot(project_id='p1'):
    make_project(project_id)
    selected_line = make_input_line(project_id, 'LINE-SELECTED')
    rejected_line = make_input_line(project_id, 'LINE-REJECTED')
    conductivity_basis = {
        'effective_method': DEFAULT_HEAT_LOSS_METHOD,
        'effective_method_label': 'Mean insulation temperature (recommended)',
        'rule_set': 'MEAN_INSULATION_TEMPERATURE_AMBIENT_SURFACE_V1',
    }

    store_calculated_results(project_id, {
        'heat_loss': [
            {
                'uid': selected_line.uid,
                'heat_loss': 12.0,
                'base_heat_loss': 10.0,
                'design_heat_loss': 12.0,
                'heat_loss_sf': 1.2,
                'pipe_size_mm': 60.3,
                'conductivity': 0.052,
                'conductivity_basis': conductivity_basis,
                'wind_correction': 1.0,
                'accessory_adders': {'total': 1.2},
                'selection_status': 'selected',
                'selection_rejection_reasons': [],
                'tracer_adder': 1.2,
            },
            {
                'uid': rejected_line.uid,
                'heat_loss': 18.0,
                'base_heat_loss': 15.0,
                'design_heat_loss': 18.0,
                'heat_loss_sf': 1.2,
                'pipe_size_mm': 60.3,
                'conductivity': 0.052,
                'conductivity_basis': conductivity_basis,
                'wind_correction': 1.0,
                'accessory_adders': {'total': 1.2},
                'selection_status': 'rejected',
                'selection_rejection_reasons': [{
                    'rule_set': 'SR_SELECTION_REJECTION_REASON_V1',
                    'code': 'NO_SR_CATALOGUE_VOLTAGE_COMPATIBILITY',
                    'message': 'No SR tracers satisfy nominal voltage compatibility.',
                    'details': {'system_voltage': 240.0},
                }],
                'tracer_adder': 1.2,
            },
        ],
        'selected_tracers': [
            {
                'uid': selected_line.uid,
                'V_UID': 'V-001',
                'A_Coeff': 0.0,
                'B_Coeff': 0.0,
                'C_Coeff': 30.0,
                'Power_at_Startup_T': 11.5,
                'Ohm_per_km': 9.5,
                'Res_corrFactor_Mica': 0.5,
                'Tracer_Family': 'Self Regulating',
                'Voltage_Float': 230.0,
                'Voltage_Correction_Factor': 1.0,
                'Power_Output': 30.0,
                'Spiral_Factor': 1.1,
                'Tracer_Length': 12.0,
                'Tracer_With_Margin': 12.6,
            },
        ],
        'alternative_tracers': [],
        'power_distribution': [
            {
                'uid': selected_line.uid,
                'total_circuits': 1,
                'branches': [
                    {
                        'type': '1phJB',
                        'circuit_count': 1,
                        'connected_to': 'Tracer',
                        'cable_length_db_to_jb': 30.0,
                        'cable_length_jb_to_jb': None,
                        'tagged_components': {
                            'MCB': 'MCB_001',
                            'Downstream': [{'circuit_index': 1, 'Tracer': 'Tracer_001'}],
                        },
                    },
                ],
            },
        ],
        'boq_per_line': {
            selected_line.uid: {'TRACER': 12.85, 'MCB': 1},
        },
        'consolidated_boq': {
            'TRACER': 12.85,
            'MCB': 1,
        },
        'tracer_power_param': [
            {
                'uid': selected_line.uid,
                'breaker_size': 10,
                'no_of_circuits': 1,
                'max_current': 2.5,
                'operating_current': 2.0,
                'operating_load': 460.0,
                'total_tracer_length': 12.85,
                'pipe_size_mm': 60.3,
            },
        ],
    })
    return selected_line, rejected_line


class SRReportingAlignmentTests(TestCase):
    def test_result_view_surfaces_heat_loss_basis_and_selection_diagnostics(self):
        _selected_line, rejected_line = make_reporting_snapshot()

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Design Heat Loss')
        self.assertContains(response, 'Base 10.00 x SF 1.20')
        self.assertContains(response, 'Current / Ckt')
        self.assertContains(response, 'Includes termination allowance')
        self.assertContains(response, 'SR Selection Diagnostics')
        self.assertContains(response, rejected_line.line_id)
        self.assertContains(response, 'NO_SR_CATALOGUE_VOLTAGE_COMPATIBILITY')

    def test_result_export_includes_basis_columns_and_selection_diagnostics_sheet(self):
        _selected_line, rejected_line = make_reporting_snapshot()

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            [
                'Line Results',
                'Selection Diagnostics',
                'Power Distribution',
                'Cold Cable Sizing',
                'Cold Cable 3C Segments',
                'Alternate Tracers',
                'MI Selection',
            ],
        )

        line_header = next(workbook['Line Results'].iter_rows(values_only=True))
        self.assertIn('Design Heat Loss (W/m)', line_header)
        self.assertIn('Starting Current / Circuit (A)', line_header)
        self.assertIn('Ordered SR Tracer Length incl. Termination Allowance (m)', line_header)

        diagnostics_rows = list(workbook['Selection Diagnostics'].iter_rows(values_only=True))
        diagnostics_header = diagnostics_rows[0]
        rejected_row = next(row for row in diagnostics_rows[1:] if row[1] == rejected_line.line_id)
        self.assertEqual(
            rejected_row[diagnostics_header.index('Reason Code')],
            'NO_SR_CATALOGUE_VOLTAGE_COMPATIBILITY',
        )
        self.assertEqual(json.loads(rejected_row[diagnostics_header.index('Reason Details')])['system_voltage'], 240.0)

    def test_boq_tracer_description_records_ordered_sr_length_basis(self):
        make_reporting_snapshot()

        tracer_item = BOQ.objects.get(project_id='p1', scope='consolidated', item_code='TRACER')

        self.assertEqual(
            tracer_item.item_description,
            'Ordered SR heating tracer length (incl. termination allowance)',
        )

    def test_sld_tracer_metadata_carries_sr_calculation_basis(self):
        make_reporting_snapshot()

        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')
        sr_calculation = tracer_node['metadata']['sr_calculation']

        self.assertEqual(sr_calculation['heat_loss']['design_heat_loss'], 12.0)
        self.assertEqual(sr_calculation['heat_loss']['heat_loss_sf'], 1.2)
        self.assertEqual(sr_calculation['electrical']['current_basis'], 'per_circuit')
        self.assertEqual(sr_calculation['electrical']['operating_current_per_circuit'], 2.0)
        self.assertEqual(
            sr_calculation['electrical']['tracer_length_basis'],
            'ordered_length_includes_termination_allowance',
        )
