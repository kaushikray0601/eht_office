import json
import math
from copy import deepcopy
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from uuid import uuid4

import pandas as pd
from django.contrib import admin
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test import SimpleTestCase, TestCase, TransactionTestCase
from django.urls import reverse
from openpyxl import load_workbook

from eht.cal import orchestrate_calculations
from eht.cable_schedule import build_cable_schedule_workspace_data
from eht.calculations.boq import compute_bill_of_quantities
from eht.calculations.heat_loss import calculate_heat_loss
from eht.calculations.power_distribution import compute_mi_power_params, compute_power_distribution, compute_power_params
from eht.calculations.tag_management import ProjectTagFactory
from eht.calculations.tracer_selection import (
    calculate_voltage_scenarios,
    filter_sr_catalogue_voltage_compatibility,
    filter_sr_catalogue_suitability,
    get_tracer_options,
)
from eht.cold_cable import (
    build_cold_cable_sizing_input,
    calculate_k_temp,
    calculate_vd,
    check_fault_l_pe,
    ColdCableSizingInput,
    conductor_density_kg_m3,
    conductor_operating_temperature,
    conductor_temperature_coefficient,
    find_minimum_cable_for_ampacity,
    get_conductor_resistance_at_temp,
    mcb_instantaneous_factor,
    optimise_cable_pair,
    resolve_cable_lengths,
    SOURCE_IMPEDANCE_ASSUMED_ZERO_NOTE,
    size_cold_cable_for_branch,
    size_cold_cables_for_project,
)
from eht.cold_cable_readiness import cold_cable_method_readiness
from eht.data_service import clear_project_workspace_data, fetch_process_lines, store_calculated_results
from eht.forms import PROJECT_FORM_COLD_CABLE_DEFAULTS, ProjectDataForm
from eht.models import (
    AlternateTracer,
    BOQ,
    CableScheduleOverride,
    ColdCableCatalogue,
    ColdCableResult,
    DEFAULT_PROJECT_ID,
    ElecEHT_ASMEB36,
    ElecEHT_ThermalConductivity,
    ElecEHT_Vendor,
    HeatLoss,
    HeatTracingInput,
    ManagedProject,
    MAX_CB_SIZE,
    MICableFamily,
    MICableHeater,
    MIColdLeadOption,
    PowerDistribution,
    PowerDistributionBranch,
    ProcessLineCalculation,
    ProjectData,
    SLDNodeLayout,
    SLDTopologyEdit,
    SelectedMIHeater,
    SelectedTracer,
    TracerSelectionOverride,
    is_default_project_id,
)
from eht.views import _build_panel_load_summary
from eht.sld_layout import get_project_sld_layout
from eht.sld_payload import SLD_GRAPH_SCHEMA_VERSION, build_project_sld_payload
from eht.sld_pdf import build_sld_pdf
from eht.sld_schema import audit_tagged_component_schema
from eht.sld_topology import payload_fingerprint
from eht.sld_topology_history import compact_sld_topology_history, compact_topology_edit_record
from eht.sld_validation import validate_project_sld_payload
from eht.pipeline import run_project_calculations
from eht.heat_loss_methods import DEFAULT_HEAT_LOSS_METHOD


BASE_DIR = Path(__file__).resolve().parent.parent


def make_line(**overrides):
    line = {
        'uid': 'L1',
        'maint_temp': 100.0,
        'insul_thick': 50.0,
        'ins_mat_type': 'MW',
        'line_size': 2.0,
        'valve_qty': 2,
        'support_qty': 3,
        'flange_qty': 1,
        'line_length': 10.0,
        'oper_temp': 80.0,
        'design_temp': 120.0,
        'service_type': 'EP',
    }
    line.update(overrides)
    return line


class SldWorkspaceJavaScriptTests(SimpleTestCase):
    def _source(self):
        return (BASE_DIR / 'static' / 'js' / 'sld_workspace.js').read_text()

    def test_render_guard_releases_after_render_exceptions(self):
        source = self._source()

        self.assertIn('function beginSldRender(root)', source)
        self.assertIn('root.__sldRenderWatchdog = setTimeout', source)
        self.assertIn('function finishSldRender(root)', source)
        self.assertIn('clearTimeout(root.__sldRenderWatchdog);', source)
        self.assertGreaterEqual(source.count('finally {\n                            finishSldRender(root);\n                        }'), 2)
        self.assertIn("console.error('[SLD] renderSldGraph failed:', renderError);", source)
        self.assertIn('SLD diagram could not be rendered. Check the browser console for details.', source)

    def test_sld_components_have_explicit_pointer_hit_targets(self):
        source = self._source()

        self.assertIn("pointerEvents: 'all'", source)
        self.assertIn("cursor: 'pointer'", source)

    def test_sld_pager_uses_safe_render_gateway(self):
        source = self._source()

        self.assertIn('function safeRenderCurrentSldPage(root, reason)', source)
        self.assertEqual(source.count('renderCurrentSldPage(root);'), 1)
        self.assertIn("safeRenderCurrentSldPage(root, 'page-size-change')", source)
        self.assertIn("safeRenderCurrentSldPage(root, 'page-prev')", source)
        self.assertIn("safeRenderCurrentSldPage(root, 'page-next')", source)

    def test_focused_sld_disables_topology_mutation_but_not_overrides(self):
        source = self._source()

        self.assertIn('function isTopologyEditDisabledForFocusedLine(root)', source)
        self.assertIn('Topology edit is not allowed in filtered SLD view', source)
        self.assertIn('payload.line_id = root.dataset.selectedLineId;', source)
        self.assertIn('function saveCableOverride(root)', source)
        self.assertIn('function saveTracerOverride(root)', source)


class SldTopologyFingerprintTests(SimpleTestCase):
    def test_payload_fingerprint_ignores_volatile_node_metadata(self):
        payload = {
            'schema_version': 1,
            'nodes': [{
                'component_id': 'mcb-1',
                'component_type': 'MCB',
                'line_uid': 'line-1',
                'branch_index': 1,
                'circuit_index': None,
                'display_tag': 'MCB_001',
                'metadata': {'breaker_size': 16, 'cold_cable': {'sizing_status': 'ok'}},
            }],
            'edges': [],
            'line_groups': [{'line_uid': 'line-1', 'line_id': 'L1', 'branch_indices': [1]}],
        }
        changed_metadata = deepcopy(payload)
        changed_metadata['nodes'][0]['metadata'] = {
            'breaker_size': 20,
            'cold_cable': {
                'sizing_status': 'review_required',
                'voltage_drop_percent': 2.5,
            },
        }
        changed_topology = deepcopy(payload)
        changed_topology['nodes'][0]['component_id'] = 'mcb-2'

        self.assertEqual(payload_fingerprint(payload), payload_fingerprint(changed_metadata))
        self.assertNotEqual(payload_fingerprint(payload), payload_fingerprint(changed_topology))


def make_project_settings(**overrides):
    settings = {
        'min_amb_t': 20.0,
        'wind_speed': 32.0,
        'voltage': 230.0,
        'voltage_var_factor': 0.0,
        'margin_on_tracer_lengths': 10.0,
        'spiral_wrap_allowed': True,
        'spiral_factor': 2.0,
        'max_cb_size': 10.0,
        'restrict_cb_current': 80.0,
        'termination_margin': 250.0,
        'ckt_ln': 30.0,
        'loop_ln': 12.0,
        'isolator_location': 'bothSides',
        'rtd_thrm': 'TI',
        'caution_label_interval': 10.0,
        'allowablevdrop': 5.0,
        'vendor': 'CHR',
        'heat_loss_method': DEFAULT_HEAT_LOSS_METHOD,
    }
    settings.update(overrides)
    return settings


def make_asme_table():
    return pd.DataFrame([
        {'Nominal_Pipe_Size': 2.0, 'Outside_Diameter_mm': 60.3},
    ])


def make_thermal_table():
    return pd.DataFrame([
        {'Ins_Mat_Type': 'MW', 'K_factor_A': 0.0, 'K_factor_B': 0.0, 'K_factor_C': 0.05},
    ])


def make_tracer_vendor_data():
    return pd.DataFrame([
        {
            'V_UID': 'T1',
            'Voltage_Float': 230.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Power_at_Startup_T': 10.0,
            'Ohm_per_km': 1.0,
            'Res_corrFactor_Mica': 1.0,
            'Tracer_Family': 'SR',
        },
        {
            'V_UID': 'T2',
            'Voltage_Float': 230.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 80.0,
            'Power_at_Startup_T': 10.0,
            'Ohm_per_km': 1.0,
            'Res_corrFactor_Mica': 1.0,
            'Tracer_Family': 'SR',
        },
        {
            'V_UID': 'T3',
            'Voltage_Float': 230.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 60.0,
            'Power_at_Startup_T': 10.0,
            'Ohm_per_km': 1.0,
            'Res_corrFactor_Mica': 1.0,
            'Tracer_Family': 'SR',
        },
    ])


def make_project_record(**overrides):
    data = {
        'proj_id': 'p1',
        'min_amb_t': 20.0,
        'max_amb_t': 45.0,
        'startup_t': 15.0,
        'area_class': 'SAFE',
        'temp_class': 'T3',
        'voltage': 230.0,
        'max_cb_size': 10,
        'restrict_cb_current': 80.0,
        'vendor': 'CHR',
        'spiral_wrap_allowed': True,
        'spiral_factor': 2.0,
        'margin_on_tracer_lengths': 10.0,
        'voltage_var_factor': 0.0,
        'res_tol': 10.0,
        'termination_margin': 250.0,
        'heat_loss_sf': 1.0,
        'heat_loss_method': DEFAULT_HEAT_LOSS_METHOD,
        'rtd_thrm': 'TI',
        'wind_speed': 32.0,
        'req_local_isolator': 'required',
        'caution_label_interval': 10.0,
        'isolator_location': 'bothSides',
        'ckt_ln': 30.0,
        'loop_ln': 12.0,
        'allowablevdrop': 5.0,
    }
    data.update(overrides)
    ManagedProject.objects.get_or_create(
        proj_id=data['proj_id'],
        defaults={'description': f"Project {data['proj_id']}", 'is_active': True},
    )
    return ProjectData.objects.create(**data)


def make_project_form_payload(**overrides):
    payload = {
        'proj_id': 'PLANT_A_001',
        'vendor': 'CHR',
        'startup_t': '15.00',
        'min_amb_t': '20.00',
        'max_amb_t': '45.00',
        'area_class': 'SAFE',
        'temp_class': 'T3',
        'voltage': '230.00',
        'eht_db_fault_rating_ka_preset': '15',
        'eht_db_fault_rating_ka_custom': '',
        'max_cb_size': '10',
        'restrict_cb_current': '80.00',
        'allowablevdrop': '5.00',
        'cable_standard': 'IEC_60502_1',
        'cable_conductor_material': 'Cu',
        'cable_insulation_type': 'XLPE',
        'cable_install_method': 'E',
        'cable_grouping_derating': '1.000',
        'min_cold_cable_size_mm2': 'CALCULATED',
        'mcb_curve': 'C',
        'rcd_provided': 'on',
        'spiral_factor': '2.00',
        'spiral_wrap_allowed': 'True',
        'margin_on_tracer_lengths': '10.00',
        'voltage_var_factor': '0.00',
        'res_tol': '10.00',
        'termination_margin': '250.00',
        'heat_loss_sf': '1.00',
        'heat_loss_method': DEFAULT_HEAT_LOSS_METHOD,
        'rtd_thrm': 'TI',
        'wind_speed': '32.00',
        'caution_label_interval': '10.00',
        'isolator_location': 'bothSides',
        'ckt_ln': '30.00',
        'loop_ln': '12.00',
    }
    payload.update(overrides)
    return payload


def make_managed_project(proj_id='PLANT_A_001', description='Plant A 001', users=None, is_active=True):
    project, _created = ManagedProject.objects.get_or_create(
        proj_id=proj_id,
        defaults={
            'description': description,
            'is_active': is_active,
        },
    )
    project.description = description
    project.is_active = is_active
    project.save(update_fields=['description', 'is_active'])
    if users:
        project.assigned_users.set(users)
    return project


def make_default_project_record(**overrides):
    managed_project, _ = ManagedProject.objects.get_or_create(
        proj_id=DEFAULT_PROJECT_ID,
        defaults={'description': 'Default project', 'is_active': True},
    )
    data = {
        'proj_id': DEFAULT_PROJECT_ID,
        'min_amb_t': 10.0,
        'max_amb_t': 40.0,
        'startup_t': 5.0,
        'area_class': 'SAFE',
        'temp_class': 'T4',
        'voltage': 110.0,
        'max_cb_size': 16,
        'restrict_cb_current': 75.0,
        'vendor': 'THR',
        'spiral_wrap_allowed': False,
        'spiral_factor': 1.5,
        'margin_on_tracer_lengths': 8.0,
        'voltage_var_factor': 2.0,
        'res_tol': 5.0,
        'termination_margin': 200.0,
        'heat_loss_sf': 1.1,
        'heat_loss_method': DEFAULT_HEAT_LOSS_METHOD,
        'rtd_thrm': 'TO',
        'wind_speed': 25.0,
        'req_local_isolator': 'required',
        'caution_label_interval': 12.0,
        'isolator_location': 'incomingOnly',
        'ckt_ln': 15.0,
        'loop_ln': 6.0,
        'allowablevdrop': 3.0,
    }
    data.update(overrides)
    return ProjectData.objects.create(**data)


def make_calculated_project_snapshot(project_id='p1'):
    make_project_record(proj_id=project_id)
    line = HeatTracingInput.objects.create(
        proj_id=project_id,
        line_id='LINE-001',
        service_type='EP',
        line_size=2.0,
        line_length=10.0,
        ins_mat_type='Mineral Wool',
        insul_thick=50.0,
        maint_temp=120.0,
        oper_temp=100.0,
        design_temp=140.0,
        status='confirmed',
    )

    aggregated_results = {
        'heat_loss': [
            {'uid': line.uid, 'heat_loss': 12.5, 'tracer_adder': 1.2},
        ],
        'selected_tracers': [
            {
                'uid': line.uid,
                'V_UID': 'V-001',
                'A_Coeff': 1.0,
                'B_Coeff': 2.0,
                'C_Coeff': 3.0,
                'Power_at_Startup_T': 11.5,
                'Ohm_per_km': 9.5,
                'Res_corrFactor_Mica': 0.5,
                'Tracer_Family': 'Self Regulating',
                'Voltage_Float': 230.0,
                'Voltage_Correction_Factor': 0.95,
                'Power_Output': 30.0,
                'Spiral_Factor': 1.1,
                'Tracer_Length': 12.0,
                'Tracer_With_Margin': 12.6,
            },
        ],
        'alternative_tracers': [
            {
                'uid': line.uid,
                'V_UID': 'V-ALT-001',
                'A_Coeff': 1.1,
                'B_Coeff': 2.1,
                'C_Coeff': 3.1,
                'Power_at_Startup_T': 12.5,
                'Ohm_per_km': 10.5,
                'Res_corrFactor_Mica': 0.6,
                'Tracer_Family': 'Self Regulating',
                'Voltage_Float': 230.0,
                'Voltage_Correction_Factor': 0.97,
                'Power_Output': 31.0,
                'Spiral_Factor': 1.2,
                'Tracer_Length': 13.0,
                'Tracer_With_Margin': 13.7,
            },
        ],
        'power_distribution': [
            {
                'uid': line.uid,
                'total_circuits': 2,
                'branches': [
                    {
                        'type': '3phJB',
                        'circuit_count': 2,
                        'connected_to': '2x 1phJB',
                        'cable_length_db_to_jb': 25.0,
                        'cable_length_jb_to_jb': 10.0,
                        'tagged_components': {
                            'MCB': 'MCB_001',
                            'JB3PH': 'JB3PH_001',
                            'Downstream': [
                                {'Tracer': 'Tracer_001'},
                                {'Tracer': 'Tracer_002'},
                            ],
                        },
                    }
                ],
            },
        ],
        'boq_per_line': {
            line.uid: {
                'TRACER': 12.6,
                'MCB': 1,
                'JB3PH': 1,
            },
        },
        'consolidated_boq': {
            'TRACER': 12.6,
            'MCB': 1,
            'JB3PH': 1,
        },
        'tracer_power_param': [
            {
                'uid': line.uid,
                'breaker_size': 10,
                'no_of_circuits': 2,
                'max_current': 2.5,
                'operating_current': 2.0,
                'operating_load': 460.0,
                'total_tracer_length': 12.6,
                'pipe_size_mm': 60.3,
            },
        ],
    }
    store_calculated_results(project_id, aggregated_results)
    return line


def make_mi_calculated_project_snapshot(project_id='p1'):
    make_project_record(proj_id=project_id)
    line = HeatTracingInput.objects.create(
        proj_id=project_id,
        line_id='LINE-MI-001',
        service_type='EP',
        line_size=2.0,
        line_length=18.0,
        ins_mat_type='Mineral Wool',
        insul_thick=50.0,
        maint_temp=260.0,
        oper_temp=280.0,
        design_temp=320.0,
        status='confirmed',
    )
    seeded_mi_result = make_mi_selection_result(line, status='selected')
    heater = seeded_mi_result.heater
    cold_lead = seeded_mi_result.cold_lead_option

    aggregated_results = {
        'heat_loss': [
            {
                'uid': line.uid,
                'heat_loss': 42.0,
                'design_heat_loss': 42.0,
                'tracer_adder': 0.0,
                'selection_status': 'rejected',
                'selection_rejection_reasons': [{
                    'code': 'SR_TEMPERATURE_LIMIT_EXCEEDED',
                    'message': 'SR temperature limit exceeded; automatic MI fallback selected.',
                }],
            },
        ],
        'selected_mi_heaters': [
            {
                'uid': line.uid,
                'heater_id': heater.id,
                'cold_lead_option_id': cold_lead.id,
                'selection_status': 'selected',
                'heater_part_number': heater.part_number,
                'cold_lead_option_code': cold_lead.option_code,
                'heated_length_m': 18.0,
                'cold_lead_length_m': cold_lead.length_m,
                'heater_resistance_ohms': 1.8,
                'cold_lead_resistance_total_ohms': 0.04,
                'power_nominal_w': 5284.8,
                'power_density_w_m': 293.6,
                'current_nominal_a': 21.82,
                'current_cold_start_a': 21.82,
                'selection_basis': {'selection_mode': 'automatic_temperature_fallback'},
            },
        ],
        'power_distribution': [
            {
                'uid': line.uid,
                'total_circuits': 1,
                'branches': [
                    {
                        'type': '1phJB',
                        'circuit_count': 1,
                        'connected_to': 'Tracer',
                        'cable_length_db_to_jb': 30.0,
                        'cable_length_jb_to_jb': None,
                        'tagged_components': {
                            'MCB': 'MCB_001',
                            'Downstream': [{'Tracer': 'Tracer_001'}],
                        },
                    },
                ],
            },
        ],
        'boq_per_line': {
            line.uid: {
                'MCB': 1,
                'JB1PH': 1,
                'MI_HEATER_SET': 1,
                'MI_HEATED_LENGTH': 18.0,
                'MI_COLD_LEAD_LENGTH': cold_lead.length_m,
            },
        },
        'consolidated_boq': {
            'MCB': 1,
            'JB1PH': 1,
            'MI_HEATER_SET': 1,
            'MI_HEATED_LENGTH': 18.0,
            'MI_COLD_LEAD_LENGTH': cold_lead.length_m,
        },
        'tracer_power_param': [
            {
                'uid': line.uid,
                'calculation_basis': 'MI_SINGLE_HEATER_MVP',
                'selected_tracer': heater.part_number,
                'breaker_size': 32,
                'no_of_circuits': 1,
                'max_current': 21.82,
                'operating_current': 21.82,
                'operating_load': 5284.8,
                'total_tracer_length': 18.0,
                'pipe_size_mm': 60.3,
            },
        ],
    }
    store_calculated_results(project_id, aggregated_results)
    return line


def make_mi_rejected_project_snapshot(project_id='p1'):
    make_project_record(proj_id=project_id, vendor='THR')
    line = HeatTracingInput.objects.create(
        proj_id=project_id,
        line_id='LINE-MI-REJECTED',
        service_type='EP',
        line_size=30.0,
        line_length=70.95,
        ins_mat_type='MO',
        insul_thick=50.0,
        maint_temp=45.0,
        oper_temp=45.0,
        design_temp=350.0,
        status='confirmed',
    )
    store_calculated_results(project_id, {
        'heat_loss': [
            {
                'uid': line.uid,
                'heat_loss': 42.0,
                'design_heat_loss': 42.0,
                'tracer_adder': 0.0,
                'selection_status': 'rejected',
                'selection_rejection_reasons': [{
                    'code': 'SR_TEMPERATURE_LIMIT_EXCEEDED',
                    'message': 'SR catalogue temperature limit exceeded.',
                }],
            },
        ],
        'selected_mi_heaters': [
            {
                'uid': line.uid,
                'mi_selection_status': 'rejected',
                'mi_selection_rejection_reasons': [{
                    'rule_set': 'MI_SELECTION_REJECTION_REASON_V1',
                    'code': 'NO_VALIDATED_MI_CATALOGUE_DATA',
                    'message': 'No validated MI catalogue rows are available for the selected vendor.',
                    'details': {'vendor': 'THR', 'catalogue_rows': 3},
                }],
                'selection_basis': {'selection_mode': 'automatic_temperature_fallback'},
            },
        ],
    })
    return line


def make_rich_sld_project_snapshot(project_id='p1', line_ids=None):
    line_ids = line_ids or ['LINE-001']
    make_project_record(proj_id=project_id)
    project_settings = make_project_settings(proj_id=project_id)
    tag_factory = ProjectTagFactory(project_id)
    selected_tracer_template = {
        'V_UID': 'V-001',
        'A_Coeff': 0.0,
        'B_Coeff': 0.0,
        'C_Coeff': 100.0,
        'Power_at_Startup_T': 10.0,
        'Ohm_per_km': 1.0,
        'Res_corrFactor_Mica': 1.0,
        'Tracer_Family': 'SR',
        'Voltage_Float': 230.0,
        'Voltage_Correction_Factor': 1.0,
        'Power_Output': 30.0,
        'Spiral_Factor': 1.0,
        'Tracer_Length': 20.0,
        'Tracer_With_Margin': 20.0,
    }
    aggregated_results = {
        'heat_loss': [],
        'selected_tracers': [],
        'alternative_tracers': [],
        'power_distribution': [],
        'boq_per_line': {},
        'consolidated_boq': {},
        'tracer_power_param': [],
    }
    created_lines = []

    for index, line_id in enumerate(line_ids, start=1):
        line = HeatTracingInput.objects.create(
            proj_id=project_id,
            line_id=line_id,
            service_type='EP',
            line_size=2.0,
            line_length=10.0 + index,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )
        created_lines.append(line)
        aggregated_results['selected_tracers'].append({
            **selected_tracer_template,
            'uid': line.uid,
        })

        power_params = compute_power_params(
            make_line(uid=str(line.uid), line_id=line.line_id),
            project_settings,
            make_asme_table(),
            selected_tracer_template,
        )
        aggregated_results['power_distribution'].append(
            compute_power_distribution(power_params, project_settings, tag_factory=tag_factory)
        )
        aggregated_results['tracer_power_param'].append(power_params)

    store_calculated_results(project_id, aggregated_results)
    return created_lines


def make_direct_sld_project_snapshot(project_id='p1', line_id='LINE-001'):
    make_project_record(proj_id=project_id)
    project_settings = make_project_settings(proj_id=project_id)
    tag_factory = ProjectTagFactory(project_id)
    line = HeatTracingInput.objects.create(
        proj_id=project_id,
        line_id=line_id,
        service_type='EP',
        line_size=2.0,
        line_length=10.0,
        ins_mat_type='Mineral Wool',
        insul_thick=50.0,
        maint_temp=120.0,
        oper_temp=100.0,
        design_temp=140.0,
        status='confirmed',
    )
    selected_tracer = {
        'V_UID': 'V-001',
        'A_Coeff': 0.0,
        'B_Coeff': 0.0,
        'C_Coeff': 10.0,
        'Power_at_Startup_T': 10.0,
        'Ohm_per_km': 1.0,
        'Res_corrFactor_Mica': 1.0,
        'Tracer_Family': 'SR',
        'Voltage_Float': 230.0,
        'Voltage_Correction_Factor': 1.0,
        'Power_Output': 10.0,
        'Spiral_Factor': 1.0,
        'Tracer_Length': 10.0,
        'Tracer_With_Margin': 10.0,
    }
    power_params = compute_power_params(
        make_line(uid=str(line.uid), line_id=line.line_id, oper_temp=100.0),
        project_settings,
        make_asme_table(),
        selected_tracer,
    )
    store_calculated_results(project_id, {
        'heat_loss': [],
        'selected_tracers': [{**selected_tracer, 'uid': line.uid}],
        'alternative_tracers': [],
        'power_distribution': [compute_power_distribution(power_params, project_settings, tag_factory=tag_factory)],
        'boq_per_line': {},
        'consolidated_boq': {},
        'tracer_power_param': [power_params],
    })
    return line


def make_mi_selection_result(line, status='available_alternative'):
    family = MICableFamily.objects.create(
        vendor='CHR',
        family_name=f'MIQ-{line.uid}',
        alloy_type='Alloy 825',
        max_voltage=600.0,
        max_sheath_temp_c=180.0,
        max_maintain_temp_c=500.0,
        max_exposure_temp_c=600.0,
        max_watt_density_w_m=80.0,
        min_circuit_length_m=1.0,
        max_circuit_length_m=250.0,
        temp_class_rating='T3',
        gas_group='IIC',
        zone_approval='Zone 1',
        source_document='Test-only vendor document',
        is_validated=True,
    )
    heater = MICableHeater.objects.create(
        family=family,
        part_number=f'MIQ-R{line.uid}',
        conductors=1,
        resistance_ohms_m=0.1,
        max_current_a=60.0,
        cold_lead_resistance_ohms_m=0.02,
        cold_lead_max_ampacity_a=60.0,
        sheath_material='Alloy 825',
        conductor_material='Nickel Chromium',
    )
    cold_lead = MIColdLeadOption.objects.create(
        heater=heater,
        option_code='CL-2M',
        length_m=2.0,
    )
    return SelectedMIHeater.objects.create(
        line=line,
        heater=heater,
        cold_lead_option=cold_lead,
        selection_status=status,
        cold_lead_option_code='CL-2M',
        cold_lead_length_m=2.0,
        heated_length_m=105.0,
        heater_resistance_ohms=10.5,
        cold_lead_resistance_total_ohms=0.04,
        power_nominal_w=5284.8,
        power_density_w_m=50.33,
        current_nominal_a=21.82,
        current_cold_start_a=21.82,
        max_sheath_temp_published_c=180.0,
        project_t_class_limit_c=200.0,
        t_class_verdict='pass',
        selection_basis={'selection_mode': 'available_alternative'},
    )


def seed_reference_data():
    ElecEHT_ThermalConductivity.objects.create(
        Ins_Mat_Type='Mineral Wool',
        K_factor_A=0.0,
        K_factor_B=0.0,
        K_factor_C=0.05,
    )
    ElecEHT_ASMEB36.objects.create(
        Nominal_Pipe_Size=2.0,
        Outside_Diameter_mm=60.3,
    )
    ElecEHT_Vendor.objects.create(
        V_UID='V-001',
        Vendor='Chromalox',
        Tracer_Family='SR',
        Voltage=230.0,
        A_Coeff=0.0,
        B_Coeff=0.0,
        C_Coeff=30.0,
        Power_at_Startup_T=10.0,
        Ohm_per_km=1.0,
        Res_corrFactor_Mica=1.0,
    )
    ElecEHT_Vendor.objects.create(
        V_UID='V-ALT-001',
        Vendor='Chromalox',
        Tracer_Family='SR',
        Voltage=230.0,
        A_Coeff=0.0,
        B_Coeff=0.0,
        C_Coeff=25.0,
        Power_at_Startup_T=10.0,
        Ohm_per_km=1.0,
        Res_corrFactor_Mica=1.0,
    )


class HeatLossCalculationTests(SimpleTestCase):
    def test_calculate_heat_loss_returns_expected_values(self):
        line = make_line()
        project_settings = make_project_settings()
        pipe_size_mm = 60.3

        result = calculate_heat_loss(line, project_settings, make_asme_table(), make_thermal_table())

        expected_heat_loss = (2 * math.pi * 0.05 * 80.0) / math.log((2 * 50.0 + pipe_size_mm) / pipe_size_mm)
        expected_tracer_adder = (
            2 * (3.5 + 0.5 * pipe_size_mm / 25.4) / 3.048
            + 3 * (2 + 0.08 * pipe_size_mm / 25.4) / 3.048
            + 0.3
        )

        self.assertEqual(result['uid'], 'L1')
        self.assertAlmostEqual(result['heat_loss'], expected_heat_loss, places=6)
        self.assertAlmostEqual(result['base_heat_loss'], expected_heat_loss, places=6)
        self.assertAlmostEqual(result['design_heat_loss'], expected_heat_loss, places=6)
        self.assertEqual(result['heat_loss_sf'], 1.0)
        self.assertAlmostEqual(result['tracer_adder'], expected_tracer_adder, places=6)

    def test_calculate_heat_loss_applies_project_safety_factor(self):
        line = make_line()
        project_settings = make_project_settings(heat_loss_sf=1.25)
        pipe_size_mm = 60.3

        result = calculate_heat_loss(line, project_settings, make_asme_table(), make_thermal_table())

        expected_base_heat_loss = (2 * math.pi * 0.05 * 80.0) / math.log((2 * 50.0 + pipe_size_mm) / pipe_size_mm)
        self.assertAlmostEqual(result['base_heat_loss'], expected_base_heat_loss, places=6)
        self.assertAlmostEqual(result['design_heat_loss'], expected_base_heat_loss * 1.25, places=6)
        self.assertAlmostEqual(result['heat_loss'], result['design_heat_loss'], places=6)
        self.assertEqual(result['heat_loss_sf'], 1.25)

    def test_calculate_heat_loss_returns_none_for_unknown_insulation(self):
        line = make_line(ins_mat_type='UNKNOWN')

        result = calculate_heat_loss(line, make_project_settings(), make_asme_table(), make_thermal_table())

        self.assertIsNone(result)


class TracerSelectionTests(SimpleTestCase):
    def _catalogue(self, *rows):
        base_row = {
            'Voltage_Float': 230.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Power_at_Startup_T': 10.0,
            'Ohm_per_km': 1.0,
            'Res_corrFactor_Mica': 1.0,
            'Tracer_Family': 'Self Regulating',
            'Zone': 'Zone1, Zone2',
            'Gas_Group': 'IIB',
            'T_Rating': 'T1,T2,T3',
            'Maint_T': 150.0,
            'Max_Op_T': 150.0,
            'Max_Exp_T_On': 232.0,
        }
        return pd.DataFrame([{**base_row, **row} for row in rows])

    def test_get_tracer_options_returns_best_and_sorted_alternatives(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0}
        best_tracer, alternatives = get_tracer_options(
            heat_loss,
            make_line(),
            make_project_settings(),
            make_tracer_vendor_data(),
        )

        self.assertEqual(best_tracer['V_UID'], 'T1')
        self.assertAlmostEqual(best_tracer['Tracer_With_Margin'], 13.20, places=2)
        self.assertEqual(best_tracer['SR_Parallel_Run_Count'], 1)
        self.assertEqual([item['V_UID'] for item in alternatives], ['T2', 'T3'])
        self.assertEqual(heat_loss['selection_status'], 'selected')
        self.assertEqual(heat_loss['selection_rejection_reasons'], [])

    def test_get_tracer_options_selects_parallel_sr_runs_before_rejecting_heat_duty(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 120.0, 'tracer_adder': 0.0}

        best_tracer, alternatives = get_tracer_options(
            heat_loss,
            make_line(line_size=4.0, line_length=10.0),
            make_project_settings(spiral_factor=1.0, sr_max_parallel_runs=4),
            self._catalogue({'V_UID': 'T70', 'C_Coeff': 70.0}),
        )

        self.assertEqual(best_tracer['V_UID'], 'T70')
        self.assertEqual(best_tracer['SR_Parallel_Run_Count'], 2)
        self.assertAlmostEqual(best_tracer['Spiral_Factor'], 120.0 / 70.0 / 2, places=6)
        self.assertAlmostEqual(best_tracer['SR_Per_Run_Tracer_Length'], 10.0, places=6)
        self.assertAlmostEqual(best_tracer['Tracer_With_Margin'], 22.0, places=6)
        self.assertEqual(best_tracer['SR_Constructability_Warning'], '')
        self.assertEqual(alternatives, [])

    def test_get_tracer_options_allows_straight_sr_overdesign_without_lower_duty_gate(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 15.94, 'tracer_adder': 0.0}

        best_tracer, alternatives = get_tracer_options(
            heat_loss,
            make_line(line_size=2.0, line_length=46.7, maint_temp=40.0, oper_temp=40.0, design_temp=120.0),
            make_project_settings(spiral_factor=1.0, spiral_wrap_allowed=False, sr_max_parallel_runs=4),
            self._catalogue(
                {'V_UID': 'UNDERPOWERED', 'C_Coeff': 15.33},
                {'V_UID': 'LOWEST_ACCEPTABLE', 'C_Coeff': 23.02},
                {'V_UID': 'HIGHER_OUTPUT', 'C_Coeff': 30.75},
            ),
        )

        self.assertEqual(best_tracer['V_UID'], 'LOWEST_ACCEPTABLE')
        self.assertEqual(best_tracer['SR_Parallel_Run_Count'], 1)
        self.assertAlmostEqual(best_tracer['Spiral_Factor'], 15.94 / 23.02, places=6)
        self.assertEqual([item['V_UID'] for item in alternatives[:2]], ['HIGHER_OUTPUT', 'UNDERPOWERED'])

    def test_get_tracer_options_flags_small_pipe_parallel_run_constructability(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 120.0, 'tracer_adder': 0.0}

        best_tracer, _alternatives = get_tracer_options(
            heat_loss,
            make_line(line_size=1.5, line_length=10.0),
            make_project_settings(spiral_factor=1.0, sr_max_parallel_runs=4),
            self._catalogue({'V_UID': 'T45', 'C_Coeff': 45.0}),
        )

        self.assertEqual(best_tracer['SR_Parallel_Run_Count'], 3)
        self.assertIn('exceed the pipe-size guided preference', best_tracer['SR_Constructability_Warning'])

    def test_get_tracer_options_rejects_spiral_wrap_candidates_when_not_allowed(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0}
        best_tracer, alternatives = get_tracer_options(
            heat_loss,
            make_line(),
            make_project_settings(spiral_wrap_allowed=False, sr_max_parallel_runs=1),
            make_tracer_vendor_data().iloc[1:].copy(),
        )

        self.assertEqual(best_tracer, {})
        self.assertEqual(alternatives, [])
        self.assertEqual(heat_loss['selection_status'], 'rejected')
        rejection = heat_loss['selection_rejection_reasons'][0]
        self.assertEqual(rejection['code'], 'NO_SPIRAL_FACTOR_MATCH')
        self.assertEqual(rejection['details']['attempted_run_counts'], [1])
        self.assertEqual(rejection['details']['best_candidate_v_uid'], 'T2')
        self.assertGreater(rejection['details']['best_per_run_duty_ratio_at_max_runs'], 1.0)

    def test_get_tracer_options_filters_catalogue_temperature_limits_and_family(self):
        vendor_data = self._catalogue(
            {'V_UID': 'GOOD'},
            {'V_UID': 'BAD_MAINT', 'Maint_T': 90.0},
            {'V_UID': 'BAD_OPERATING', 'Max_Op_T': 70.0},
            {'V_UID': 'BAD_EXPOSURE', 'Max_Exp_T_On': 110.0},
            {'V_UID': 'MI_ROW', 'Tracer_Family': 'Constant Wattage'},
        )

        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0},
            make_line(maint_temp=100.0, oper_temp=80.0, design_temp=120.0),
            make_project_settings(area_class='SAFE', temp_class='T3'),
            vendor_data,
        )

        self.assertEqual(best_tracer['V_UID'], 'GOOD')
        self.assertEqual(alternatives, [])

    def test_get_tracer_options_filters_zone_and_temperature_class_when_declared(self):
        vendor_data = self._catalogue(
            {'V_UID': 'BAD_ZONE', 'Zone': 'Zone2'},
            {'V_UID': 'BAD_T_CLASS', 'T_Rating': 'T1,T2'},
            {'V_UID': 'GOOD', 'Zone': 'Zone1, Zone2', 'T_Rating': 'T1,T2,T3'},
        )

        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0},
            make_line(),
            make_project_settings(area_class='Zone1', temp_class='T3'),
            vendor_data,
        )

        self.assertEqual(best_tracer['V_UID'], 'GOOD')
        self.assertEqual(alternatives, [])

    def test_get_tracer_options_parses_combined_zone_and_gas_group_area_class(self):
        vendor_data = self._catalogue(
            {'V_UID': 'BAD_ZONE', 'Zone': 'Zone1'},
            {'V_UID': 'GOOD', 'Zone': 'Zone1, Zone2', 'Gas_Group': 'NA'},
        )

        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0},
            make_line(),
            make_project_settings(area_class='Zone-II, IIB', temp_class='T3'),
            vendor_data,
        )

        self.assertEqual(best_tracer['V_UID'], 'GOOD')
        self.assertEqual(alternatives, [])

    def test_get_tracer_options_keeps_nec_gas_group_text_for_iec_project_area(self):
        vendor_data = self._catalogue(
            {
                'V_UID': 'CHROMALOX_STYLE',
                'Zone': 'Zone1, Zone2',
                'Gas_Group': 'Class I Div.2, Gr. A, B, C, D',
            },
        )

        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0},
            make_line(),
            make_project_settings(area_class='Zone-II, IIB', temp_class='T3'),
            vendor_data,
        )

        self.assertEqual(best_tracer['V_UID'], 'CHROMALOX_STYLE')
        self.assertEqual(alternatives, [])

    def test_get_tracer_options_handles_empty_vendor_catalogue(self):
        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0},
            make_line(),
            make_project_settings(),
            pd.DataFrame(),
        )

        self.assertEqual(best_tracer, {})
        self.assertEqual(alternatives, [])

    def test_calculate_voltage_scenarios_separates_low_nominal_and_high_voltage(self):
        scenarios = calculate_voltage_scenarios(230.0, 0.10, 230.0)

        self.assertAlmostEqual(scenarios['low_voltage'], 207.0, places=6)
        self.assertAlmostEqual(scenarios['nominal_voltage'], 230.0, places=6)
        self.assertAlmostEqual(scenarios['high_voltage'], 253.0, places=6)
        self.assertAlmostEqual(scenarios['heat_delivery_correction'], 0.81, places=6)
        self.assertAlmostEqual(scenarios['nominal_correction'], 1.0, places=6)
        self.assertAlmostEqual(scenarios['max_current_correction'], 1.21, places=6)

    def test_calculate_voltage_scenarios_rejects_non_positive_voltages(self):
        with self.assertRaises(ValueError):
            calculate_voltage_scenarios(230.0, 0.10, 0.0)

        with self.assertRaises(ValueError):
            calculate_voltage_scenarios(0.0, 0.10, 230.0)

    def test_get_tracer_options_rejects_catalogue_rows_with_invalid_voltage(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0}

        best_tracer, alternatives = get_tracer_options(
            heat_loss,
            make_line(),
            make_project_settings(),
            self._catalogue({'V_UID': 'BAD_VOLTAGE', 'Voltage_Float': 0.0}),
        )

        self.assertEqual(best_tracer, {})
        self.assertEqual(alternatives, [])
        self.assertEqual(heat_loss['selection_status'], 'rejected')
        self.assertEqual(
            heat_loss['selection_rejection_reasons'][0]['code'],
            'NO_SR_CATALOGUE_VOLTAGE_COMPATIBILITY',
        )

    def test_get_tracer_options_rejects_rows_with_invalid_power_coefficients(self):
        heat_loss = {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0}

        best_tracer, alternatives = get_tracer_options(
            heat_loss,
            make_line(),
            make_project_settings(),
            self._catalogue({'V_UID': 'BAD_COEFF', 'A_Coeff': 'not-a-number'}),
        )

        self.assertEqual(best_tracer, {})
        self.assertEqual(alternatives, [])
        self.assertEqual(heat_loss['selection_status'], 'rejected')
        rejection = heat_loss['selection_rejection_reasons'][0]
        self.assertEqual(rejection['code'], 'NO_SR_POWER_COEFFICIENT_DATA')
        self.assertEqual(rejection['details']['invalid_by_column']['A_Coeff'], 1)

    def test_get_tracer_options_sizes_heat_delivery_at_low_voltage(self):
        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 81.0, 'tracer_adder': 2.0},
            make_line(),
            make_project_settings(voltage_var_factor=10.0, margin_on_tracer_lengths=0.0),
            self._catalogue({'V_UID': 'LOW_VOLTAGE_CASE'}),
        )

        self.assertEqual(best_tracer['V_UID'], 'LOW_VOLTAGE_CASE')
        self.assertAlmostEqual(best_tracer['Power_Output'], 100.0, places=6)
        self.assertAlmostEqual(best_tracer['Power_Output_Heat_Delivery'], 81.0, places=6)
        self.assertAlmostEqual(best_tracer['Spiral_Factor'], 1.0, places=6)
        self.assertAlmostEqual(best_tracer['Voltage_Correction_Factor_Heat_Delivery'], 0.81, places=6)
        self.assertAlmostEqual(best_tracer['Voltage_Correction_Factor_Max_Current'], 1.21, places=6)
        self.assertEqual(alternatives, [])

    def test_filter_sr_catalogue_voltage_prefers_rows_at_or_above_system_voltage(self):
        filtered = filter_sr_catalogue_voltage_compatibility(
            self._catalogue(
                {'V_UID': 'LOWER_NOMINAL', 'Voltage_Float': 230.0},
                {'V_UID': 'MATCHING_NOMINAL', 'Voltage_Float': 240.0},
                {'V_UID': 'HIGHER_NOMINAL', 'Voltage_Float': 300.0},
            ),
            240.0,
        )

        self.assertEqual(filtered['V_UID'].tolist(), ['MATCHING_NOMINAL', 'HIGHER_NOMINAL'])
        self.assertEqual(
            set(filtered['Catalogue_Voltage_Compatibility']),
            {'rated_at_or_above_system_nominal'},
        )

    def test_filter_sr_catalogue_voltage_allows_nearby_nominal_class_when_needed(self):
        filtered = filter_sr_catalogue_voltage_compatibility(
            self._catalogue(
                {'V_UID': 'SAME_CLASS_230', 'Voltage_Float': 230.0},
                {'V_UID': 'TOO_LOW_200', 'Voltage_Float': 200.0},
            ),
            240.0,
        )

        self.assertEqual(filtered['V_UID'].tolist(), ['SAME_CLASS_230'])
        self.assertEqual(filtered.iloc[0]['Catalogue_Voltage_Compatibility'], 'nearby_nominal_voltage_class')

    def test_get_tracer_options_uses_nearby_230v_catalogue_for_240v_project(self):
        best_tracer, alternatives = get_tracer_options(
            {'uid': 'L1', 'heat_loss': 90.0, 'tracer_adder': 2.0},
            make_line(),
            make_project_settings(voltage=240.0),
            self._catalogue({'V_UID': 'SST_STYLE_230V', 'Voltage_Float': 230.0}),
        )

        self.assertEqual(best_tracer['V_UID'], 'SST_STYLE_230V')
        self.assertEqual(best_tracer['Catalogue_Voltage_Compatibility'], 'nearby_nominal_voltage_class')
        self.assertEqual(alternatives, [])

    def test_filter_sr_catalogue_suitability_handles_iec_gas_group_hierarchy(self):
        filtered = filter_sr_catalogue_suitability(
            self._catalogue(
                {'V_UID': 'BAD_IIA', 'Gas_Group': 'IIA'},
                {'V_UID': 'GOOD_IIC', 'Gas_Group': 'IIC'},
            ),
            make_line(),
            make_project_settings(area_class='SAFE', gas_group='IIB'),
        )

        self.assertEqual(filtered['V_UID'].tolist(), ['GOOD_IIC'])


class PowerDistributionCalculationTests(SimpleTestCase):
    def test_compute_mi_power_params_sizes_single_heater_set(self):
        selected_mi_heater = {
            'heater_part_number': 'MIQ-R001',
            'heated_length_m': 105.0,
            'cold_lead_length_m': 2.0,
            'current_nominal_a': 4.5,
            'current_cold_start_a': 5.0,
            'power_nominal_w': 1035.0,
        }

        power_params = compute_mi_power_params(
            make_line(),
            make_project_settings(),
            make_asme_table(),
            selected_mi_heater,
        )

        self.assertEqual(power_params['calculation_basis'], 'MI_SINGLE_HEATER_MVP')
        self.assertEqual(power_params['selected_tracer'], 'MIQ-R001')
        self.assertEqual(power_params['no_of_circuits'], 1)
        self.assertEqual(power_params['breaker_size'], 10)
        self.assertEqual(power_params['operating_current'], 4.5)
        self.assertEqual(power_params['max_current'], 5.0)
        self.assertEqual(power_params['operating_load'], 1035.0)
        self.assertEqual(power_params['total_tracer_length'], 105.0)
        self.assertEqual(
            power_params['termination_margin_basis']['semantics'],
            'factory_terminated_mi_heater_set_no_sr_field_termination_allowance',
        )

    def test_compute_mi_power_params_treats_multi_set_mi_as_parallel_circuits(self):
        selected_mi_heater = {
            'heater_part_number': 'MIQ-R002',
            'heater_set_count': 3,
            'heated_length_m': 90.0,
            'cold_lead_length_m': 2.0,
            'current_nominal_a': 15.0,
            'current_cold_start_a': 18.0,
            'power_nominal_w': 10800.0,
        }

        power_params = compute_mi_power_params(
            make_line(),
            make_project_settings(max_cb_size=25.0, restrict_cb_current=80.0),
            make_asme_table(),
            selected_mi_heater,
        )
        distribution = compute_power_distribution(
            power_params,
            make_project_settings(max_cb_size=25.0, restrict_cb_current=80.0),
        )

        self.assertEqual(power_params['calculation_basis'], 'MI_MULTI_HEATER_SET_MVP')
        self.assertEqual(power_params['heater_set_count'], 3)
        self.assertEqual(power_params['no_of_circuits'], 3)
        self.assertEqual(power_params['breaker_size'], 25)
        self.assertEqual(power_params['operating_current'], 15.0)
        self.assertEqual(power_params['max_current'], 18.0)
        self.assertEqual(power_params['line_operating_current'], 45.0)
        self.assertEqual(power_params['line_max_current'], 54.0)
        self.assertEqual(power_params['total_tracer_length'], 270.0)
        self.assertEqual(power_params['total_heated_tracer_length'], 270.0)
        self.assertEqual(
            power_params['breaker_sizing']['rule_set'],
            'MI_MULTI_HEATER_SET_BREAKER_SIZING_MVP_V1',
        )
        self.assertEqual(distribution['total_circuits'], 3)
        self.assertEqual(len(distribution['branches']), 3)
        self.assertTrue(all(branch['type'] == '1phJB' for branch in distribution['branches']))
        self.assertTrue(all(branch['circuit_count'] == 1 for branch in distribution['branches']))
        self.assertEqual(
            [branch['tagged_components']['MCB'] for branch in distribution['branches']],
            ['MCB_001', 'MCB_002', 'MCB_003'],
        )
        self.assertEqual(
            [
                branch['tagged_components']['component_details']['MCB']['metadata']['mi_heater_set_index']
                for branch in distribution['branches']
            ],
            [1, 2, 3],
        )
        self.assertTrue(all(
            branch['tagged_components']['component_details']['MCB']['metadata']['mi_independent_protection']
            for branch in distribution['branches']
        ))

        boq = compute_bill_of_quantities(
            distribution,
            make_project_settings(max_cb_size=25.0, restrict_cb_current=80.0),
            tracer_qty=0,
            line_length=90.0,
            pipe_size_mm=60.3,
            is_process_temp_controlled=True,
        )
        self.assertEqual(boq['MCB'], 3)
        self.assertEqual(boq['JB1PH'], 3)
        self.assertEqual(boq['JB3PH'], 0)
        self.assertEqual(boq['ENDTRM'], 3)

    def test_compute_power_params_and_distribution_for_two_circuits(self):
        selected_tracer = {
            'Tracer_With_Margin': 20.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Voltage_Correction_Factor': 1.0,
        }

        power_params = compute_power_params(
            make_line(),
            make_project_settings(),
            make_asme_table(),
            selected_tracer,
        )
        distribution = compute_power_distribution(power_params, make_project_settings())
        branch = distribution['branches'][0]
        expected_line_current = 2000.0 / 230.0
        expected_per_circuit_current = expected_line_current / 2

        self.assertEqual(power_params['no_of_circuits'], 2)
        self.assertEqual(power_params['breaker_size'], 6)
        self.assertAlmostEqual(power_params['line_max_current'], expected_line_current, places=6)
        self.assertAlmostEqual(power_params['max_current'], expected_per_circuit_current, places=6)
        self.assertAlmostEqual(power_params['operating_current'], expected_per_circuit_current, places=6)
        self.assertAlmostEqual(power_params['operating_load'], 2000.0, places=6)
        self.assertAlmostEqual(power_params['heated_tracer_length'], 20.0, places=6)
        self.assertAlmostEqual(power_params['termination_margin_length'], 0.5, places=6)
        self.assertAlmostEqual(power_params['total_tracer_length'], 20.5, places=6)
        self.assertEqual(
            power_params['termination_margin_basis']['semantics'],
            'installation_allowance_excluded_from_electrical_load',
        )
        self.assertAlmostEqual(power_params['pipe_size_mm'], 60.3, places=6)

        self.assertEqual(distribution['total_circuits'], 2)
        self.assertEqual(branch['type'], '3phJB')
        self.assertEqual(branch['connected_to'], '2x 1phJB')
        self.assertEqual(branch['tagged_components']['component_details']['MCB']['metadata']['breaker_size'], 6)
        self.assertAlmostEqual(
            branch['tagged_components']['component_details']['MCB']['metadata']['max_current'],
            expected_per_circuit_current,
            places=6,
        )
        self.assertEqual(branch['tagged_components']['Isolator3PH'], 'ISOL_3PH_001')
        self.assertEqual(branch['tagged_components']['component_details']['MCB']['component_type'], 'MCB')
        self.assertEqual(branch['tagged_components']['component_details']['MCB']['display_tag'], 'MCB_001')
        self.assertEqual(branch['tagged_components']['Downstream'][0]['Isolator1PH'], 'ISOL_1PH_001')
        self.assertEqual(len(branch['tagged_components']['Downstream']), 2)
        self.assertEqual(len(branch['tagged_components']['connections']), 13)

    def test_compute_power_distribution_uses_project_wide_tags_across_lines(self):
        selected_tracer = {
            'Tracer_With_Margin': 20.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Voltage_Correction_Factor': 1.0,
        }
        project_settings = make_project_settings(proj_id='p1')
        tag_factory = ProjectTagFactory('p1')

        first_distribution = compute_power_distribution(
            compute_power_params(
                make_line(uid='L1', line_id='LINE-001'),
                project_settings,
                make_asme_table(),
                selected_tracer,
            ),
            project_settings,
            tag_factory=tag_factory,
        )
        second_distribution = compute_power_distribution(
            compute_power_params(
                make_line(uid='L2', line_id='LINE-002'),
                project_settings,
                make_asme_table(),
                selected_tracer,
            ),
            project_settings,
            tag_factory=tag_factory,
        )

        first_branch = first_distribution['branches'][0]
        second_branch = second_distribution['branches'][0]

        self.assertEqual(first_branch['tagged_components']['MCB'], 'MCB_001')
        self.assertEqual(second_branch['tagged_components']['MCB'], 'MCB_002')
        self.assertEqual(
            second_branch['tagged_components']['component_details']['MCB']['line_id'],
            'LINE-002',
        )
        self.assertNotEqual(
            first_branch['tagged_components']['component_details']['MCB']['component_uid'],
            second_branch['tagged_components']['component_details']['MCB']['component_uid'],
        )

    def test_compute_power_distribution_for_single_circuit_avoids_fake_3ph_components(self):
        selected_tracer = {
            'Tracer_With_Margin': 10.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 40.0,
            'Voltage_Correction_Factor': 1.0,
        }
        project_settings = make_project_settings(max_cb_size=40.0, proj_id='p1')

        power_params = compute_power_params(
            make_line(line_id='LINE-001'),
            project_settings,
            make_asme_table(),
            selected_tracer,
        )
        distribution = compute_power_distribution(power_params, project_settings)
        branch = distribution['branches'][0]

        self.assertEqual(branch['type'], '1phJB')
        self.assertIsNone(branch['tagged_components']['JB3PH'])
        self.assertIsNone(branch['tagged_components']['Cable4C'])
        self.assertIsNone(branch['tagged_components']['Isolator3PH'])
        self.assertEqual(branch['tagged_components']['Downstream'][0]['JB1PH'], 'JB1PH_001')

    def test_compute_power_params_uses_high_voltage_for_max_current(self):
        selected_tracer = {
            'Tracer_With_Margin': 20.0,
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Voltage_Correction_Factor': 1.0,
            'Voltage_Correction_Factor_Nominal': 1.0,
            'Voltage_Correction_Factor_Max_Current': 1.21,
            'Max_Current_Voltage': 253.0,
        }

        power_params = compute_power_params(
            make_line(),
            make_project_settings(voltage_var_factor=10.0),
            make_asme_table(),
            selected_tracer,
        )

        self.assertEqual(power_params['no_of_circuits'], 2)
        self.assertAlmostEqual(power_params['line_operating_current'], 2000.0 / 230.0, places=6)
        self.assertAlmostEqual(power_params['line_max_current'], 2420.0 / 253.0, places=6)
        self.assertAlmostEqual(power_params['operating_current'], (2000.0 / 230.0) / 2, places=6)
        self.assertAlmostEqual(power_params['max_current'], (2420.0 / 253.0) / 2, places=6)
        self.assertEqual(power_params['breaker_size'], 6)
        self.assertAlmostEqual(power_params['voltage_scenarios']['max_current_voltage'], 253.0, places=6)

    def test_compute_power_params_groups_parallel_sr_runs_on_shared_mcb(self):
        selected_tracer = {
            'V_UID': 'SR-70',
            'Tracer_Family': 'Self Regulating',
            'Tracer_With_Margin': 22.0,
            'SR_Per_Run_Tracer_Length': 10.0,
            'SR_Parallel_Run_Count': 2,
            'SR_Parallel_Run_Basis': 'SR_PARALLEL_STRAIGHT_RUN_SELECTION_MVP_V1',
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 70.0,
            'Voltage_Correction_Factor': 1.0,
            'Voltage_Correction_Factor_Nominal': 1.0,
            'Voltage_Correction_Factor_Max_Current': 1.0,
        }

        power_params = compute_power_params(
            make_line(line_length=10.0),
            make_project_settings(margin_on_tracer_lengths=10.0),
            make_asme_table(),
            selected_tracer,
        )
        distribution = compute_power_distribution(power_params, make_project_settings())

        self.assertEqual(power_params['no_of_circuits'], 2)
        self.assertFalse(power_params['sr_independent_parallel_runs'])
        self.assertAlmostEqual(power_params['line_max_current'], 1540.0 / 230.0, places=6)
        self.assertAlmostEqual(power_params['max_current'], 770.0 / 230.0, places=6)
        self.assertEqual(distribution['total_circuits'], 2)
        self.assertEqual(len(distribution['branches']), 1)
        branch = distribution['branches'][0]
        self.assertEqual(branch['type'], '3phJB')
        self.assertEqual(branch['circuit_count'], 2)
        self.assertEqual(branch['tagged_components']['heating_cable_type'], 'SR')
        self.assertEqual(branch['tagged_components']['sr_parallel_run_count'], 2)
        self.assertEqual(branch['tagged_components']['sr_parallel_run_basis'], 'SR_PARALLEL_STRAIGHT_RUN_SELECTION_MVP_V1')
        self.assertTrue(branch['tagged_components']['sr_shared_mcb'])
        self.assertNotIn('sr_parallel_run_index', branch['tagged_components'])
        self.assertEqual(
            branch['tagged_components']['component_details']['MCB']['metadata']['sr_parallel_run_count'],
            2,
        )


class BoqCalculationTests(SimpleTestCase):
    def test_compute_bill_of_quantities_counts_key_components(self):
        power_distribution = {
            'uid': 'L1',
            'total_circuits': 2,
            'branches': [
                {
                    'type': '3phJB',
                    'connected_to': '2x 1phJB',
                    'circuit_count': 2,
                    'cable_length_db_to_jb': 30.0,
                    'cable_length_jb_to_jb': 12.0,
                },
            ],
        }

        boq = compute_bill_of_quantities(
            power_distribution=power_distribution,
            project_settings=make_project_settings(),
            tracer_qty=20.5,
            line_length=10.0,
            pipe_size_mm=60.3,
            is_process_temp_controlled=True,
        )

        self.assertEqual(boq['MCB'], 1)
        self.assertEqual(boq['JB3PH'], 1)
        self.assertEqual(boq['JB1PH'], 2)
        self.assertEqual(boq['CCMCB-3PHJB'], 30.0)
        self.assertEqual(boq['CC3PHJB-1PHJB'], 24.0)
        self.assertEqual(boq['ISOLATOR_3PH'], 1)
        self.assertEqual(boq['ISOLATOR_1PH'], 2)
        self.assertEqual(boq['THERMOSTAT'], 2)
        self.assertEqual(boq['ENDTRM'], 2)

    def test_compute_bill_of_quantities_aggregates_all_branches_in_active_payload(self):
        power_distribution = {
            'uid': 'L2',
            'total_circuits': 4,
            'branches': [
                {
                    'type': '3phJB',
                    'connected_to': '3x 1phJB',
                    'circuit_count': 3,
                    'cable_length_db_to_jb': 30.0,
                    'cable_length_jb_to_jb': 12.0,
                },
                {
                    'type': '1phJB',
                    'connected_to': 'Tracer',
                    'circuit_count': 1,
                    'cable_length_db_to_jb': 30.0,
                    'cable_length_jb_to_jb': None,
                },
            ],
        }

        boq = compute_bill_of_quantities(
            power_distribution=power_distribution,
            project_settings=make_project_settings(),
            tracer_qty=41.0,
            line_length=20.0,
            pipe_size_mm=60.3,
            is_process_temp_controlled=True,
        )

        self.assertEqual(boq['MCB'], 2)
        self.assertEqual(boq['JB3PH'], 1)
        self.assertEqual(boq['JB1PH'], 4)
        self.assertEqual(boq['CCMCB-3PHJB'], 30.0)
        self.assertEqual(boq['CC3PHJB-1PHJB'], 66.0)
        self.assertEqual(boq['ISOLATOR_3PH'], 1)
        self.assertEqual(boq['ISOLATOR_1PH'], 4)
        self.assertEqual(boq['THERMOSTAT'], 4)
        self.assertEqual(boq['ENDTRM'], 4)


class OrchestrationTests(SimpleTestCase):
    def test_orchestrate_calculations_builds_expected_aggregates(self):
        line = make_line(valve_qty=0, support_qty=0, flange_qty=0)
        vendor_data = make_tracer_vendor_data().assign(
            C_Coeff=[30.0, 25.0, 20.0]
        )

        result = orchestrate_calculations(
            process_lines=[line],
            vendor_data=vendor_data,
            project_settings=make_project_settings(),
            asme_b36_table=make_asme_table(),
            thermal_cond_data=make_thermal_table(),
        )

        self.assertEqual(len(result['heat_loss']), 1)
        self.assertEqual(result['selected_tracers'][0]['uid'], 'L1')
        self.assertEqual(len(result['alternative_tracers']), 2)
        self.assertEqual({item['uid'] for item in result['alternative_tracers']}, {'L1'})
        self.assertEqual(len(result['power_distribution']), 1)
        self.assertEqual(len(result['tracer_power_param']), 1)
        self.assertIn('TRACER', result['boq_per_line']['L1'])
        self.assertIn('MCB', result['consolidated_boq'])

    def test_orchestrate_calculations_keeps_display_tags_stable_for_same_line_set(self):
        vendor_data = make_tracer_vendor_data().assign(C_Coeff=[30.0, 25.0, 20.0])
        project_settings = make_project_settings(proj_id='p1')
        lines = [
            make_line(uid='L2', line_id='LINE-002', xlid=2, valve_qty=0, support_qty=0, flange_qty=0),
            make_line(uid='L1', line_id='LINE-001', xlid=1, valve_qty=0, support_qty=0, flange_qty=0),
        ]

        first_result = orchestrate_calculations(
            process_lines=lines,
            vendor_data=vendor_data,
            project_settings=project_settings,
            asme_b36_table=make_asme_table(),
            thermal_cond_data=make_thermal_table(),
        )
        second_result = orchestrate_calculations(
            process_lines=list(reversed(lines)),
            vendor_data=vendor_data,
            project_settings=project_settings,
            asme_b36_table=make_asme_table(),
            thermal_cond_data=make_thermal_table(),
        )

        def key_tags(result):
            return [
                (
                    distribution['uid'],
                    distribution['branches'][0]['tagged_components']['MCB'],
                    distribution['branches'][0]['tagged_components']['Downstream'][0]['Tracer'],
                )
                for distribution in result['power_distribution']
            ]

        self.assertEqual(key_tags(first_result), key_tags(second_result))
        self.assertEqual(key_tags(first_result), [('L1', 'MCB_001', 'Tracer_001'), ('L2', 'MCB_002', 'Tracer_002')])


class ProcessLineFetchTests(TestCase):
    def create_line(self, **overrides):
        defaults = {
            'proj_id': 'p1',
            'line_id': 'LINE-001',
            'service_type': 'EP',
            'line_size': 2.0,
            'line_length': 10.0,
            'ins_mat_type': 'Mineral Wool',
            'insul_thick': 50.0,
            'maint_temp': 120.0,
            'oper_temp': 100.0,
            'design_temp': 140.0,
            'status': 'confirmed',
        }
        defaults.update(overrides)
        return HeatTracingInput.objects.create(**defaults)

    def test_fetch_process_lines_returns_only_confirmed_rows(self):
        confirmed_line = self.create_line(line_id='LINE-CONFIRMED', status='confirmed')
        self.create_line(line_id='LINE-PENDING', status='pending')

        process_lines = fetch_process_lines('p1')

        self.assertEqual(list(process_lines['uid']), [confirmed_line.uid])
        self.assertEqual(list(process_lines['status']), ['confirmed'])

    def test_run_project_calculations_rejects_pending_only_project(self):
        make_project_record()
        self.create_line(status='pending')

        with self.assertRaisesMessage(ValidationError, 'No confirmed input data found for this project.'):
            run_project_calculations('p1')


class StoreCalculatedResultsTests(TestCase):
    def test_store_calculated_results_handles_current_orchestration_payload(self):
        make_project_record()
        line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-001',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )

        aggregated_results = {
            'heat_loss': [
                {
                    'uid': line.uid,
                    'heat_loss': 13.75,
                    'base_heat_loss': 12.5,
                    'design_heat_loss': 13.75,
                    'heat_loss_sf': 1.1,
                    'tracer_adder': 1.2,
                },
            ],
            'selected_tracers': [
                {
                    'uid': line.uid,
                    'V_UID': 'V-001',
                    'A_Coeff': 1.0,
                    'B_Coeff': 2.0,
                    'C_Coeff': 3.0,
                    'Power_at_Startup_T': 11.5,
                    'Ohm_per_km': 9.5,
                    'Res_corrFactor_Mica': 0.5,
                    'Tracer_Family': 'Self Regulating',
                    'Voltage_Float': 230.0,
                    'Voltage_Correction_Factor': 0.95,
                    'Power_Output': 30.0,
                    'Spiral_Factor': 1.1,
                    'Tracer_Length': 12.0,
                    'Tracer_With_Margin': 12.6,
                },
            ],
            'alternative_tracers': [
                {
                    'uid': line.uid,
                    'V_UID': 'V-ALT-001',
                    'A_Coeff': 1.1,
                    'B_Coeff': 2.1,
                    'C_Coeff': 3.1,
                    'Power_at_Startup_T': 12.5,
                    'Ohm_per_km': 10.5,
                    'Res_corrFactor_Mica': 0.6,
                    'Tracer_Family': 'Self Regulating',
                    'Voltage_Float': 230.0,
                    'Voltage_Correction_Factor': 0.97,
                    'Power_Output': 31.0,
                    'Spiral_Factor': 1.2,
                    'Tracer_Length': 13.0,
                    'Tracer_With_Margin': 13.7,
                },
            ],
            'power_distribution': [
                {
                    'uid': line.uid,
                    'total_circuits': 1,
                    'branches': [
                        {
                            'type': '1phJB',
                            'circuit_count': 1,
                            'connected_to': 'Tracer',
                            'cable_length_db_to_jb': 25.0,
                            'cable_length_jb_to_jb': None,
                            'tagged_components': {
                                'MCB': 'MCB_001',
                                'Downstream': [{'Tracer': 'Tracer_001'}],
                            },
                        }
                    ],
                },
            ],
            'boq_per_line': {
                line.uid: {'TRACER': 12.6, 'MCB': 1},
            },
            'consolidated_boq': {
                'TRACER': 12.6,
                'MCB': 1,
            },
            'tracer_power_param': [
                {
                    'uid': line.uid,
                    'max_current': 2.5,
                    'operating_current': 2.0,
                    'operating_load': 460.0,
                    'total_tracer_length': 12.6,
                    'pipe_size_mm': 60.3,
                },
            ],
        }

        self.assertTrue(store_calculated_results('p1', aggregated_results))

        heat_loss_result = HeatLoss.objects.get(line=line)
        self.assertEqual(heat_loss_result.heat_loss, 13.75)
        self.assertEqual(heat_loss_result.base_heat_loss, 12.5)
        self.assertEqual(heat_loss_result.design_heat_loss, 13.75)
        self.assertEqual(heat_loss_result.heat_loss_sf, 1.1)
        self.assertEqual(SelectedTracer.objects.get(line=line).power_output, 30.0)
        alternate_tracer = AlternateTracer.objects.get(line=line, option_rank=1)
        self.assertEqual(alternate_tracer.tracer_with_margin, 13.7)
        self.assertEqual(PowerDistribution.objects.get(line=line).total_circuits, 1)

        branch = PowerDistributionBranch.objects.get(distribution__line=line, branch_index=1)
        self.assertEqual(branch.branch_type, '1phJB')
        self.assertEqual(branch.tagged_components['MCB'], 'MCB_001')

        process_line_calc = ProcessLineCalculation.objects.get(line=line)
        self.assertEqual(process_line_calc.heat_loss, 13.75)
        self.assertEqual(process_line_calc.selected_tracer, 'V-001')
        self.assertEqual(process_line_calc.breaker_size, 0)
        self.assertEqual(process_line_calc.total_circuits, 0)
        self.assertEqual(process_line_calc.total_power_consumption, 460.0)
        self.assertEqual(process_line_calc.total_tracer_length, 12.6)
        self.assertEqual(process_line_calc.pipe_size_mm, 60.3)

        self.assertEqual(
            BOQ.objects.filter(project_id='p1', scope='line', line=line).count(),
            2,
        )
        self.assertEqual(
            BOQ.objects.filter(project_id='p1', scope='consolidated', line__isnull=True).count(),
            2,
        )
        self.assertEqual(
            BOQ.objects.get(project_id='p1', scope='line', line=line, item_code='TRACER').unit,
            'm',
        )

    def test_store_calculated_results_rolls_back_if_a_late_write_fails(self):
        make_project_record()
        line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-001',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )

        store_calculated_results('p1', {
            'heat_loss': [
                {'uid': line.uid, 'heat_loss': 12.5, 'tracer_adder': 1.2},
            ],
            'selected_tracers': [
                {
                    'uid': line.uid,
                    'V_UID': 'V-001',
                    'A_Coeff': 1.0,
                    'B_Coeff': 2.0,
                    'C_Coeff': 3.0,
                    'Power_at_Startup_T': 11.5,
                    'Ohm_per_km': 9.5,
                    'Res_corrFactor_Mica': 0.5,
                    'Tracer_Family': 'Self Regulating',
                    'Voltage_Float': 230.0,
                    'Voltage_Correction_Factor': 0.95,
                    'Power_Output': 30.0,
                    'Spiral_Factor': 1.1,
                    'Tracer_Length': 12.0,
                    'Tracer_With_Margin': 12.6,
                },
            ],
            'alternative_tracers': [],
            'power_distribution': [],
            'boq_per_line': {},
            'consolidated_boq': {},
            'tracer_power_param': [],
        })

        with patch('eht.data_service.ProcessLineCalculation.objects.bulk_create', side_effect=RuntimeError('boom')):
            with self.assertRaises(RuntimeError):
                store_calculated_results('p1', {
                    'heat_loss': [
                        {'uid': line.uid, 'heat_loss': 99.9, 'tracer_adder': 9.9},
                    ],
                    'selected_tracers': [
                        {
                            'uid': line.uid,
                            'V_UID': 'V-002',
                            'A_Coeff': 5.0,
                            'B_Coeff': 6.0,
                            'C_Coeff': 7.0,
                            'Power_at_Startup_T': 22.5,
                            'Ohm_per_km': 19.5,
                            'Res_corrFactor_Mica': 1.5,
                            'Tracer_Family': 'Updated',
                            'Voltage_Float': 230.0,
                            'Voltage_Correction_Factor': 0.99,
                            'Power_Output': 66.0,
                            'Spiral_Factor': 1.4,
                            'Tracer_Length': 14.0,
                            'Tracer_With_Margin': 14.7,
                        },
                    ],
                    'alternative_tracers': [],
                    'power_distribution': [],
                    'boq_per_line': {},
                    'consolidated_boq': {},
                    'tracer_power_param': [
                        {
                            'uid': line.uid,
                            'max_current': 4.5,
                            'operating_current': 4.0,
                            'operating_load': 920.0,
                            'total_tracer_length': 14.7,
                            'pipe_size_mm': 60.3,
                        },
                    ],
                })

        self.assertEqual(HeatLoss.objects.get(line=line).heat_loss, 12.5)
        self.assertEqual(SelectedTracer.objects.get(line=line).v_uid, 'V-001')

    def test_same_tracer_vendor_uid_can_be_stored_for_multiple_lines(self):
        make_project_record()
        line_one = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-001',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )
        line_two = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-002',
            service_type='EP',
            line_size=3.0,
            line_length=12.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )

        aggregated_results = {
            'heat_loss': [],
            'selected_tracers': [
                {
                    'uid': line_one.uid,
                    'V_UID': 'V-001',
                    'A_Coeff': 1.0,
                    'B_Coeff': 2.0,
                    'C_Coeff': 3.0,
                    'Power_at_Startup_T': 11.5,
                    'Ohm_per_km': 9.5,
                    'Res_corrFactor_Mica': 0.5,
                    'Tracer_Family': 'Self Regulating',
                    'Voltage_Float': 230.0,
                    'Voltage_Correction_Factor': 0.95,
                    'Power_Output': 30.0,
                    'Spiral_Factor': 1.1,
                    'Tracer_Length': 12.0,
                    'Tracer_With_Margin': 12.6,
                },
                {
                    'uid': line_two.uid,
                    'V_UID': 'V-001',
                    'A_Coeff': 1.0,
                    'B_Coeff': 2.0,
                    'C_Coeff': 3.0,
                    'Power_at_Startup_T': 11.5,
                    'Ohm_per_km': 9.5,
                    'Res_corrFactor_Mica': 0.5,
                    'Tracer_Family': 'Self Regulating',
                    'Voltage_Float': 230.0,
                    'Voltage_Correction_Factor': 0.95,
                    'Power_Output': 30.0,
                    'Spiral_Factor': 1.3,
                    'Tracer_Length': 14.0,
                    'Tracer_With_Margin': 14.7,
                },
            ],
            'alternative_tracers': [],
            'power_distribution': [],
            'boq_per_line': {},
            'consolidated_boq': {},
            'tracer_power_param': [],
        }

        store_calculated_results('p1', aggregated_results)

        self.assertEqual(SelectedTracer.objects.filter(v_uid='V-001').count(), 2)
        self.assertEqual(SelectedTracer.objects.get(line=line_one).tracer_with_margin, 12.6)
        self.assertEqual(SelectedTracer.objects.get(line=line_two).tracer_with_margin, 14.7)


class SldPayloadTests(TestCase):
    def test_build_project_sld_payload_reads_rich_branch_json_for_multiple_lines(self):
        make_project_record(proj_id='p1')
        line_one = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-001',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )
        line_two = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-002',
            service_type='EP',
            line_size=2.0,
            line_length=12.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )

        selected_tracer_template = {
            'V_UID': 'V-001',
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Power_at_Startup_T': 10.0,
            'Ohm_per_km': 1.0,
            'Res_corrFactor_Mica': 1.0,
            'Tracer_Family': 'SR',
            'Voltage_Float': 230.0,
            'Voltage_Correction_Factor': 1.0,
            'Power_Output': 30.0,
            'Spiral_Factor': 1.0,
            'Tracer_Length': 20.0,
            'Tracer_With_Margin': 20.0,
        }
        project_settings = make_project_settings(proj_id='p1')
        tag_factory = ProjectTagFactory('p1')
        aggregated_results = {
            'heat_loss': [],
            'selected_tracers': [],
            'alternative_tracers': [],
            'power_distribution': [],
            'boq_per_line': {},
            'consolidated_boq': {},
            'tracer_power_param': [],
        }

        for line in [line_one, line_two]:
            selected_tracer = {**selected_tracer_template, 'uid': line.uid}
            power_params = compute_power_params(
                make_line(uid=str(line.uid), line_id=line.line_id),
                project_settings,
                make_asme_table(),
                selected_tracer_template,
            )
            aggregated_results['selected_tracers'].append(selected_tracer)
            aggregated_results['power_distribution'].append(
                compute_power_distribution(power_params, project_settings, tag_factory=tag_factory)
            )
            aggregated_results['tracer_power_param'].append(power_params)

        self.assertTrue(store_calculated_results('p1', aggregated_results))

        payload = build_project_sld_payload('p1')

        self.assertEqual(payload['project_id'], 'p1')
        self.assertEqual(payload['schema_version'], SLD_GRAPH_SCHEMA_VERSION)
        self.assertEqual(payload['meta']['branch_count'], 2)
        self.assertEqual(payload['meta']['node_count'], 28)
        self.assertEqual(payload['meta']['edge_count'], 26)
        self.assertEqual(
            {group['line_id'] for group in payload['line_groups']},
            {'LINE-001', 'LINE-002'},
        )
        self.assertEqual(len({group['line_uid'] for group in payload['line_groups']}), 2)

        component_ids = [node['component_id'] for node in payload['nodes']]
        component_uids = [node['component_uid'] for node in payload['nodes']]
        display_tags = [node['display_tag'] for node in payload['nodes']]

        self.assertEqual(len(component_ids), len(set(component_ids)))
        self.assertEqual(len(component_uids), len(set(component_uids)))
        self.assertEqual(len(display_tags), len(set(display_tags)))
        self.assertTrue(all(len(component_uid) == 32 for component_uid in component_uids))
        self.assertIn('MCB_001', display_tags)
        self.assertIn('MCB_002', display_tags)
        self.assertTrue(any('line:LINE-001' in component_id for component_id in component_ids))
        self.assertTrue(any('line:LINE-002' in component_id for component_id in component_ids))
        self.assertTrue(all('line_uid:' in component_id for component_id in component_ids))

    def test_build_project_sld_payload_represents_multi_set_mi_as_independent_breakers(self):
        make_project_record(proj_id='p1')
        line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-MI-MULTI',
            service_type='EP',
            line_size=30.0,
            line_length=70.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=45.0,
            oper_temp=45.0,
            design_temp=350.0,
            status='confirmed',
        )
        mi_result = make_mi_selection_result(line, status='selected')
        selected_mi_heater = {
            'uid': line.uid,
            'heater_id': mi_result.heater_id,
            'cold_lead_option_id': mi_result.cold_lead_option_id,
            'selection_status': 'selected',
            'heater_part_number': mi_result.heater.part_number,
            'cold_lead_option_code': mi_result.cold_lead_option_code,
            'heater_set_count': 3,
            'heated_length_m': 90.0,
            'cold_lead_length_m': 2.0,
            'power_nominal_w': 10800.0,
            'power_density_w_m': 120.0,
            'current_nominal_a': 15.0,
            'current_cold_start_a': 18.0,
            'selection_basis': {
                'selection_mode': 'automatic_temperature_fallback',
                'heater_set_count': 3,
                'mvp_multi_set_selection': True,
            },
        }
        project_settings = make_project_settings(
            proj_id='p1',
            max_cb_size=25.0,
            restrict_cb_current=80.0,
            isolator_location='bothSides',
        )
        power_params = compute_mi_power_params(
            make_line(uid=str(line.uid), line_id=line.line_id, line_length=70.0),
            project_settings,
            make_asme_table(),
            selected_mi_heater,
        )
        distribution = compute_power_distribution(
            power_params,
            project_settings,
            tag_factory=ProjectTagFactory('p1'),
        )
        boq = compute_bill_of_quantities(
            distribution,
            project_settings,
            tracer_qty=0,
            line_length=line.line_length,
            pipe_size_mm=60.3,
            is_process_temp_controlled=True,
        )
        boq.update({
            'MI_HEATER_SET': 3,
            'MI_HEATED_LENGTH': 270.0,
            'MI_COLD_LEAD_LENGTH': 6.0,
        })

        self.assertTrue(store_calculated_results('p1', {
            'heat_loss': [{
                'uid': line.uid,
                'heat_loss': 102.0,
                'design_heat_loss': 102.0,
                'tracer_adder': 20.0,
                'selection_status': 'rejected',
            }],
            'selected_mi_heaters': [selected_mi_heater],
            'power_distribution': [distribution],
            'boq_per_line': {line.uid: boq},
            'consolidated_boq': boq,
            'tracer_power_param': [power_params],
        }))

        payload = build_project_sld_payload('p1')
        mcb_nodes = [node for node in payload['nodes'] if node['component_type'] == 'MCB']
        tracer_nodes = [node for node in payload['nodes'] if node['component_type'] == 'Tracer']
        jb3ph_nodes = [node for node in payload['nodes'] if node['component_type'] == 'JB3PH']
        mi_group_ids = {node['metadata'].get('mi_group_id') for node in mcb_nodes}

        self.assertEqual(payload['meta']['branch_count'], 3)
        self.assertEqual(payload['line_groups'][0]['branch_indices'], [1, 2, 3])
        self.assertEqual(len(mcb_nodes), 3)
        self.assertEqual(len(tracer_nodes), 3)
        self.assertEqual(jb3ph_nodes, [])
        self.assertEqual({node['display_name'] for node in tracer_nodes}, {mi_result.heater.part_number})
        self.assertEqual(
            [node['metadata']['mi_heater_set_index'] for node in mcb_nodes],
            [1, 2, 3],
        )
        self.assertEqual(len(mi_group_ids), 1)
        self.assertTrue(all(
            node['metadata']['mi_independent_protection']
            for node in mcb_nodes
        ))

    def test_build_project_sld_payload_keeps_duplicate_line_ids_distinct_by_line_uid(self):
        lines = make_rich_sld_project_snapshot('p1', ['LINE-DUP', 'LINE-DUP'])
        expected_line_uids = {str(line.uid) for line in lines}

        payload = build_project_sld_payload('p1')

        self.assertEqual(payload['meta']['branch_count'], 2)
        self.assertEqual(payload['meta']['node_count'], 28)
        self.assertEqual([group['line_id'] for group in payload['line_groups']], ['LINE-DUP', 'LINE-DUP'])
        self.assertEqual({group['line_uid'] for group in payload['line_groups']}, expected_line_uids)
        self.assertTrue(all(group['branch_indices'] == [1] for group in payload['line_groups']))

        component_ids = [node['component_id'] for node in payload['nodes']]
        component_uids = [node['component_uid'] for node in payload['nodes']]
        node_line_uids = {node['line_uid'] for node in payload['nodes']}
        edge_line_uids = {edge['line_uid'] for edge in payload['edges']}

        self.assertEqual(len(component_ids), len(set(component_ids)))
        self.assertEqual(len(component_uids), len(set(component_uids)))
        self.assertEqual(node_line_uids, expected_line_uids)
        self.assertEqual(edge_line_uids, expected_line_uids)
        for line_uid in expected_line_uids:
            self.assertTrue(any(f'line_uid:{line_uid}' in component_id for component_id in component_ids))
        self.assertTrue(any('line:LINE-DUP' in component_id for component_id in component_ids))

    def test_build_project_sld_payload_filters_line_at_query_layer(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        payload = build_project_sld_payload('p1', line_id='line-002')

        self.assertEqual(payload['meta']['branch_count'], 1)
        self.assertEqual([group['line_id'] for group in payload['line_groups']], ['LINE-002'])
        self.assertTrue(payload['nodes'])
        self.assertTrue(all(node['line_id'] == 'LINE-002' for node in payload['nodes']))
        self.assertTrue(all(edge['line_ids'] == ['LINE-002'] for edge in payload['edges']))

    def test_build_project_sld_payload_supports_partial_line_filtering(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'UNIT-003'])

        payload = build_project_sld_payload('p1', line_id='002')

        self.assertEqual(payload['meta']['branch_count'], 1)
        self.assertEqual([group['line_id'] for group in payload['line_groups']], ['LINE-002'])
        self.assertTrue(payload['nodes'])
        self.assertTrue(all(node['line_id'] == 'LINE-002' for node in payload['nodes']))

    def test_build_project_sld_payload_applies_cable_length_override_metadata(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        generated_payload = build_project_sld_payload('p1')
        cable_node = next(node for node in generated_payload['nodes'] if node['component_type'] == 'Cable3C')

        CableScheduleOverride.objects.create(
            project_id='p1',
            component_id=cable_node['component_id'],
            component_uid=cable_node['component_uid'],
            display_tag=cable_node['display_tag'],
            component_type=cable_node['component_type'],
            line_id=cable_node['line_id'],
            line_uid=cable_node['line_uid'],
            branch_index=cable_node['branch_index'],
            circuit_index=cable_node['circuit_index'],
            generated_length_m=(cable_node['metadata'] or {}).get('length_m'),
            manual_length_m=123.45,
            manual_cable_size='4C x 2.5',
            remarks='Routed via field JB rack.',
        )

        payload = build_project_sld_payload('p1')
        adjusted = next(node for node in payload['nodes'] if node['component_id'] == cable_node['component_id'])

        self.assertEqual(adjusted['metadata']['length_m'], 123.45)
        self.assertEqual(adjusted['metadata']['manual_length_m'], 123.45)
        self.assertEqual(adjusted['metadata']['cable_size'], '4C x 2.5')
        self.assertTrue(adjusted['metadata']['cable_override_active'])
        self.assertEqual(adjusted['metadata']['cable_override_remarks'], 'Routed via field JB rack.')

    def test_build_project_sld_payload_adds_cold_cable_metadata_to_cable_nodes(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        size_cold_cables_for_project('p1')

        payload = build_project_sld_payload('p1')
        cable4c = next(node for node in payload['nodes'] if node['component_type'] == 'Cable4C')
        cable3c = next(node for node in payload['nodes'] if node['component_type'] == 'Cable3C')

        self.assertEqual(cable4c['metadata']['cold_cable_calculated_size'], '4C x 2.5 mm2')
        self.assertEqual(cable3c['metadata']['cold_cable_calculated_size'], '3C x 2.5 mm2')
        self.assertEqual(cable4c['metadata']['cold_cable']['calculated_size'], '4C x 2.5 mm2')
        self.assertEqual(cable3c['metadata']['cold_cable']['calculated_size'], '3C x 2.5 mm2')
        self.assertIn(cable3c['metadata']['cold_cable_status'], {'selected', 'review_required'})
        self.assertEqual(cable3c['metadata']['cold_cable_vd_status'], 'pass')
        self.assertEqual(cable3c['metadata']['cold_cable_fault_status'], 'pass')
        self.assertGreater(cable3c['metadata']['cold_cable']['conductor_mass_mt'], 0)

    def test_build_project_sld_payload_flags_manual_cable_size_below_calculated_size(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        size_cold_cables_for_project('p1')
        generated_payload = build_project_sld_payload('p1')
        cable_node = next(node for node in generated_payload['nodes'] if node['component_type'] == 'Cable3C')
        CableScheduleOverride.objects.create(
            project_id='p1',
            component_id=cable_node['component_id'],
            component_uid=cable_node['component_uid'],
            display_tag=cable_node['display_tag'],
            component_type=cable_node['component_type'],
            line_id=cable_node['line_id'],
            line_uid=cable_node['line_uid'],
            branch_index=cable_node['branch_index'],
            circuit_index=cable_node['circuit_index'],
            generated_length_m=(cable_node['metadata'] or {}).get('length_m'),
            manual_cable_size='3C x 1.5',
        )

        payload = build_project_sld_payload('p1')
        adjusted = next(node for node in payload['nodes'] if node['component_id'] == cable_node['component_id'])

        self.assertEqual(adjusted['metadata']['manual_size_review_status'], 'undersized')
        self.assertIn('below the calculated cold cable size', adjusted['metadata']['manual_size_review_note'])

    def test_build_project_sld_payload_adds_tracer_selection_metadata(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]
        AlternateTracer.objects.create(
            line=line,
            option_rank=1,
            v_uid='V-ALT-001',
            a_coeff=0.0,
            b_coeff=0.0,
            c_coeff=80.0,
            power_at_startup_t=10.0,
            ohm_per_km=1.0,
            res_corrFactor_mica=1.0,
            tracer_family='SR-ALT',
            voltage_float=230.0,
            voltage_correction_factor=1.0,
            power_output=24.5,
            spiral_factor=1.15,
            tracer_length=21.25,
            tracer_with_margin=23.38,
        )

        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')
        tracer_selection = tracer_node['metadata']['tracer_selection']

        self.assertEqual(tracer_selection['selected']['v_uid'], 'V-001')
        self.assertEqual(tracer_selection['selected']['tracer_family'], 'SR')
        self.assertEqual(tracer_selection['selected']['power_output'], 30.0)
        self.assertEqual(tracer_node['display_name'], 'V-001')
        self.assertEqual(tracer_selection['alternate_count'], 1)
        self.assertTrue(tracer_selection['override_supported'])
        self.assertEqual(tracer_selection['alternatives'][0]['v_uid'], 'V-ALT-001')
        self.assertEqual(tracer_selection['alternatives'][0]['option_rank'], 1)
        self.assertEqual(tracer_selection['alternatives'][0]['tracer_family'], 'SR-ALT')

    def test_build_project_sld_payload_adds_mi_alternative_metadata(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]
        mi_result = make_mi_selection_result(line)
        expected_uid = f'MI:{mi_result.heater.part_number}:{mi_result.cold_lead_option_code}'

        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')
        tracer_selection = tracer_node['metadata']['tracer_selection']

        self.assertTrue(tracer_selection['override_supported'])
        self.assertEqual(tracer_selection['alternate_count'], 1)
        self.assertEqual(tracer_selection['alternatives'][0]['v_uid'], expected_uid)
        self.assertEqual(tracer_selection['alternatives'][0]['tracer_family'], 'MI')
        self.assertEqual(tracer_selection['alternatives'][0]['heater_part_number'], mi_result.heater.part_number)

    def test_build_project_sld_payload_applies_tracer_override_metadata(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]
        AlternateTracer.objects.create(
            line=line,
            option_rank=1,
            v_uid='V-ALT-001',
            a_coeff=0.0,
            b_coeff=0.0,
            c_coeff=80.0,
            power_at_startup_t=10.0,
            ohm_per_km=1.0,
            res_corrFactor_mica=1.0,
            tracer_family='SR-ALT',
            voltage_float=230.0,
            voltage_correction_factor=1.0,
            power_output=24.5,
            spiral_factor=1.15,
            tracer_length=21.25,
            tracer_with_margin=23.38,
        )
        TracerSelectionOverride.objects.create(
            project_id='p1',
            line=line,
            selected_v_uid='V-ALT-001',
            selected_option_rank=1,
            remarks='Use alternate after engineering review.',
        )

        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')
        tracer_selection = tracer_node['metadata']['tracer_selection']

        self.assertTrue(tracer_selection['override_active'])
        self.assertEqual(tracer_selection['selected']['v_uid'], 'V-ALT-001')
        self.assertEqual(tracer_selection['generated_selected']['v_uid'], 'V-001')
        self.assertEqual(tracer_selection['override_remarks'], 'Use alternate after engineering review.')

    def test_build_project_sld_payload_is_deterministic_for_repeated_builds(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        first_payload = build_project_sld_payload('p1')
        second_payload = build_project_sld_payload('p1')

        self.assertEqual(first_payload, second_payload)
        self.assertEqual(
            json.dumps(first_payload, sort_keys=True),
            json.dumps(second_payload, sort_keys=True),
        )
        self.assertEqual(
            [group['line_id'] for group in first_payload['line_groups']],
            ['LINE-001', 'LINE-002'],
        )

    def test_build_project_sld_payload_applies_active_topology_edit(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        generated_payload = build_project_sld_payload('p1')
        edited_payload = deepcopy(generated_payload)
        edited_payload['nodes'][0]['display_tag'] = f"{edited_payload['nodes'][0]['display_tag']}-M"

        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={'sld_payload': edited_payload},
            validation_summary={'status': 'passed'},
        )

        payload = build_project_sld_payload('p1')

        self.assertTrue(payload['meta']['has_topology_edit'])
        self.assertEqual(payload['meta']['topology_edit_type'], 'combine_feeders')
        self.assertTrue(any(node['display_tag'].endswith('-M') for node in payload['nodes']))

    def test_build_project_sld_payload_falls_back_for_legacy_branch_json(self):
        line = make_calculated_project_snapshot()

        payload = build_project_sld_payload('p1')

        self.assertEqual(payload['project_id'], 'p1')
        self.assertEqual(payload['schema_version'], SLD_GRAPH_SCHEMA_VERSION)
        self.assertEqual(payload['meta']['branch_count'], 1)
        self.assertEqual(payload['meta']['node_count'], 4)
        self.assertEqual(payload['meta']['edge_count'], 3)
        self.assertEqual(
            payload['line_groups'],
            [{'line_id': line.line_id, 'line_uid': str(line.uid), 'branch_indices': [1]}],
        )

        nodes_by_tag = {node['display_tag']: node for node in payload['nodes']}
        self.assertEqual(nodes_by_tag['MCB_001']['component_type'], 'MCB')
        self.assertEqual(nodes_by_tag['JB3PH_001']['component_type'], 'JB3PH')
        self.assertEqual(nodes_by_tag['Tracer_001']['component_type'], 'Tracer')
        self.assertEqual(nodes_by_tag['Tracer_002']['component_type'], 'Tracer')

        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in payload['edges']
        }
        self.assertIn(
            (nodes_by_tag['MCB_001']['component_id'], nodes_by_tag['JB3PH_001']['component_id']),
            edge_pairs,
        )
        self.assertIn(
            (nodes_by_tag['JB3PH_001']['component_id'], nodes_by_tag['Tracer_001']['component_id']),
            edge_pairs,
        )
        self.assertIn(
            (nodes_by_tag['JB3PH_001']['component_id'], nodes_by_tag['Tracer_002']['component_id']),
            edge_pairs,
        )


class SldValidationTests(TestCase):
    def test_audit_tagged_component_schema_passes_rich_generated_branches(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        audit = audit_tagged_component_schema('p1')

        self.assertEqual(audit['branch_count'], 2)
        self.assertEqual(audit['issue_count'], 0)
        self.assertTrue(audit['ready_for_strict_schema'])

    def test_audit_tagged_component_schema_reports_legacy_branches(self):
        make_calculated_project_snapshot()

        audit = audit_tagged_component_schema('p1')

        self.assertEqual(audit['branch_count'], 1)
        self.assertFalse(audit['ready_for_strict_schema'])
        self.assertGreaterEqual(audit['issue_count'], 1)
        self.assertIn('missing_schema_version', {issue['code'] for issue in audit['issues']})

    def test_validate_project_sld_payload_passes_for_rich_project_distribution(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        report = validate_project_sld_payload('p1')

        self.assertEqual(report['project_id'], 'p1')
        self.assertEqual(report['status'], 'passed')
        self.assertEqual(report['summary']['failed_count'], 0)
        self.assertEqual(report['summary']['warning_count'], 0)
        self.assertGreater(report['summary']['passed_count'], 0)
        self.assertTrue(any(check['code'] == 'schema_version_supported' for check in report['checks']))
        self.assertTrue(any(check['code'] == 'tagged_component_schema_coverage' for check in report['checks']))
        self.assertTrue(all(check['status'] == 'passed' for check in report['checks']))
        self.assertTrue(all(check['status'] == 'passed' for check in report['branch_checks']))

    def test_validate_project_sld_payload_warns_for_legacy_schema_coverage(self):
        make_calculated_project_snapshot()

        report = validate_project_sld_payload('p1')

        schema_check = next(
            check for check in report['checks']
            if check['code'] == 'tagged_component_schema_coverage'
        )
        self.assertIn(report['status'], {'warning', 'failed'})
        self.assertEqual(schema_check['status'], 'warning')

    def test_validate_project_sld_payload_flags_unsupported_schema_version(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        payload['schema_version'] = 999

        report = validate_project_sld_payload('p1', payload=payload)

        schema_check = next(check for check in report['checks'] if check['code'] == 'schema_version_supported')
        self.assertEqual(report['status'], 'failed')
        self.assertEqual(schema_check['status'], 'failed')

    def test_validate_project_sld_payload_accepts_duplicate_line_ids_with_distinct_line_uids(self):
        lines = make_rich_sld_project_snapshot('p1', ['LINE-DUP', 'LINE-DUP'])
        expected_line_uids = {str(line.uid) for line in lines}

        report = validate_project_sld_payload('p1')

        self.assertEqual(report['status'], 'passed')
        self.assertEqual({check['line_uid'] for check in report['branch_checks']}, expected_line_uids)
        self.assertTrue(all(check['status'] == 'passed' for check in report['branch_checks']))

    def test_validate_project_sld_payload_flags_duplicate_display_tags(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        payload['nodes'][1]['display_tag'] = payload['nodes'][0]['display_tag']

        report = validate_project_sld_payload('p1', payload=payload)

        duplicate_check = next(check for check in report['checks'] if check['code'] == 'unique_display_tags')
        self.assertEqual(report['status'], 'failed')
        self.assertEqual(duplicate_check['status'], 'failed')

    def test_validate_project_sld_payload_accepts_applied_topology_edit_layer(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        edited_payload = deepcopy(build_project_sld_payload('p1'))
        removed_component_id = edited_payload['nodes'][0]['component_id']
        edited_payload['nodes'] = [
            node for node in edited_payload['nodes'] if node['component_id'] != removed_component_id
        ]
        edited_payload['edges'] = [
            edge for edge in edited_payload['edges']
            if edge['from_component_id'] != removed_component_id and edge['to_component_id'] != removed_component_id
        ]
        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='split_circuits',
            status='applied',
            baseline_fingerprint='baseline-b',
            edit_payload={'sld_payload': edited_payload},
            validation_summary={'status': 'passed'},
        )

        report = validate_project_sld_payload('p1')

        self.assertEqual(report['status'], 'passed')
        self.assertTrue(any(check['code'] == 'topology_edit_applied' for check in report['checks']))
        self.assertTrue(all(check['status'] == 'passed' for check in report['branch_checks']))


class SldLayoutTests(TestCase):
    def test_sld_layout_view_saves_loads_and_resets_component_positions(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        positions = {
            node['component_id']: {'x': 100 + index * 25, 'y': 200 + index * 10}
            for index, node in enumerate(payload['nodes'], start=1)
        }

        save_response = self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': positions}),
            content_type='application/json',
        )

        self.assertEqual(save_response.status_code, 200)
        save_payload = save_response.json()
        self.assertEqual(save_payload['saved_count'], len(payload['nodes']))
        self.assertEqual(SLDNodeLayout.objects.filter(project_id='p1').count(), len(payload['nodes']))

        load_response = self.client.get(reverse('sld_layout_view'), {'project_id': 'p1'})

        self.assertEqual(load_response.status_code, 200)
        layout_payload = load_response.json()
        self.assertTrue(layout_payload['meta']['has_saved_layout'])
        first_component_id = payload['nodes'][0]['component_id']
        self.assertEqual(layout_payload['positions'][first_component_id]['x'], positions[first_component_id]['x'])
        self.assertEqual(layout_payload['positions'][first_component_id]['y'], positions[first_component_id]['y'])

        reset_response = self.client.post(
            reverse('sld_layout_reset_view'),
            data=json.dumps({'project_id': 'p1'}),
            content_type='application/json',
        )

        self.assertEqual(reset_response.status_code, 200)
        self.assertEqual(reset_response.json()['deleted_count'], len(payload['nodes']))
        self.assertFalse(SLDNodeLayout.objects.filter(project_id='p1').exists())

        reloaded_layout = get_project_sld_layout('p1', payload=payload)
        self.assertFalse(reloaded_layout['meta']['has_saved_layout'])
        self.assertEqual(reloaded_layout['meta']['saved_count'], 0)

    def test_sld_layout_view_merges_partial_position_updates_without_deleting_other_saved_nodes(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        all_positions = {
            node['component_id']: {'x': 100 + index * 10, 'y': 200 + index * 5}
            for index, node in enumerate(payload['nodes'], start=1)
        }
        self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': all_positions}),
            content_type='application/json',
        )

        preserved_component_id = payload['nodes'][0]['component_id']
        updated_component_id = payload['nodes'][1]['component_id']
        partial_update = {
            updated_component_id: {'x': 999, 'y': 777},
        }

        response = self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': partial_update}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        layout_payload = self.client.get(reverse('sld_layout_view'), {'project_id': 'p1'}).json()
        self.assertEqual(layout_payload['positions'][updated_component_id]['x'], 999)
        self.assertEqual(layout_payload['positions'][updated_component_id]['y'], 777)
        self.assertEqual(layout_payload['positions'][preserved_component_id]['x'], all_positions[preserved_component_id]['x'])
        self.assertEqual(layout_payload['positions'][preserved_component_id]['y'], all_positions[preserved_component_id]['y'])
        self.assertEqual(layout_payload['meta']['save_mode'], 'merge')

    def test_sld_layout_view_filters_saved_positions_for_selected_line(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        payload = build_project_sld_payload('p1')
        all_positions = {
            node['component_id']: {'x': 100 + index * 10, 'y': 200 + index * 5}
            for index, node in enumerate(payload['nodes'], start=1)
        }
        line_two_component_ids = {
            node['component_id']
            for node in payload['nodes']
            if 'LINE-002' in node.get('line_ids', [])
        }

        save_response = self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': all_positions}),
            content_type='application/json',
        )
        self.assertEqual(save_response.status_code, 200)

        load_response = self.client.get(
            reverse('sld_layout_view'),
            {'project_id': 'p1', 'line_id': 'line-002'},
        )

        self.assertEqual(load_response.status_code, 200)
        layout_payload = load_response.json()
        self.assertEqual(layout_payload['meta']['node_count'], len(line_two_component_ids))
        self.assertEqual(set(layout_payload['positions']), line_two_component_ids)
        self.assertTrue(layout_payload['meta']['has_saved_layout'])

    def test_sld_layout_view_saves_focused_line_without_pruning_other_lines(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        payload = build_project_sld_payload('p1')
        all_positions = {
            node['component_id']: {'x': 100 + index * 10, 'y': 200 + index * 5}
            for index, node in enumerate(payload['nodes'], start=1)
        }
        line_one_component_id = next(
            node['component_id'] for node in payload['nodes'] if node['line_id'] == 'LINE-001'
        )
        line_two_component_id = next(
            node['component_id'] for node in payload['nodes'] if node['line_id'] == 'LINE-002'
        )
        self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': all_positions}),
            content_type='application/json',
        )

        response = self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({
                'project_id': 'p1',
                'line_id': 'line-002',
                'positions': {line_two_component_id: {'x': 555, 'y': 666}},
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        full_layout = self.client.get(reverse('sld_layout_view'), {'project_id': 'p1'}).json()
        self.assertEqual(full_layout['positions'][line_two_component_id]['x'], 555)
        self.assertEqual(full_layout['positions'][line_two_component_id]['y'], 666)
        self.assertEqual(
            full_layout['positions'][line_one_component_id]['x'],
            all_positions[line_one_component_id]['x'],
        )
        self.assertEqual(
            full_layout['positions'][line_one_component_id]['y'],
            all_positions[line_one_component_id]['y'],
        )

    def test_sld_layout_save_keeps_generated_topology_unchanged(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        before_payload = build_project_sld_payload('p1')
        shifted_positions = {
            node['component_id']: {'x': 400 + index * 20, 'y': 300 + index * 15}
            for index, node in enumerate(before_payload['nodes'], start=1)
        }

        response = self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': shifted_positions}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        after_payload = build_project_sld_payload('p1')
        self.assertEqual(after_payload['line_groups'], before_payload['line_groups'])
        self.assertEqual(after_payload['edges'], before_payload['edges'])
        self.assertEqual(
            [node['component_id'] for node in after_payload['nodes']],
            [node['component_id'] for node in before_payload['nodes']],
        )

    def test_sld_layout_view_returns_unknown_line_filter_error(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(
            reverse('sld_layout_view'),
            {'project_id': 'p1', 'line_id': 'LINE-999'},
        )

        self.assertEqual(response.status_code, 404)
        self.assertIn('LINE-999', response.json()['error'])

    def test_saved_layout_survives_recalculation_when_component_ids_are_stable(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        original_payload = build_project_sld_payload('p1')
        positions = {
            node['component_id']: {'x': 120 + index * 15, 'y': 240 + index * 8}
            for index, node in enumerate(original_payload['nodes'], start=1)
        }
        save_project_response = self.client.post(
            reverse('sld_layout_view'),
            data=json.dumps({'project_id': 'p1', 'positions': positions}),
            content_type='application/json',
        )
        self.assertEqual(save_project_response.status_code, 200)

        project_settings = make_project_settings(proj_id='p1')
        tag_factory = ProjectTagFactory('p1')
        aggregated_results = {
            'heat_loss': [],
            'selected_tracers': [],
            'alternative_tracers': [],
            'power_distribution': [],
            'boq_per_line': {},
            'consolidated_boq': {},
            'tracer_power_param': [],
        }
        line = HeatTracingInput.objects.get(proj_id='p1', line_id='LINE-001')
        selected_tracer_template = {
            'V_UID': 'V-001',
            'A_Coeff': 0.0,
            'B_Coeff': 0.0,
            'C_Coeff': 100.0,
            'Power_at_Startup_T': 10.0,
            'Ohm_per_km': 1.0,
            'Res_corrFactor_Mica': 1.0,
            'Tracer_Family': 'SR',
            'Voltage_Float': 230.0,
            'Voltage_Correction_Factor': 1.0,
            'Power_Output': 30.0,
            'Spiral_Factor': 1.0,
            'Tracer_Length': 20.0,
            'Tracer_With_Margin': 20.0,
        }
        power_params = compute_power_params(
            make_line(uid=str(line.uid), line_id=line.line_id),
            project_settings,
            make_asme_table(),
            selected_tracer_template,
        )
        aggregated_results['selected_tracers'].append({**selected_tracer_template, 'uid': line.uid})
        aggregated_results['power_distribution'].append(
            compute_power_distribution(power_params, project_settings, tag_factory=tag_factory)
        )
        aggregated_results['tracer_power_param'].append(power_params)

        self.assertTrue(store_calculated_results('p1', aggregated_results))

        reloaded_payload = build_project_sld_payload('p1')
        reloaded_layout = get_project_sld_layout('p1', payload=reloaded_payload)
        self.assertTrue(reloaded_layout['meta']['has_saved_layout'])
        first_component_id = original_payload['nodes'][0]['component_id']
        self.assertEqual(reloaded_layout['positions'][first_component_id]['x'], positions[first_component_id]['x'])
        self.assertEqual(reloaded_layout['positions'][first_component_id]['y'], positions[first_component_id]['y'])


class ColdCableFoundationTests(TestCase):
    def test_project_data_defaults_include_cold_cable_basis(self):
        project = make_project_record(proj_id='p-cold')

        self.assertEqual(project.cable_standard, 'IEC_60502_1')
        self.assertEqual(project.cable_conductor_material, 'Cu')
        self.assertEqual(project.cable_insulation_type, 'XLPE')
        self.assertEqual(project.cable_install_method, 'E')
        self.assertEqual(float(project.cable_grouping_derating), 1.0)
        self.assertEqual(project.min_cold_cable_size_mm2, 'CALCULATED')
        self.assertEqual(project.mcb_curve, 'C')
        self.assertTrue(project.rcd_provided)

    def test_project_data_install_method_excludes_single_core_method(self):
        form = ProjectDataForm()
        install_methods = {value for value, _label in form.fields['cable_install_method'].choices}

        self.assertIn('E', install_methods)
        self.assertIn('D2', install_methods)
        self.assertNotIn('B2', install_methods)
        self.assertNotIn('C', install_methods)
        self.assertNotIn('D1', install_methods)
        self.assertNotIn('F', install_methods)

    def test_project_data_form_shows_d2_as_under_development(self):
        form = ProjectDataForm()
        rendered = str(form['cable_install_method'])

        self.assertIn('D2 - Direct buried in ground (coming soon)', rendered)
        self.assertIn('value="D2" disabled', rendered)
        self.assertIn('Method E is active for this phase', form.fields['cable_install_method'].help_text)
        self.assertIn('under development', form.fields['cable_install_method'].help_text)

    def test_project_data_form_rejects_forced_d2_install_method(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        form = ProjectDataForm(data=make_project_form_payload(cable_install_method='D2'))

        self.assertFalse(form.is_valid())
        self.assertIn('under development', form.errors['cable_install_method'][0])

    def test_project_data_rejects_invalid_cold_cable_grouping_derating(self):
        with self.assertRaises(ValidationError):
            make_project_record(proj_id='p-cold-invalid', cable_grouping_derating=0.24)

    def test_project_data_form_guides_cold_cable_grouping_derating_range(self):
        form = ProjectDataForm()
        widget_attrs = form.fields['cable_grouping_derating'].widget.attrs

        self.assertEqual(widget_attrs['min'], '0.25')
        self.assertEqual(widget_attrs['max'], '1.0')
        self.assertEqual(widget_attrs['step'], '0.001')

    def test_populate_cold_cable_catalogue_command_is_idempotent(self):
        call_command('populate_cold_cable_catalogue')
        call_command('populate_cold_cable_catalogue')

        rows = ColdCableCatalogue.objects.filter(
            cable_standard='IEC_60502_1',
            conductor_material='Cu',
            insulation_type='XLPE',
            installation_method='E',
        )
        self.assertEqual(rows.count(), 14)
        row = rows.get(core_count=4, conductor_size_mm2=10)
        self.assertEqual(row.ampacity_a, 52)
        self.assertEqual(row.resistance_mohm_per_m, 1.83)
        self.assertTrue(row.is_validated)

    def test_cold_cable_catalogue_readiness_summarises_methods(self):
        call_command('populate_cold_cable_catalogue')

        readiness = cold_cable_method_readiness('IEC_60502_1', 'Cu', 'XLPE')

        self.assertEqual(readiness['E']['status'], 'ready')
        self.assertEqual(readiness['E']['validated_rows'], 14)
        self.assertEqual(readiness['B2']['status'], 'unavailable')
        self.assertEqual(readiness['C']['status'], 'unavailable')
        self.assertEqual(readiness['D1']['status'], 'unavailable')
        self.assertEqual(readiness['D2']['status'], 'unavailable')

    def test_project_data_form_guides_install_method_catalogue_readiness(self):
        call_command('populate_cold_cable_catalogue')

        form = ProjectDataForm()
        help_text = form.fields['cable_install_method'].help_text

        self.assertIn('Method E is active for this phase', help_text)
        self.assertIn('Method D2 direct buried', help_text)
        self.assertIn('under development', help_text)

    def test_resolve_cable_lengths_uses_generated_branch_defaults(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related('distribution__line').get(distribution__line__proj_id='p1')

        result = resolve_cable_lengths(branch, project)

        self.assertEqual(result.branch_type, '3phJB')
        self.assertEqual(result.length_4c_m, 30.0)
        self.assertEqual(result.length_3c_m, 12.0)
        self.assertEqual(result.length_3c_total_m, 24.0)
        self.assertEqual(result.length_basis, 'project_default')
        self.assertFalse(result.length_missing)

    def test_resolve_cable_lengths_prefers_manual_cable_override(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related('distribution__line').get(distribution__line__proj_id='p1')
        cable3c_detail = branch.tagged_components['Downstream'][0]['component_details']['Cable3C']
        CableScheduleOverride.objects.create(
            project=project,
            component_id=cable3c_detail['component_id'],
            component_uid=cable3c_detail['component_uid'],
            display_tag=cable3c_detail['display_tag'],
            component_type='Cable3C',
            line_id='LINE-001',
            line_uid=str(branch.distribution.line.uid),
            branch_index=branch.branch_index,
            circuit_index=1,
            generated_length_m=12.0,
            manual_length_m=88.5,
        )

        result = resolve_cable_lengths(branch, project)

        self.assertEqual(result.length_4c_m, 30.0)
        self.assertEqual(result.length_3c_m, 88.5)
        self.assertEqual(result.length_3c_total_m, 100.5)
        self.assertEqual(result.length_3c_basis, 'manual_override')
        self.assertEqual(result.length_basis, 'manual_override')
        self.assertEqual(
            sorted((segment.length_m, segment.length_basis) for segment in result.length_3c_segments),
            [(12.0, 'project_default'), (88.5, 'manual_override')],
        )

    def test_resolve_cable_lengths_can_use_applied_topology_schedule_row(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related('distribution__line').get(distribution__line__proj_id='p1')
        generated_payload = build_project_sld_payload('p1')
        SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'cable_schedule_rows': [{
                    'distribution': {'line': {'line_id': 'LINE-001'}},
                    'branch_index': branch.branch_index,
                    'branch_type': 'manual_topology_edit',
                    'connected_to': 'manual',
                    'circuit_count': 2,
                    'cable_length_db_to_jb': 45.0,
                    'branch_cable_length_total_m': 40.0,
                    'tagged_components': {'MCB': branch.tagged_components['MCB']},
                }],
            },
            validation_summary={'status': 'passed'},
        )

        result = resolve_cable_lengths(branch, project)

        self.assertEqual(result.length_4c_m, 45.0)
        self.assertEqual(result.length_3c_m, 20.0)
        self.assertEqual(result.length_3c_total_m, 40.0)
        self.assertEqual(result.length_basis, 'topology_edit')

    def test_resolve_cable_lengths_uses_topology_db_route_for_direct_1ph_branch(self):
        make_direct_sld_project_snapshot('p-direct', 'LINE-001')
        project = ProjectData.objects.get(proj_id='p-direct')
        branch = PowerDistributionBranch.objects.select_related('distribution__line').get(distribution__line__proj_id='p-direct')
        generated_payload = build_project_sld_payload('p-direct')
        SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'cable_schedule_rows': [{
                    'distribution': {'line': {'line_id': 'LINE-001'}},
                    'branch_index': branch.branch_index,
                    'branch_type': 'manual_topology_edit',
                    'connected_to': 'manual',
                    'circuit_count': 1,
                    'cable_length_db_to_jb': 123.0,
                    'cable_length_jb_to_jb': None,
                    'branch_cable_length_total_m': 0,
                    'tagged_components': {'MCB': branch.tagged_components['MCB']},
                }],
            },
            validation_summary={'status': 'passed'},
        )

        result = resolve_cable_lengths(branch, project)

        self.assertEqual(result.branch_type, '1phJB')
        self.assertIsNone(result.length_4c_m)
        self.assertEqual(result.length_3c_m, 123.0)
        self.assertEqual(result.length_3c_total_m, 123.0)
        self.assertEqual(result.length_basis, 'topology_edit')
        self.assertFalse(result.length_missing)

    def test_resolve_cable_lengths_ignores_zero_topology_3c_length_and_uses_project_default(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related('distribution__line').get(distribution__line__proj_id='p1')
        generated_payload = build_project_sld_payload('p1')
        SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'cable_schedule_rows': [{
                    'distribution': {'line': {'line_id': 'LINE-001'}},
                    'branch_index': branch.branch_index,
                    'branch_type': 'manual_topology_edit',
                    'connected_to': 'manual',
                    'circuit_count': 2,
                    'cable_length_db_to_jb': 45.0,
                    'cable_length_jb_to_jb': None,
                    'branch_cable_length_total_m': 0,
                    'tagged_components': {'MCB': branch.tagged_components['MCB']},
                }],
            },
            validation_summary={'status': 'passed'},
        )

        result = resolve_cable_lengths(branch, project)

        self.assertEqual(result.length_4c_m, 45.0)
        self.assertEqual(result.length_4c_basis, 'topology_edit')
        self.assertEqual(result.length_3c_m, 12.0)
        self.assertEqual(result.length_3c_total_m, 24.0)
        self.assertEqual(result.length_3c_basis, 'project_default')
        self.assertEqual(result.length_basis, 'topology_edit')
        self.assertFalse(result.length_missing)

    def test_resolve_cable_lengths_does_not_substring_match_topology_line_id(self):
        make_rich_sld_project_snapshot('p1', ['LINE-1', 'LINE-10'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related('distribution__line').get(distribution__line__line_id='LINE-1')
        generated_payload = build_project_sld_payload('p1')
        SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'cable_schedule_rows': [{
                    'distribution': {'line': {'line_id': 'LINE-10'}},
                    'branch_index': branch.branch_index,
                    'branch_type': 'manual_topology_edit',
                    'connected_to': 'manual',
                    'circuit_count': 2,
                    'cable_length_db_to_jb': 99.0,
                    'branch_cable_length_total_m': 88.0,
                    'tagged_components': {'MCB': 'NOT-THE-BRANCH-MCB'},
                }],
            },
            validation_summary={'status': 'passed'},
        )

        result = resolve_cable_lengths(branch, project)

        self.assertEqual(result.length_4c_m, 30.0)
        self.assertEqual(result.length_3c_m, 12.0)
        self.assertEqual(result.length_basis, 'project_default')

    def test_build_cold_cable_sizing_input_normalises_branch_current_and_lengths(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related(
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p1')

        sizing_input = build_cold_cable_sizing_input(branch, project)

        self.assertEqual(sizing_input.project_id, 'p1')
        self.assertEqual(sizing_input.branch_type, '3phJB')
        self.assertEqual(sizing_input.heating_cable_type, 'SR')
        self.assertGreater(sizing_input.per_circuit_operating_current_a, 0)
        self.assertEqual(sizing_input.length_4c_m, 30.0)
        self.assertEqual(sizing_input.length_3c_m, 12.0)

    def test_calculate_k_temp_uses_conductor_and_site_temperatures(self):
        self.assertAlmostEqual(calculate_k_temp(40.0, 30.0, 90.0), math.sqrt(50.0 / 60.0), places=6)

    def test_calculate_k_temp_rejects_site_ambient_at_conductor_limit(self):
        with self.assertRaises(ValueError):
            calculate_k_temp(90.0, 30.0, 90.0)

    def test_get_conductor_resistance_at_temp_applies_copper_alpha(self):
        call_command('populate_cold_cable_catalogue')
        row = ColdCableCatalogue.objects.get(core_count=4, conductor_size_mm2=2.5)

        resistance = get_conductor_resistance_at_temp(row, 75.0)

        self.assertAlmostEqual(resistance, 7.41 * (1 + 0.00393 * 55.0), places=6)

    def test_conductor_temperature_uses_insulation_rating_basis(self):
        xlpe = ColdCableCatalogue(
            cable_standard='IEC_60502_1',
            cable_type_code='XLPE-test',
            conductor_material='Cu',
            insulation_type='XLPE',
            core_count=3,
            conductor_size_mm2=2.5,
            installation_method='E',
            ampacity_a=26.0,
            resistance_mohm_per_m=7.41,
            max_conductor_temp_c=90.0,
        )
        pvc = ColdCableCatalogue(
            cable_standard='IEC_60502_1',
            cable_type_code='PVC-test',
            conductor_material='Cu',
            insulation_type='PVC',
            core_count=3,
            conductor_size_mm2=2.5,
            installation_method='E',
            ampacity_a=26.0,
            resistance_mohm_per_m=7.41,
            max_conductor_temp_c=70.0,
        )

        self.assertEqual(conductor_operating_temperature(xlpe), 90.0)
        self.assertEqual(conductor_operating_temperature(pvc), 70.0)

    def test_conductor_temperature_rejects_unknown_insulation_type(self):
        unknown = ColdCableCatalogue(
            cable_standard='IEC_60502_1',
            cable_type_code='unknown-test',
            conductor_material='Cu',
            insulation_type='UNKNOWN',
            core_count=3,
            conductor_size_mm2=2.5,
            installation_method='E',
            ampacity_a=26.0,
            resistance_mohm_per_m=7.41,
            max_conductor_temp_c=90.0,
        )

        with self.assertRaisesMessage(ValueError, "Insulation type must be 'XLPE' or 'PVC'."):
            conductor_operating_temperature(unknown)

    def test_conductor_material_coefficients_and_density_are_available(self):
        self.assertEqual(conductor_temperature_coefficient('Cu'), 0.00393)
        self.assertEqual(conductor_density_kg_m3('Cu'), 8960.0)
        with self.assertRaisesMessage(ValueError, "Conductor material must be 'Cu'."):
            conductor_temperature_coefficient('Al')
        with self.assertRaisesMessage(ValueError, "Conductor material must be 'Cu'."):
            conductor_density_kg_m3('Al')

    def test_calculate_vd_uses_single_and_three_phase_factors(self):
        one_phase = calculate_vd(10.0, 10.0, 100.0, '1phase', 230.0)
        three_phase = calculate_vd(10.0, 10.0, 100.0, '3phase', 230.0)

        self.assertAlmostEqual(one_phase.vd_v, 20.0, places=6)
        self.assertAlmostEqual(one_phase.vd_pct, 20.0 / 230.0 * 100.0, places=6)
        self.assertAlmostEqual(three_phase.vd_v, math.sqrt(3) * 10.0, places=6)

    def test_calculate_vd_exact_allowable_threshold_is_not_rounded_over(self):
        result = calculate_vd(10.0, 5.75, 100.0, '1phase', 230.0)

        self.assertAlmostEqual(result.vd_v, 11.5, places=6)
        self.assertAlmostEqual(result.vd_pct, 5.0, places=6)

    def test_find_minimum_cable_for_ampacity_applies_derating_and_minimum_size(self):
        call_command('populate_cold_cable_catalogue')
        project = make_project_record(
            proj_id='p-ampacity',
            max_amb_t=40.0,
            cable_grouping_derating=0.8,
            min_cold_cable_size_mm2='4',
        )

        selection = find_minimum_cable_for_ampacity(project, 4, 20.0)

        self.assertEqual(selection.status, 'selected')
        self.assertEqual(selection.catalogue.core_count, 4)
        self.assertEqual(selection.catalogue.conductor_size_mm2, 4.0)
        self.assertAlmostEqual(selection.k_temp, math.sqrt(50.0 / 60.0), places=6)
        self.assertAlmostEqual(selection.k_group, 0.8)
        self.assertGreaterEqual(selection.derated_ampacity_a, 20.0)

    def test_cold_cable_sizing_clamps_stale_grouping_derating_below_current_floor(self):
        call_command('populate_cold_cable_catalogue')
        project = make_project_record(proj_id='p-stale-derating', max_amb_t=40.0)
        ProjectData.objects.filter(proj_id='p-stale-derating').update(cable_grouping_derating=0.10)
        project.refresh_from_db()

        selection = find_minimum_cable_for_ampacity(project, 4, 5.0)

        self.assertEqual(selection.status, 'selected')
        self.assertEqual(selection.k_group, 0.25)
        self.assertTrue(any('0.1 is outside the allowed range' in note for note in selection.review_notes))

    def test_find_minimum_cable_for_ampacity_rejects_thermally_unsuitable_site(self):
        call_command('populate_cold_cable_catalogue')
        project = make_project_record(proj_id='p-hot-site', max_amb_t=95.0)

        selection = find_minimum_cable_for_ampacity(project, 3, 1.0)

        self.assertEqual(selection.status, 'unsizeable')
        self.assertIsNone(selection.catalogue)
        self.assertTrue(any('Site ambient temperature' in note for note in selection.review_notes))

    def test_optimise_cable_pair_finds_minimum_cost_voltage_drop_pair(self):
        call_command('populate_cold_cable_catalogue')
        project = make_project_record(proj_id='p-opt', allowablevdrop=5.0)
        sizing_input = ColdCableSizingInput(
            project_id='p-opt',
            line_id='LINE-001',
            line_uid='1',
            branch_index=1,
            branch_type='3phJB',
            circuit_count=2,
            heating_cable_type='SR',
            per_circuit_operating_current_a=8.0,
            line_operating_current_a=16.0,
            breaker_size_a=10.0,
            length_4c_m=150.0,
            length_3c_m=10.0,
            length_basis='manual_override',
            length_missing=False,
        )

        result = optimise_cable_pair(project, sizing_input)

        self.assertEqual(result.status, 'selected')
        self.assertEqual(result.selection_4c.catalogue.conductor_size_mm2, 10.0)
        self.assertEqual(result.selection_3c.catalogue.conductor_size_mm2, 16.0)
        self.assertLessEqual(result.vd_total_pct, 5.0)
        self.assertEqual(result.conductor_volume_proxy, 5460.0)

    def test_size_cold_cable_for_branch_marks_total_voltage_drop_unsizeable(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        project.allowablevdrop = 0.01
        project.save()
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p1')

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.sizing_status, 'unsizeable')
        self.assertEqual(result.vd_status, 'fail')
        self.assertTrue(any('voltage-drop' in note for note in result.review_notes))

    def test_size_cold_cable_for_branch_guides_unavailable_install_method(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        project.cable_install_method = 'D2'
        project.save(update_fields=['cable_install_method'])
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p1')

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.sizing_status, 'unsizeable')
        notes = ' '.join(result.review_notes)
        self.assertIn('installation method D2', notes)
        self.assertIn('Select a catalogue-ready method', notes)

    def test_size_cold_cable_for_direct_1ph_branch_skips_optimisation(self):
        call_command('populate_cold_cable_catalogue')
        make_direct_sld_project_snapshot('p-direct', 'LINE-001')
        project = ProjectData.objects.get(proj_id='p-direct')
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-direct')

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(branch.branch_type, '1phJB')
        self.assertIsNone(result.cable_4c_size_mm2)
        self.assertEqual(result.cable_3c_size_mm2, 2.5)
        self.assertFalse(result.optimization_run)
        self.assertEqual(result.vd_status, 'pass')
        self.assertIsNotNone(result.cable_3c_vd_pct)

    def test_mcb_curve_threshold_uses_lower_instantaneous_bound(self):
        self.assertEqual(mcb_instantaneous_factor('B'), 3.0)
        self.assertEqual(mcb_instantaneous_factor('C'), 5.0)
        self.assertEqual(mcb_instantaneous_factor('D'), 10.0)

    def test_l_pe_fault_loop_check_returns_expected_pass_fail_statuses(self):
        call_command('populate_cold_cable_catalogue')
        row = ColdCableCatalogue.objects.get(core_count=3, conductor_size_mm2=2.5)

        passing = check_fault_l_pe(
            row, 30.0, row, 10.0, 230.0, 10.0, 'C', True,
            feeder_conductor_temp_c=90.0,
            branch_conductor_temp_c=90.0,
        )
        rcd_review = check_fault_l_pe(
            row, 600.0, row, 600.0, 230.0, 10.0, 'C', True,
            feeder_conductor_temp_c=90.0,
            branch_conductor_temp_c=90.0,
        )
        no_rcd_fail = check_fault_l_pe(
            row, 600.0, row, 600.0, 230.0, 10.0, 'C', False,
            feeder_conductor_temp_c=90.0,
            branch_conductor_temp_c=90.0,
        )

        self.assertEqual(passing.status, 'pass')
        self.assertGreaterEqual(passing.fault_current_a, passing.threshold_current_a)
        self.assertEqual(rcd_review.status, 'review_required')
        self.assertEqual(no_rcd_fail.status, 'fail')

    def test_size_cold_cable_for_branch_flags_feeder_l_pe_fault_review_with_rcd(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p-fault-4c', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p-fault-4c')
        project.ckt_ln = 600.0
        project.loop_ln = 10.0
        project.allowablevdrop = 50.0
        project.mcb_curve = 'D'
        project.save()
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-fault-4c')
        branch.cable_length_db_to_jb = 600.0
        branch.cable_length_jb_to_jb = 10.0
        tagged_components = branch.tagged_components
        tagged_components['component_details']['MCB']['metadata']['breaker_size'] = 10.0
        branch.tagged_components = tagged_components
        branch.save(update_fields=['cable_length_db_to_jb', 'cable_length_jb_to_jb', 'tagged_components'])

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.fault_loop_status, 'review_required')
        self.assertEqual(result.cable_4c_size_mm2, 2.5)
        self.assertEqual(result.cable_3c_size_mm2, 2.5)
        self.assertEqual(result.cable_4c_conductor_temp_c, 90.0)
        self.assertLess(result.fault_current_l_pe_a, result.breaker_size_a * 10.0)
        self.assertEqual(result.fault_loop_basis['basis'], 'single_phase_l_pe_complete_cold_cable_path')

    def test_size_cold_cable_assumes_zero_source_impedance_when_project_fault_rating_invalid(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p-source-missing', ['LINE-001'])
        ProjectData.objects.filter(proj_id='p-source-missing').update(eht_db_fault_rating_ka=0)
        project = ProjectData.objects.get(proj_id='p-source-missing')
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-source-missing')

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.fault_loop_basis['source_impedance_ohm'], 0.0)
        self.assertEqual(result.fault_loop_basis['source_impedance_note'], SOURCE_IMPEDANCE_ASSUMED_ZERO_NOTE)
        self.assertIn(SOURCE_IMPEDANCE_ASSUMED_ZERO_NOTE, result.review_notes)

    def test_size_cold_cable_for_3c_fault_review_required_with_rcd(self):
        call_command('populate_cold_cable_catalogue')
        make_direct_sld_project_snapshot('p-rcd-review', 'LINE-001')
        project = ProjectData.objects.get(proj_id='p-rcd-review')
        project.loop_ln = 900.0
        project.allowablevdrop = 10.0
        project.rcd_provided = True
        project.save()
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-rcd-review')
        branch.cable_length_db_to_jb = 900.0
        tagged_components = branch.tagged_components
        tagged_components['component_details']['MCB']['metadata']['breaker_size'] = 10.0
        branch.tagged_components = tagged_components
        branch.save(update_fields=['cable_length_db_to_jb', 'tagged_components'])

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.sizing_status, 'review_required')
        self.assertEqual(result.fault_loop_status, 'review_required')
        self.assertEqual(result.vd_status, 'pass')
        self.assertTrue(any('Source impedance' in note for note in result.review_notes))

    def test_size_cold_cable_for_3c_fault_upsizes_without_rcd_when_larger_cable_can_pass(self):
        call_command('populate_cold_cable_catalogue')
        make_direct_sld_project_snapshot('p-no-rcd-upsize', 'LINE-001')
        project = ProjectData.objects.get(proj_id='p-no-rcd-upsize')
        project.loop_ln = 150.0
        project.allowablevdrop = 100.0
        project.mcb_curve = 'D'
        project.rcd_provided = False
        project.save()
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-no-rcd-upsize')
        branch.cable_length_db_to_jb = 150.0
        tagged_components = branch.tagged_components
        tagged_components['component_details']['MCB']['metadata']['breaker_size'] = 10.0
        branch.tagged_components = tagged_components
        branch.save(update_fields=['cable_length_db_to_jb', 'tagged_components'])

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.sizing_status, 'review_required')
        self.assertEqual(result.fault_loop_status, 'pass')
        self.assertEqual(result.cable_3c_size_mm2, 6.0)

    def test_size_cold_cable_for_3c_fault_hard_fails_without_rcd(self):
        call_command('populate_cold_cable_catalogue')
        make_direct_sld_project_snapshot('p-no-rcd-fail', 'LINE-001')
        project = ProjectData.objects.get(proj_id='p-no-rcd-fail')
        project.loop_ln = 999.0
        project.allowablevdrop = 100.0
        project.mcb_curve = 'D'
        project.rcd_provided = False
        project.save()
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-no-rcd-fail')
        branch.cable_length_db_to_jb = 999.0
        tagged_components = branch.tagged_components
        tagged_components['component_details']['MCB']['metadata']['breaker_size'] = 10.0
        branch.tagged_components = tagged_components
        branch.save(update_fields=['cable_length_db_to_jb', 'tagged_components'])

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(result.sizing_status, 'unsizeable')
        self.assertEqual(result.fault_loop_status, 'fail')
        self.assertTrue(any('without RCD' in note for note in result.review_notes))

    def test_size_cold_cable_for_branch_stores_ampacity_result(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p1')
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p1')

        result = size_cold_cable_for_branch(branch, project)

        self.assertEqual(ColdCableResult.objects.count(), 1)
        self.assertEqual(result.branch, branch)
        self.assertEqual(result.sizing_status, 'review_required')
        self.assertEqual(result.cable_4c_size_mm2, 2.5)
        self.assertEqual(result.cable_3c_size_mm2, 2.5)
        self.assertGreater(result.cable_4c_ampacity_derated_a, result.per_circuit_operating_current_a)
        self.assertGreater(result.cable_3c_ampacity_derated_a, result.per_circuit_operating_current_a)
        self.assertEqual(result.cable_4c_conductor_temp_c, 90.0)
        self.assertEqual(result.cable_3c_conductor_temp_c, 90.0)
        self.assertEqual(result.conductor_material_density_kg_m3, 8960.0)
        self.assertIsNotNone(result.cable_4c_conductor_mass_mt)
        self.assertIsNotNone(result.cable_3c_conductor_mass_mt)
        self.assertAlmostEqual(
            result.conductor_mass_total_mt,
            result.cable_4c_conductor_mass_mt + result.cable_3c_conductor_mass_mt,
            places=9,
        )
        self.assertEqual(result.vd_status, 'pass')
        self.assertIsNotNone(result.vd_total_pct)
        self.assertEqual(result.fault_loop_status, 'pass')
        self.assertIsNotNone(result.fault_current_l_pe_a)
        self.assertTrue(result.optimization_run)
        self.assertTrue(any('project default' in note for note in result.review_notes))

    def test_sld_cold_cable_metadata_sizes_each_3c_outgoing_by_own_length(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p-variable-3c', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p-variable-3c')
        project.allowablevdrop = 5.0
        project.save(update_fields=['allowablevdrop'])
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-variable-3c')
        branch.cable_length_db_to_jb = 30.0
        tagged_components = branch.tagged_components
        tagged_components['component_details']['MCB']['metadata']['operating_current'] = 8.0
        tagged_components['component_details']['MCB']['metadata']['per_circuit_operating_current'] = 8.0
        tagged_components['component_details']['MCB']['metadata']['breaker_size'] = 10.0
        branch.tagged_components = tagged_components
        branch.save(update_fields=['cable_length_db_to_jb', 'tagged_components'])
        cable3c_details = [
            downstream['component_details']['Cable3C']
            for downstream in branch.tagged_components['Downstream']
        ]
        for detail, length in zip(cable3c_details, [15.0, 150.0]):
            CableScheduleOverride.objects.create(
                project=project,
                component_id=detail['component_id'],
                component_uid=detail['component_uid'],
                display_tag=detail['display_tag'],
                component_type='Cable3C',
                line_id='LINE-001',
                line_uid=str(branch.distribution.line.uid),
                branch_index=branch.branch_index,
                circuit_index=detail['circuit_index'],
                generated_length_m=12.0,
                manual_length_m=length,
            )

        result = size_cold_cable_for_branch(branch, project)
        payload = build_project_sld_payload('p-variable-3c')
        cable_nodes = sorted(
            [node for node in payload['nodes'] if node['component_type'] == 'Cable3C'],
            key=lambda node: node['metadata']['cold_cable']['length_m'],
        )

        self.assertEqual(len(result.cable_3c_segments), 2)
        self.assertEqual(cable_nodes[0]['metadata']['cold_cable']['length_m'], 15.0)
        self.assertEqual(cable_nodes[1]['metadata']['cold_cable']['length_m'], 150.0)
        self.assertEqual(cable_nodes[0]['metadata']['cold_cable_calculated_size'], '3C x 2.5 mm2')
        self.assertEqual(cable_nodes[1]['metadata']['cold_cable_calculated_size'], '3C x 6 mm2')
        self.assertEqual(result.cable_3c_size_mm2, 6.0)
        self.assertEqual(result.cable_3c_segments[0]['phase_label'], 'L1')
        self.assertEqual(result.cable_3c_segments[1]['phase_label'], 'L2')
        self.assertEqual(cable_nodes[0]['metadata']['cold_cable']['phase_label'], 'L1')
        self.assertEqual(cable_nodes[1]['metadata']['cold_cable']['phase_label'], 'L2')


class SldTopologyWorkflowTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username=f'sld-{uuid4().hex[:8]}',
            email='sld@example.com',
            password='password123',
        )
        self.client.force_login(self.user)

    def _mcb_component_ids(self):
        payload = build_project_sld_payload('p1')
        return [
            node['component_id']
            for node in payload['nodes']
            if node['component_type'] == 'MCB'
        ]

    def _downstream_component_ids(self):
        payload = build_project_sld_payload('p1')
        return [
            node['component_id']
            for node in payload['nodes']
            if node['component_type'] == 'Tracer'
        ]

    def test_combine_feeders_preview_returns_recommended_breaker(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        component_ids = self._mcb_component_ids()

        response = self.client.post(
            reverse('sld_topology_combine_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': component_ids}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        preview = response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['edit_type'], 'combine_feeders')
        self.assertEqual(preview['recommended_breaker_rating'], 10)
        self.assertEqual(len(preview['removed_component_ids']), 1)
        self.assertEqual(preview['added_component_types'], ['Cable4C', 'Isolator3PH', 'JB3PH'])
        self.assertEqual(len(preview['added_display_tags']), 3)

    def test_combine_feeders_preview_uses_starting_current_not_existing_breaker_sum(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        lines = list(HeatTracingInput.objects.filter(proj_id='p1').order_by('line_id'))
        ProcessLineCalculation.objects.filter(line=lines[0]).update(starting_current=0.5)
        ProcessLineCalculation.objects.filter(line=lines[1]).update(starting_current=0.8)
        component_ids = self._mcb_component_ids()

        response = self.client.post(
            reverse('sld_topology_combine_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': component_ids}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        preview = response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['breaker_rating_basis'], 'starting_current')
        self.assertAlmostEqual(preview['combined_feeder_current'], 1.3)
        self.assertEqual(preview['recommended_breaker_rating'], 2)

    def test_combine_feeders_preview_uses_line_current_for_legacy_manual_combines(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003'])
        lines = list(HeatTracingInput.objects.filter(proj_id='p1').order_by('line_id'))
        for line in lines:
            ProcessLineCalculation.objects.filter(line=line).update(starting_current=0.5)
        generated_payload = build_project_sld_payload('p1')
        first_two_mcbs = [
            node['component_id']
            for node in generated_payload['nodes']
            if node['component_type'] == 'MCB'
        ][:2]
        apply_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': first_two_mcbs}),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)

        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        edit_payload = deepcopy(edit.edit_payload)
        for node in edit_payload['sld_payload']['nodes']:
            if node.get('component_type') == 'MCB' and str(node.get('display_tag')).endswith('-M'):
                metadata = dict(node.get('metadata') or {})
                metadata['breaker_size'] = 40
                metadata.pop('starting_current', None)
                metadata.pop('combined_feeder_current', None)
                node['metadata'] = metadata
        edit.edit_payload = edit_payload
        edit.save(update_fields=['edit_payload'])

        active_payload = build_project_sld_payload('p1')
        combined_mcb = next(
            node for node in active_payload['nodes']
            if node['component_type'] == 'MCB' and str(node['display_tag']).endswith('-M')
        )
        remaining_mcb = next(
            node for node in active_payload['nodes']
            if node['component_type'] == 'MCB' and node.get('line_id') == 'LINE-003'
        )
        response = self.client.post(
            reverse('sld_topology_combine_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': [combined_mcb['component_id'], remaining_mcb['component_id']],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        preview = response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['breaker_rating_basis'], 'starting_current')
        self.assertAlmostEqual(preview['combined_feeder_current'], 1.5)
        self.assertEqual(preview['recommended_breaker_rating'], 2)

    def test_combine_feeders_apply_persists_active_topology_edit(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        component_ids = self._mcb_component_ids()

        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': component_ids,
                'trunk_length_m': 42.5,
                'cable_size': '4C x 10',
                'remarks': 'Combined after engineering review.',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(edit.edit_type, 'combine_feeders')
        self.assertIn('engineering review', edit.remarks)
        self.assertEqual(len(edit.baseline_fingerprint), 64)
        self.assertEqual(edit.validation_summary['status'], 'needs_review')
        self.assertTrue(edit.generated_snapshot['nodes'])
        self.assertEqual(edit.edit_payload['downstream_summaries']['boq']['mcb_total'], 1)
        self.assertEqual(edit.edit_payload['downstream_summaries']['result']['branch_count'], 1)
        self.assertEqual(len(edit.edit_payload['cable_schedule_rows']), 1)
        schedule_row = edit.edit_payload['cable_schedule_rows'][0]
        self.assertEqual(schedule_row['tagged_components']['MCB'], 'MCB_001-M')
        self.assertGreater(schedule_row['cable_length_db_to_jb'], 0)
        self.assertGreater(schedule_row['branch_cable_length_total_m'], 0)
        self.assertTrue(schedule_row['tagged_components']['CableRoles']['mcb_trunks'])
        self.assertTrue(schedule_row['tagged_components']['CableRoles']['branch_cables'])

        edited_payload = build_project_sld_payload('p1')
        self.assertTrue(edited_payload['meta']['has_topology_edit'])
        self.assertFalse(edited_payload['meta']['topology_baseline_changed'])
        nodes_by_type = {}
        for node in edited_payload['nodes']:
            nodes_by_type.setdefault(node['component_type'], []).append(node)
        self.assertEqual(len(nodes_by_type['MCB']), 1)
        self.assertTrue(any(
            node['component_type'] == 'MCB' and node['display_tag'].endswith('-M')
            for node in edited_payload['nodes']
        ))
        combined_mcb = nodes_by_type['MCB'][0]
        manual_trunk_cables = [
            node for node in nodes_by_type['Cable4C']
            if (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        ]
        manual_distribution_jbs = [
            node for node in nodes_by_type['JB3PH']
            if (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        ]
        manual_isolators = [
            node for node in nodes_by_type['Isolator3PH']
            if (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        ]
        self.assertEqual(len(manual_trunk_cables), 1)
        self.assertEqual(len(manual_isolators), 1)
        self.assertEqual(len(manual_distribution_jbs), 1)
        trunk_cable = manual_trunk_cables[0]
        isolator = manual_isolators[0]
        distribution_jb = manual_distribution_jbs[0]
        self.assertEqual((trunk_cable.get('metadata') or {}).get('length_m'), 42.5)
        self.assertEqual((trunk_cable.get('metadata') or {}).get('cable_size'), '4C x 10')
        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in edited_payload['edges']
        }
        self.assertIn((combined_mcb['component_id'], trunk_cable['component_id']), edge_pairs)
        self.assertIn((trunk_cable['component_id'], isolator['component_id']), edge_pairs)
        self.assertIn((isolator['component_id'], distribution_jb['component_id']), edge_pairs)
        self.assertEqual(
            sum(
                1
                for node in edited_payload['nodes']
                if (distribution_jb['component_id'], node['component_id']) in edge_pairs
            ),
            2,
        )
        self.assertEqual(
            [
                edge['to_component_id']
                for edge in edited_payload['edges']
                if edge['from_component_id'] == combined_mcb['component_id']
            ],
            [trunk_cable['component_id']],
        )

    def test_single_outgoing_3ph_collapse_does_not_bypass_to_tracer(self):
        from eht.sld_topology_workflows import _collapse_single_outgoing_3ph_jbs

        payload = {
            'project_id': 'p1',
            'nodes': [
                {'component_id': 'mcb', 'component_type': 'MCB', 'display_tag': 'MCB_001', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1},
                {'component_id': 'c4', 'component_type': 'Cable4C', 'display_tag': 'CCAB4C_001', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1, 'metadata': {'length_m': 10}},
                {'component_id': 'jb3', 'component_type': 'JB3PH', 'display_tag': 'JB3PH_001', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1},
                {'component_id': 'tr', 'component_type': 'Tracer', 'display_tag': 'Tracer_001', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1},
            ],
            'edges': [
                {'from_component_id': 'mcb', 'to_component_id': 'c4', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1, 'circuit_index': None},
                {'from_component_id': 'c4', 'to_component_id': 'jb3', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1, 'circuit_index': None},
                {'from_component_id': 'jb3', 'to_component_id': 'tr', 'line_ids': ['L1'], 'line_uid': 'line-1', 'branch_index': 1, 'circuit_index': 1},
            ],
            'line_groups': [],
            'meta': {},
        }

        collapsed = _collapse_single_outgoing_3ph_jbs(payload)

        self.assertEqual({node['component_id'] for node in collapsed['nodes']}, {'mcb', 'c4', 'jb3', 'tr'})
        self.assertNotIn(('mcb', 'tr'), {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in collapsed['edges']
        })

    def test_combine_feeders_can_extend_active_combine_edit(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003'])

        first_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()[:2]}),
            content_type='application/json',
        )
        self.assertEqual(first_response.status_code, 200)

        active_mcb_ids = self._mcb_component_ids()
        preview_response = self.client.post(
            reverse('sld_topology_combine_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': active_mcb_ids}),
            content_type='application/json',
        )

        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertTrue(preview['extends_existing_combine'])
        self.assertEqual(preview['added_component_types'], [])
        self.assertEqual(preview['recommended_breaker_rating'], 16)

        second_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': active_mcb_ids}),
            content_type='application/json',
        )

        self.assertEqual(second_response.status_code, 200)
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='applied').count(), 1)
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='superseded').count(), 1)
        edited_payload = build_project_sld_payload('p1')
        manual_trunks = [
            node for node in edited_payload['nodes']
            if node['component_type'] == 'Cable4C'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        ]
        self.assertEqual(sum(1 for node in edited_payload['nodes'] if node['component_type'] == 'MCB'), 1)
        self.assertEqual(len(manual_trunks), 1)

    def test_downstream_jb_apply_moves_selected_outgoing_branches(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003', 'LINE-004'])
        combine_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(combine_response.status_code, 200)

        combined_payload = build_project_sld_payload('p1')
        parent_jb = next(
            node for node in combined_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        )
        direct_children = [
            edge['to_component_id']
            for edge in combined_payload['edges']
            if edge['from_component_id'] == parent_jb['component_id']
        ]
        self.assertEqual(len(direct_children), 4)

        preview_response = self.client.post(
            reverse('sld_topology_downstream_jb_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'parent_component_id': parent_jb['component_id'],
                'branch_component_ids': direct_children[:3],
                'trunk_length_m': 18.5,
            }),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['parent_outgoing_before'], 4)
        self.assertEqual(preview['parent_outgoing_after'], 2)
        self.assertEqual(preview['downstream_outgoing_count'], 3)

        apply_response = self.client.post(
            reverse('sld_topology_downstream_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'parent_component_id': parent_jb['component_id'],
                'branch_component_ids': direct_children[:3],
                'trunk_length_m': 18.5,
                'cable_size': '4C x 6',
                'remarks': 'Add downstream JB to limit outgoing feeders.',
            }),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(edit.edit_type, 'downstream_jb')
        self.assertIn('limit outgoing', edit.remarks)

        edited_payload = build_project_sld_payload('p1')
        node_by_id = {node['component_id']: node for node in edited_payload['nodes']}
        downstream_jb = next(
            node for node in edited_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'downstream_jb'
        )
        downstream_trunk = next(
            node for node in edited_payload['nodes']
            if node['component_type'] == 'Cable4C'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'downstream_jb'
        )
        downstream_isolator = next(
            node for node in edited_payload['nodes']
            if node['component_type'] == 'Isolator3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'downstream_jb'
        )
        self.assertEqual((downstream_trunk.get('metadata') or {}).get('length_m'), 18.5)
        self.assertEqual((downstream_trunk.get('metadata') or {}).get('cable_size'), '4C x 6')
        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in edited_payload['edges']
        }
        self.assertEqual(downstream_trunk['metadata']['length_m'], 18.5)
        self.assertIn((parent_jb['component_id'], downstream_trunk['component_id']), edge_pairs)
        self.assertIn((downstream_trunk['component_id'], downstream_isolator['component_id']), edge_pairs)
        self.assertIn((downstream_isolator['component_id'], downstream_jb['component_id']), edge_pairs)
        self.assertEqual(
            len([edge for edge in edited_payload['edges'] if edge['from_component_id'] == parent_jb['component_id']]),
            2,
        )
        self.assertEqual(
            len([edge for edge in edited_payload['edges'] if edge['from_component_id'] == downstream_jb['component_id']]),
            3,
        )
        self.assertTrue(all(component_id in node_by_id for component_id in direct_children[:3]))

    def test_attach_to_jb_moves_standalone_mcb_feeder_to_target_jb(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003'])
        first_two_mcb_ids = self._mcb_component_ids()[:2]
        combine_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': first_two_mcb_ids}),
            content_type='application/json',
        )
        self.assertEqual(combine_response.status_code, 200)

        active_payload = build_project_sld_payload('p1')
        target_jb = next(
            node for node in active_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        )
        source_mcb = next(
            node for node in active_payload['nodes']
            if node['component_type'] == 'MCB' and node.get('line_id') == 'LINE-003'
        )

        preview_response = self.client.post(
            reverse('sld_topology_attach_jb_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': source_mcb['component_id'],
                'target_jb_component_id': target_jb['component_id'],
            }),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['edit_type'], 'attach_to_jb')
        self.assertEqual(preview['target_outgoing_before'], 2)
        self.assertEqual(preview['target_outgoing_after'], 3)
        self.assertEqual(preview['recommended_breaker_rating'], 16)

        apply_response = self.client.post(
            reverse('sld_topology_attach_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': source_mcb['component_id'],
                'target_jb_component_id': target_jb['component_id'],
                'remarks': 'Attach line 3 to spare outgoing on combined JB.',
            }),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(edit.edit_type, 'attach_to_jb')
        self.assertIn('spare outgoing', edit.remarks)

        edited_payload = build_project_sld_payload('p1')
        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in edited_payload['edges']
        }
        self.assertFalse(any(node['component_id'] == source_mcb['component_id'] for node in edited_payload['nodes']))
        self.assertEqual(sum(1 for node in edited_payload['nodes'] if node['component_type'] == 'MCB'), 1)
        moved_entry_id = preview['moved_component_ids'][0]
        self.assertIn((target_jb['component_id'], moved_entry_id), edge_pairs)
        combined_mcb = next(node for node in edited_payload['nodes'] if node['component_type'] == 'MCB')
        self.assertEqual((combined_mcb.get('metadata') or {}).get('breaker_size'), 16)

    def test_attach_to_jb_moves_downstream_branch_between_jbs(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003', 'LINE-004'])
        combine_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(combine_response.status_code, 200)

        combined_payload = build_project_sld_payload('p1')
        upstream_jb = next(
            node for node in combined_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        )
        direct_children = [
            edge['to_component_id']
            for edge in combined_payload['edges']
            if edge['from_component_id'] == upstream_jb['component_id']
        ]
        downstream_response = self.client.post(
            reverse('sld_topology_downstream_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'parent_component_id': upstream_jb['component_id'],
                'branch_component_ids': direct_children[:3],
                'trunk_length_m': 12,
            }),
            content_type='application/json',
        )
        self.assertEqual(downstream_response.status_code, 200)

        moved_payload = build_project_sld_payload('p1')
        downstream_jb = next(
            node for node in moved_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'downstream_jb'
        )
        branch_root_id = next(
            edge['to_component_id']
            for edge in moved_payload['edges']
            if edge['from_component_id'] == downstream_jb['component_id']
        )

        preview_response = self.client.post(
            reverse('sld_topology_attach_jb_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': branch_root_id,
                'target_jb_component_id': upstream_jb['component_id'],
            }),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['edit_type'], 'move_branch_to_jb')
        self.assertEqual(preview['source_outgoing_before'], 3)
        self.assertEqual(preview['source_outgoing_after'], 2)
        self.assertEqual(preview['target_outgoing_before'], 2)
        self.assertEqual(preview['target_outgoing_after'], 3)

        apply_response = self.client.post(
            reverse('sld_topology_attach_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': branch_root_id,
                'target_jb_component_id': upstream_jb['component_id'],
                'remarks': 'Move branch to spare outgoing on upstream JB.',
            }),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(edit.edit_type, 'move_branch_to_jb')
        edited_payload = build_project_sld_payload('p1')
        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in edited_payload['edges']
        }
        self.assertIn((upstream_jb['component_id'], branch_root_id), edge_pairs)
        self.assertNotIn((downstream_jb['component_id'], branch_root_id), edge_pairs)

    def test_attach_to_jb_moves_downstream_branch_between_different_mcb_trees(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        payload = build_project_sld_payload('p1')
        source_jb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'JB3PH' and node.get('line_id') == 'LINE-001'
        )
        target_jb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'JB3PH' and node.get('line_id') == 'LINE-002'
        )
        branch_root_id = next(
            edge['to_component_id']
            for edge in payload['edges']
            if edge['from_component_id'] == source_jb['component_id']
        )
        remaining_source_child_id = next(
            edge['to_component_id']
            for edge in payload['edges']
            if edge['from_component_id'] == source_jb['component_id']
            and edge['to_component_id'] != branch_root_id
        )

        preview_response = self.client.post(
            reverse('sld_topology_attach_jb_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': branch_root_id,
                'target_jb_component_id': target_jb['component_id'],
            }),
            content_type='application/json',
        )

        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['edit_type'], 'move_branch_to_jb')
        self.assertTrue(preview['cross_mcb_move'])
        self.assertEqual(preview['target_outgoing_before'], 2)
        self.assertEqual(preview['target_outgoing_after'], 3)
        self.assertEqual(preview['source_breaker_rating'], 6)
        self.assertEqual(preview['target_breaker_rating'], 6)
        self.assertEqual(preview['recommended_source_breaker_rating'], 4)
        self.assertEqual(preview['recommended_target_breaker_rating'], 10)

        apply_response = self.client.post(
            reverse('sld_topology_attach_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': branch_root_id,
                'target_jb_component_id': target_jb['component_id'],
                'remarks': 'Move one branch to another feeder tree.',
            }),
            content_type='application/json',
        )

        self.assertEqual(apply_response.status_code, 200)
        edited_payload = build_project_sld_payload('p1')
        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in edited_payload['edges']
        }
        self.assertIn((target_jb['component_id'], branch_root_id), edge_pairs)
        self.assertNotIn((source_jb['component_id'], branch_root_id), edge_pairs)
        self.assertFalse(any(node['component_id'] == source_jb['component_id'] for node in edited_payload['nodes']))
        source_mcb = next(
            node for node in edited_payload['nodes']
            if node['component_id'] == preview['upstream_mcb_component_id']
        )
        target_mcb = next(
            node for node in edited_payload['nodes']
            if node['component_id'] == preview['target_mcb_component_id']
        )
        self.assertEqual((source_mcb.get('metadata') or {}).get('breaker_size'), 4)
        self.assertEqual((target_mcb.get('metadata') or {}).get('breaker_size'), 10)
        self.assertIn((source_mcb['component_id'], remaining_source_child_id), edge_pairs)
        self.assertEqual((source_mcb.get('metadata') or {}).get('previous_breaker_size'), 6)
        self.assertEqual((target_mcb.get('metadata') or {}).get('previous_breaker_size'), 6)
        self.assertEqual((source_mcb.get('metadata') or {}).get('recommended_breaker_size'), 4)
        self.assertEqual((target_mcb.get('metadata') or {}).get('recommended_breaker_size'), 10)

    def test_attach_to_jb_can_promote_standalone_mcb_target_to_3ph_distribution(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        payload = build_project_sld_payload('p1')
        edges_by_source = {}
        for edge in payload['edges']:
            edges_by_source.setdefault(edge['from_component_id'], []).append(edge)

        source_jb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'JB3PH' and node.get('line_id') == 'LINE-001'
        )
        source_branch_root_id = edges_by_source[source_jb['component_id']][0]['to_component_id']
        target_mcb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'MCB' and node.get('line_id') == 'LINE-002'
        )
        target_jb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'JB3PH' and node.get('line_id') == 'LINE-002'
        )
        target_existing_child_id = edges_by_source[target_jb['component_id']][0]['to_component_id']
        disconnected_child_id = edges_by_source[target_jb['component_id']][1]['to_component_id']

        remove_ids = {
            node['component_id']
            for node in payload['nodes']
            if node.get('line_id') == 'LINE-002'
            and node['component_type'] in {'Cable4C', 'Isolator3PH', 'JB3PH'}
        }
        stack = [disconnected_child_id]
        while stack:
            component_id = stack.pop()
            if component_id in remove_ids:
                continue
            remove_ids.add(component_id)
            stack.extend(edge['to_component_id'] for edge in edges_by_source.get(component_id, []))

        edited_payload = deepcopy(payload)
        edited_payload['nodes'] = [
            node for node in edited_payload['nodes']
            if node['component_id'] not in remove_ids
        ]
        edited_payload['edges'] = [
            edge for edge in edited_payload['edges']
            if edge['from_component_id'] not in remove_ids
            and edge['to_component_id'] not in remove_ids
        ]
        edited_payload['edges'].append({
            'from_component_id': target_mcb['component_id'],
            'to_component_id': target_existing_child_id,
            'line_ids': ['LINE-002'],
            'line_uid': target_mcb['line_uid'],
            'branch_index': target_mcb['branch_index'],
            'circuit_index': 1,
        })
        edited_payload['meta']['node_count'] = len(edited_payload['nodes'])
        edited_payload['meta']['edge_count'] = len(edited_payload['edges'])
        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='move_branch_to_jb',
            status='applied',
            generated_snapshot=payload,
            baseline_fingerprint=payload_fingerprint(payload),
            edit_payload={'sld_payload': edited_payload},
        )

        preview_response = self.client.post(
            reverse('sld_topology_attach_jb_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': source_branch_root_id,
                'target_jb_component_id': target_mcb['component_id'],
            }),
            content_type='application/json',
        )

        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertTrue(preview['insert_target_distribution_jb'])
        self.assertEqual(preview['target_component_type'], 'MCB')
        self.assertEqual(preview['target_outgoing_before'], 1)
        self.assertEqual(preview['target_outgoing_after'], 2)
        self.assertEqual(preview['source_breaker_rating'], 6)
        self.assertEqual(preview['target_breaker_rating'], 6)
        self.assertEqual(preview['recommended_source_breaker_rating'], 4)
        self.assertEqual(preview['recommended_target_breaker_rating'], 10)

        apply_response = self.client.post(
            reverse('sld_topology_attach_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': source_branch_root_id,
                'target_jb_component_id': target_mcb['component_id'],
                'trunk_length_m': 77.5,
                'cable_size': '4C x 16',
                'remarks': 'Promote target MCB to 3PH distribution.',
            }),
            content_type='application/json',
        )

        self.assertEqual(apply_response.status_code, 200)
        final_payload = build_project_sld_payload('p1')
        edge_pairs = {
            (edge['from_component_id'], edge['to_component_id'])
            for edge in final_payload['edges']
        }
        manual_jb = next(
            node for node in final_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('target_mcb_distribution')
        )
        manual_cable = next(
            node for node in final_payload['nodes']
            if node['component_type'] == 'Cable4C'
            and (node.get('metadata') or {}).get('target_mcb_distribution')
        )
        manual_isolator = next(
            node for node in final_payload['nodes']
            if node['component_type'] == 'Isolator3PH'
            and (node.get('metadata') or {}).get('target_mcb_distribution')
        )
        self.assertEqual((manual_cable.get('metadata') or {}).get('length_m'), 77.5)
        self.assertEqual((manual_cable.get('metadata') or {}).get('cable_size'), '4C x 16')
        self.assertIn((target_mcb['component_id'], manual_cable['component_id']), edge_pairs)
        self.assertIn((manual_cable['component_id'], manual_isolator['component_id']), edge_pairs)
        self.assertIn((manual_isolator['component_id'], manual_jb['component_id']), edge_pairs)
        self.assertIn((manual_jb['component_id'], target_existing_child_id), edge_pairs)
        self.assertIn((manual_jb['component_id'], source_branch_root_id), edge_pairs)

    def test_combine_feeders_preview_requires_two_mcbs(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        component_ids = self._mcb_component_ids()

        response = self.client.post(
            reverse('sld_topology_combine_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': component_ids}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('Select at least two', response.json()['error'])

    def test_split_circuits_preview_returns_recommended_breaker(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        component_ids = self._mcb_component_ids()[:1]

        response = self.client.post(
            reverse('sld_topology_split_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': component_ids}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        preview = response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['edit_type'], 'split_circuits')
        self.assertEqual(preview['recommended_breaker_rating'], 4)
        self.assertEqual(preview['selected_circuit_count'], 2)
        self.assertEqual(preview['new_mcb_count'], 1)
        self.assertIn('JB3PH_001', preview['removed_display_tags'])

    def test_split_circuits_apply_persists_active_topology_edit(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        component_ids = self._mcb_component_ids()[:1]

        response = self.client.post(
            reverse('sld_topology_split_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': component_ids,
                'remarks': 'Split after circuit review.',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(edit.edit_type, 'split_circuits')
        self.assertIn('circuit review', edit.remarks)
        self.assertEqual(len(edit.baseline_fingerprint), 64)
        self.assertEqual(edit.validation_summary['status'], 'needs_review')
        self.assertEqual(edit.edit_payload['downstream_summaries']['boq']['mcb_total'], 2)
        self.assertEqual(edit.edit_payload['downstream_summaries']['boq']['junction_box_total'], 2)
        self.assertEqual(edit.edit_payload['downstream_summaries']['result']['branch_count'], 2)
        self.assertEqual(
            [row['distribution']['line']['line_id'] for row in edit.edit_payload['cable_schedule_rows']],
            ['LINE-001-part1', 'LINE-001-part2'],
        )

        edited_payload = build_project_sld_payload('p1')
        self.assertTrue(edited_payload['meta']['has_topology_edit'])
        self.assertEqual(
            sum(1 for node in edited_payload['nodes'] if node['component_type'] == 'MCB'),
            2,
        )
        self.assertTrue(any(
            node['component_type'] == 'MCB' and node['display_tag'].endswith('-S1')
            for node in edited_payload['nodes']
        ))
        self.assertTrue(any(
            node['component_type'] == 'MCB' and node['display_tag'].endswith('-S2')
            for node in edited_payload['nodes']
        ))
        self.assertEqual(
            [group['line_id'] for group in edited_payload['line_groups']],
            ['LINE-001-part1', 'LINE-001-part2'],
        )
        self.assertFalse(any(node['component_type'] == 'JB3PH' for node in edited_payload['nodes']))
        self.assertFalse(any(node['component_type'] == 'Cable4C' for node in edited_payload['nodes']))
        split_tracers = [node for node in edited_payload['nodes'] if node['component_type'] == 'Tracer']
        self.assertTrue(all((node.get('metadata') or {}).get('tracer_selection') for node in split_tracers))
        self.assertEqual(
            {node['metadata']['tracer_selection']['selected']['v_uid'] for node in split_tracers},
            {'V-001'},
        )

        incoming_sources_by_target = {
            edge['to_component_id']: edge['from_component_id']
            for edge in edited_payload['edges']
        }
        mcb_by_line = {
            node['line_id']: node['component_id']
            for node in edited_payload['nodes']
            if node['component_type'] == 'MCB'
        }
        for line_id, mcb_id in mcb_by_line.items():
            first_load = next(
                node for node in edited_payload['nodes']
                if node['line_id'] == line_id and node['component_type'] == 'Isolator1PH'
            )
            self.assertEqual(incoming_sources_by_target[first_load['component_id']], mcb_id)

        filtered_payload = build_project_sld_payload('p1', line_id='LINE-001')
        self.assertEqual(
            [group['line_id'] for group in filtered_payload['line_groups']],
            ['LINE-001-part1', 'LINE-001-part2'],
        )

    def test_scoped_reset_restores_split_circuit_without_full_reset(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        component_ids = self._mcb_component_ids()[:1]

        split_response = self.client.post(
            reverse('sld_topology_split_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': component_ids}),
            content_type='application/json',
        )
        self.assertEqual(split_response.status_code, 200)

        split_payload = build_project_sld_payload('p1')
        split_mcb = next(
            node for node in split_payload['nodes']
            if node['component_type'] == 'MCB'
            and node['display_tag'].endswith('-S2')
        )
        reset_response = self.client.post(
            reverse('sld_topology_reset_selected_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_id': split_mcb['component_id'],
                'remarks': 'Reset split only.',
            }),
            content_type='application/json',
        )

        self.assertEqual(reset_response.status_code, 200)
        self.assertTrue(reset_response.json()['ok'])
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='applied').count(), 1)
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='superseded').count(), 1)
        edited_payload = build_project_sld_payload('p1')
        self.assertEqual(
            [group['line_id'] for group in edited_payload['line_groups']],
            ['LINE-001'],
        )
        self.assertEqual(
            [node['display_tag'] for node in edited_payload['nodes'] if node['component_type'] == 'MCB'],
            ['MCB_001'],
        )
        self.assertTrue(any(node['component_type'] == 'JB3PH' for node in edited_payload['nodes']))
        self.assertFalse(any(
            (node.get('metadata') or {}).get('manual_topology_edit') == 'split_circuits'
            for node in edited_payload['nodes']
        ))

    def test_split_circuits_can_split_active_manual_combine_topology(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        combine_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(combine_response.status_code, 200)

        active_mcb_ids = self._mcb_component_ids()
        self.assertEqual(len(active_mcb_ids), 1)
        preview_response = self.client.post(
            reverse('sld_topology_split_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': active_mcb_ids}),
            content_type='application/json',
        )

        self.assertEqual(preview_response.status_code, 200)
        preview = preview_response.json()
        self.assertTrue(preview['ok'])
        self.assertEqual(preview['selected_circuit_count'], 2)
        self.assertTrue(any(tag.startswith('JB3PH_001') for tag in preview['removed_display_tags']))

        apply_response = self.client.post(
            reverse('sld_topology_split_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': active_mcb_ids,
                'remarks': 'Undo combined feeder after review.',
            }),
            content_type='application/json',
        )

        self.assertEqual(apply_response.status_code, 200)
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='applied').count(), 1)
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='superseded').count(), 1)
        applied_edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        edited_payload = build_project_sld_payload('p1')
        self.assertEqual(
            sorted(group['line_id'] for group in edited_payload['line_groups']),
            ['LINE-001', 'LINE-002'],
        )
        self.assertEqual(sum(1 for node in edited_payload['nodes'] if node['component_type'] == 'MCB'), 2)
        self.assertFalse(any(
            node['component_type'] in {'Cable4C', 'Isolator3PH', 'JB3PH'}
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
            for node in edited_payload['nodes']
        ))
        self.assertEqual(
            [row['distribution']['line']['line_id'] for row in applied_edit.edit_payload['cable_schedule_rows']],
            ['LINE-001', 'LINE-002'],
        )

    def test_split_circuits_after_branch_move_avoids_line_identity_collision(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        payload = build_project_sld_payload('p1')
        source_jb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'JB3PH' and node.get('line_id') == 'LINE-001'
        )
        target_jb = next(
            node for node in payload['nodes']
            if node['component_type'] == 'JB3PH' and node.get('line_id') == 'LINE-002'
        )
        source_branch_root_id = next(
            edge['to_component_id']
            for edge in payload['edges']
            if edge['from_component_id'] == source_jb['component_id']
        )
        target_mcb_id = next(
            node['component_id']
            for node in payload['nodes']
            if node['component_type'] == 'MCB' and node.get('line_id') == 'LINE-002'
        )

        move_response = self.client.post(
            reverse('sld_topology_attach_jb_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'source_component_id': source_branch_root_id,
                'target_jb_component_id': target_jb['component_id'],
            }),
            content_type='application/json',
        )
        self.assertEqual(move_response.status_code, 200)

        split_response = self.client.post(
            reverse('sld_topology_split_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': [target_mcb_id]}),
            content_type='application/json',
        )

        self.assertEqual(split_response.status_code, 200)
        edited_payload = build_project_sld_payload('p1')
        line_ids = sorted(group['line_id'] for group in edited_payload['line_groups'])
        self.assertIn('LINE-001', line_ids)
        self.assertIn('LINE-001-part1', line_ids)
        self.assertIn('LINE-002-part1', line_ids)
        self.assertIn('LINE-002-part2', line_ids)
        self.assertEqual(
            sum(1 for node in edited_payload['nodes'] if node['component_type'] == 'MCB'),
            4,
        )
        self.assertFalse(any(
            node.get('component_type') == 'JB3PH'
            and node.get('line_id') == 'LINE-002'
            for node in edited_payload['nodes']
        ))

    def test_scoped_reset_restores_selected_tree_without_resetting_other_manual_edit(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003', 'LINE-004'])
        initial_mcb_ids = self._mcb_component_ids()
        first_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': initial_mcb_ids[:2]}),
            content_type='application/json',
        )
        self.assertEqual(first_response.status_code, 200)

        active_payload = build_project_sld_payload('p1')
        second_pair_ids = [
            node['component_id']
            for node in active_payload['nodes']
            if node['component_type'] == 'MCB' and node.get('line_id') in {'LINE-003', 'LINE-004'}
        ]
        second_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': second_pair_ids}),
            content_type='application/json',
        )
        self.assertEqual(second_response.status_code, 200)

        combined_payload = build_project_sld_payload('p1')
        first_combined_component = next(
            node for node in combined_payload['nodes']
            if node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
            and {'LINE-001', 'LINE-002'}.issubset(set(node.get('line_ids') or []))
        )
        reset_response = self.client.post(
            reverse('sld_topology_reset_selected_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_id': first_combined_component['component_id'],
                'remarks': 'Reset first reviewed feeder only.',
            }),
            content_type='application/json',
        )

        self.assertEqual(reset_response.status_code, 200)
        self.assertTrue(reset_response.json()['ok'])
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='applied').count(), 1)
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='superseded').count(), 2)
        edited_payload = build_project_sld_payload('p1')
        mcb_line_sets = [
            set(node.get('line_ids') or [node.get('line_id')])
            for node in edited_payload['nodes']
            if node['component_type'] == 'MCB'
        ]
        self.assertIn({'LINE-001'}, mcb_line_sets)
        self.assertIn({'LINE-002'}, mcb_line_sets)
        self.assertTrue(any({'LINE-003', 'LINE-004'}.issubset(line_set) for line_set in mcb_line_sets))
        self.assertTrue(any(
            node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
            and {'LINE-003', 'LINE-004'}.issubset(set(node.get('line_ids') or []))
            for node in edited_payload['nodes']
        ))
        self.assertFalse(any(
            node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
            and {'LINE-001', 'LINE-002'}.issubset(set(node.get('line_ids') or []))
            for node in edited_payload['nodes']
        ))

    def test_split_circuits_preview_rejects_downstream_selection(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        component_ids = self._downstream_component_ids()[:1]

        response = self.client.post(
            reverse('sld_topology_split_preview_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': component_ids}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('MCB feeder source', response.json()['error'])

    def test_topology_edit_audit_trail_records_authenticated_user(self):
        user = User.objects.create_user(username='reviewer', password='password123')
        self.client.force_login(user)
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        ManagedProject.objects.get(proj_id='p1').assigned_users.add(user)
        component_ids = self._mcb_component_ids()

        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': component_ids,
                'remarks': 'Approved by electrical lead.',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(edit.created_by, user)
        self.assertIn('electrical lead', edit.remarks)
        self.assertEqual(edit.generated_snapshot['schema_version'], SLD_GRAPH_SCHEMA_VERSION)
        self.assertIn('sld_payload', edit.edit_payload)
        self.assertEqual(edit.edit_payload['topology_operations'][0]['schema_version'], 1)
        self.assertEqual(edit.edit_payload['topology_operations'][0]['operation_type'], 'combine_feeders')
        self.assertEqual(
            edit.edit_payload['topology_operations'][0]['inputs']['component_ids'],
            component_ids,
        )
        self.assertIn('warnings', edit.validation_summary)

    def test_applied_topology_edit_feeds_result_cable_schedule_rows(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        result_response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(result_response.status_code, 200)
        self.assertContains(result_response, '1 branch row')
        self.assertContains(result_response, 'MCB_001-M')
        self.assertContains(result_response, 'LINE-001, LINE-002')

    def test_applied_topology_edit_reports_recalculated_baseline(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        edit.baseline_fingerprint = 'stale-baseline'
        edit.save(update_fields=['baseline_fingerprint'])

        payload = build_project_sld_payload('p1')
        self.assertTrue(payload['meta']['topology_baseline_changed'])
        self.assertFalse(payload['meta']['topology_edit_review_required'])
        self.assertTrue(payload['meta']['topology_edit_replayed_on_current_baseline'])
        self.assertIn('replayed from audited operation records', payload['meta']['manual_topology_warning'])
        self.assertEqual(sum(1 for node in payload['nodes'] if node['component_type'] == 'MCB'), 1)
        self.assertTrue(any(
            node['component_type'] == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
            for node in payload['nodes']
        ))

        result_response = self.client.get(reverse('result_view'), {'project_id': 'p1'})
        self.assertEqual(result_response.status_code, 200)
        self.assertContains(result_response, '1 branch row')
        self.assertContains(result_response, 'MCB_001-M')

    def test_replay_failure_after_baseline_change_requires_review(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        edit.baseline_fingerprint = 'stale-baseline'
        edit.edit_payload['topology_operations'][0]['inputs']['component_ids'] = [
            'missing-mcb',
            *edit.edit_payload['topology_operations'][0]['inputs']['component_ids'][1:],
        ]
        edit.save(update_fields=['baseline_fingerprint', 'edit_payload'])

        payload = build_project_sld_payload('p1')

        self.assertTrue(payload['meta']['topology_baseline_changed'])
        self.assertTrue(payload['meta']['topology_edit_review_required'])
        self.assertIn('requires review', payload['meta']['manual_topology_warning'])
        self.assertEqual(sum(1 for node in payload['nodes'] if node['component_type'] == 'MCB'), 2)

    def test_new_topology_edit_drops_unreplayable_active_operation_chain(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003'])
        first_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()[:2]}),
            content_type='application/json',
        )
        self.assertEqual(first_response.status_code, 200)

        stale_edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        stale_edit.baseline_fingerprint = 'stale-baseline'
        stale_edit.edit_payload['topology_operations'][0]['inputs']['component_ids'] = [
            'missing-mcb',
            *stale_edit.edit_payload['topology_operations'][0]['inputs']['component_ids'][1:],
        ]
        stale_edit.save(update_fields=['baseline_fingerprint', 'edit_payload'])

        stale_payload = build_project_sld_payload('p1')
        self.assertTrue(stale_payload['meta']['topology_edit_review_required'])

        second_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()[:2]}),
            content_type='application/json',
        )

        self.assertEqual(second_response.status_code, 200)
        active_edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        self.assertEqual(active_edit.edit_type, 'combine_feeders')
        self.assertEqual(len(active_edit.edit_payload['topology_operations']), 1)
        self.assertNotIn(
            'missing-mcb',
            active_edit.edit_payload['topology_operations'][0]['inputs']['component_ids'],
        )
        repaired_payload = build_project_sld_payload('p1')
        self.assertTrue(repaired_payload['meta']['has_topology_edit'])
        self.assertFalse(repaired_payload['meta'].get('topology_edit_review_required', False))
        self.assertEqual(SLDTopologyEdit.objects.filter(project_id='p1', status='superseded').count(), 1)
        self.assertEqual(
            active_edit.edit_payload['topology_chain_audit']['inheritance']['reason'],
            'active_chain_failed_replay',
        )

    def test_topology_mutation_is_rejected_from_filtered_sld_view(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        response = self.client.post(
            reverse('sld_topology_combine_preview_view'),
            data=json.dumps({
                'project_id': 'p1',
                'line_id': 'LINE-001',
                'component_ids': self._mcb_component_ids()[:2],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('filtered SLD view', response.json()['error'])
        self.assertFalse(SLDTopologyEdit.objects.filter(project_id='p1').exists())

    def test_saved_topology_invariant_failure_falls_back_to_generated_payload(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        edited = deepcopy(edit.edit_payload)
        jb_node = next(
            node for node in edited['sld_payload']['nodes']
            if node.get('component_type') == 'JB3PH'
            and (node.get('metadata') or {}).get('manual_topology_edit') == 'combine_feeders'
        )
        target_id = next(
            edge['to_component_id']
            for edge in edited['sld_payload']['edges']
            if edge['from_component_id'] == jb_node['component_id']
        )
        edited['sld_payload']['edges'].append({
            'from_component_id': target_id,
            'to_component_id': jb_node['component_id'],
            'line_uid': 'cycle',
            'branch_index': 99,
            'circuit_index': 1,
        })
        edit.edit_payload = edited
        edit.save(update_fields=['edit_payload'])

        payload = build_project_sld_payload('p1')

        self.assertTrue(payload['meta']['topology_edit_review_required'])
        self.assertIn('violates topology guard rails', payload['meta']['manual_topology_warning'])

    def test_compacted_topology_chain_requires_review_after_baseline_change(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        edit.edit_payload['topology_chain_compacted'] = True
        edit.baseline_fingerprint = 'stale-baseline'
        edit.save(update_fields=['baseline_fingerprint', 'edit_payload'])

        payload = build_project_sld_payload('p1')

        self.assertTrue(payload['meta']['topology_edit_review_required'])
        self.assertIn('operation chain was compacted', payload['meta']['manual_topology_warning'])

    def test_long_topology_operation_chain_compaction_keeps_tail_and_audit(self):
        from eht.sld_topology_workflows import (
            SLD_OPERATION_CHAIN_COMPACT_THRESHOLD,
            SLD_OPERATION_CHAIN_KEEP_COUNT,
            _compact_topology_operation_chain,
        )

        operations = [
            {
                'schema_version': 1,
                'operation_type': 'split_circuits',
                'inputs': {'component_ids': [f'mcb-{index}']},
            }
            for index in range(SLD_OPERATION_CHAIN_COMPACT_THRESHOLD + 3)
        ]

        compacted, audit = _compact_topology_operation_chain(operations)

        self.assertEqual(len(compacted), SLD_OPERATION_CHAIN_KEEP_COUNT)
        self.assertEqual(compacted[0]['inputs']['component_ids'], ['mcb-31'])
        self.assertTrue(audit['operation_chain_compacted'])
        self.assertEqual(audit['reason'], 'operation_chain_length_threshold')

    def test_invalid_saved_topology_payload_falls_back_to_generated_payload(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        generated_payload = build_project_sld_payload('p1')
        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'sld_payload': {
                    **generated_payload,
                    'edges': [{
                        'from_component_id': 'missing-source',
                        'to_component_id': generated_payload['nodes'][0]['component_id'],
                    }],
                },
            },
            validation_summary={'status': 'needs_review'},
        )

        payload = build_project_sld_payload('p1')

        self.assertTrue(payload['meta']['topology_edit_review_required'])
        self.assertFalse(payload['meta']['topology_baseline_changed'])
        self.assertIn('saved graph references are invalid', payload['meta']['manual_topology_warning'])
        self.assertNotIn('missing-source', {
            edge.get('from_component_id')
            for edge in payload['edges']
        })

    def test_invalid_topology_operation_records_fall_back_to_generated_payload(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        generated_payload = build_project_sld_payload('p1')
        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'sld_payload': generated_payload,
                'topology_operations': {'operation_type': 'combine_feeders'},
            },
            validation_summary={'status': 'needs_review'},
        )

        payload = build_project_sld_payload('p1')

        self.assertTrue(payload['meta']['topology_edit_review_required'])
        self.assertFalse(payload['meta']['topology_baseline_changed'])
        self.assertIn('operation records are invalid', payload['meta']['manual_topology_warning'])

    def test_topology_reset_restores_generated_payload(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        apply_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': self._mcb_component_ids()}),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)
        self.assertTrue(SLDTopologyEdit.objects.filter(project_id='p1', status='applied').exists())

        response = self.client.post(
            reverse('sld_topology_reset_view'),
            data=json.dumps({'project_id': 'p1'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['reset_count'], 1)
        self.assertFalse(SLDTopologyEdit.objects.filter(project_id='p1', status='applied').exists())
        self.assertTrue(SLDTopologyEdit.objects.filter(project_id='p1', status='reset').exists())
        payload = build_project_sld_payload('p1')
        self.assertFalse(payload['meta'].get('has_topology_edit', False))
        self.assertEqual(sum(1 for node in payload['nodes'] if node['component_type'] == 'MCB'), 2)


class SldTopologyAdminTests(TestCase):
    def test_sld_topology_edit_admin_shows_readonly_history_and_replay_diagnostic(self):
        import eht.admin  # noqa: F401 - ensure admin registrations are loaded.
        from eht.sld_topology_workflows import apply_combine_feeders

        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        payload = build_project_sld_payload('p1')
        mcb_ids = [
            node['component_id']
            for node in payload['nodes']
            if node['component_type'] == 'MCB'
        ]

        result = apply_combine_feeders('p1', mcb_ids)

        self.assertTrue(result['ok'])
        edit = SLDTopologyEdit.objects.get(project_id='p1', status='applied')
        model_admin = admin.site._registry[SLDTopologyEdit]
        self.assertEqual(model_admin.operation_count(edit), 1)
        self.assertIn('combine_feeders', str(model_admin.operation_history(edit)))
        self.assertIn('&quot;ok&quot;: true', str(model_admin.replay_diagnostic(edit)))

        user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='password123',
        )
        self.client.force_login(user)
        response = self.client.get(reverse('admin:eht_sldtopologyedit_change', args=[edit.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Operation history')
        self.assertContains(response, 'combine_feeders')
        self.assertContains(response, 'Replay diagnostic')

    def test_sld_topology_edit_admin_can_compact_payload_and_delete_only_non_active_history(self):
        import eht.admin  # noqa: F401 - ensure admin registrations are loaded.

        project = make_project_record(proj_id=f'padmin-{uuid4().hex[:6]}')
        active = SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status='applied',
            generated_snapshot={'nodes': [{'component_id': 'active'}]},
            edit_payload={'sld_payload': {'nodes': [{'component_id': 'active'}]}},
            validation_summary={'status': 'needs_review'},
        )
        old_history = SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status='superseded',
            generated_snapshot={'nodes': [{'component_id': 'old'}]},
            edit_payload={
                'sld_payload': {'nodes': [{'component_id': 'old'}]},
                'topology_operations': [{
                    'schema_version': 1,
                    'operation_type': 'combine_feeders',
                    'inputs': {'component_ids': ['MCB-1', 'MCB-2']},
                    'preview': {'warning': 'Review required.'},
                }],
                'cable_schedule_rows': [{'tag': 'CBL-1'}],
            },
            validation_summary={'status': 'needs_review'},
        )
        needs_review = SLDTopologyEdit.objects.create(
            project=project,
            edit_type='split_circuits',
            status='needs_review',
            generated_snapshot={'nodes': [{'component_id': 'review'}]},
            edit_payload={'sld_payload': {'nodes': [{'component_id': 'review'}]}},
            validation_summary={'status': 'needs_review'},
        )
        model_admin = admin.site._registry[SLDTopologyEdit]

        class DummyUser:
            def has_perm(self, _permission):
                return True

        request = type('Request', (), {'user': DummyUser()})()
        messages_seen = []
        original_message_user = model_admin.message_user
        model_admin.message_user = lambda _request, message, level=None: messages_seen.append((message, level))
        try:
            model_admin.compact_selected_history_records(
                request,
                SLDTopologyEdit.objects.filter(pk=old_history.pk),
            )
            old_history.refresh_from_db()
            self.assertTrue(old_history.edit_payload['history_payload_compacted'])
            self.assertFalse(old_history.generated_snapshot)
            self.assertIn('topology_operation_audit_summary', old_history.edit_payload)
            self.assertIn('audit_only', str(model_admin.operation_history(old_history)))

            model_admin.emergency_delete_selected_non_active_history_records(
                request,
                SLDTopologyEdit.objects.filter(pk__in=[active.pk, old_history.pk, needs_review.pk]),
            )
        finally:
            model_admin.message_user = original_message_user

        self.assertFalse(SLDTopologyEdit.objects.filter(pk=old_history.pk).exists())
        self.assertTrue(SLDTopologyEdit.objects.filter(pk=active.pk).exists())
        self.assertTrue(SLDTopologyEdit.objects.filter(pk=needs_review.pk).exists())
        self.assertTrue(any('Skipped 2 protected' in message for message, _level in messages_seen))


class SldTopologyHistoryRetentionTests(TestCase):
    def _project_id(self, prefix):
        return f'{prefix}-{uuid4().hex[:6]}'

    def _create_history_row(self, project, index, *, status='superseded'):
        return SLDTopologyEdit.objects.create(
            project=project,
            edit_type='combine_feeders',
            status=status,
            baseline_fingerprint=f'fingerprint-{index}',
            generated_snapshot={
                'schema_version': SLD_GRAPH_SCHEMA_VERSION,
                'nodes': [{'component_id': f'MCB-{index}', 'metadata': {'large': 'x' * 2000}}],
            },
            edit_payload={
                'sld_payload': {
                    'nodes': [{'component_id': f'MCB-{index}', 'metadata': {'large': 'y' * 2000}}],
                    'edges': [],
                },
                'topology_operations': [{
                    'schema_version': 1,
                    'operation_type': 'combine_feeders',
                    'inputs': {'component_ids': [f'MCB-{index}', f'MCB-{index + 1}']},
                    'preview': {
                        'warning': 'Manual topology edit requires engineering review.',
                        'recommended_breaker_rating': 16,
                    },
                }],
                'cable_schedule_rows': [{'line_id': f'LINE-{index}', 'large': 'z' * 2000}],
            },
            validation_summary={'status': 'needs_review', 'operation_count': 1},
        )

    def test_compaction_turns_old_history_into_audit_only_payload(self):
        project = make_project_record(proj_id=self._project_id('hcmp'))
        edit = self._create_history_row(project, 1)
        original_size = len(json.dumps(edit.generated_snapshot)) + len(json.dumps(edit.edit_payload))

        result = compact_topology_edit_record(edit, reason='unit_test')

        edit.refresh_from_db()
        self.assertTrue(result['ok'])
        self.assertTrue(result['changed'])
        self.assertGreater(result['saved_size_bytes'], 0)
        self.assertFalse(edit.generated_snapshot)
        self.assertTrue(edit.edit_payload['history_payload_compacted'])
        self.assertEqual(edit.edit_payload['history_payload_compaction']['reason'], 'unit_test')
        self.assertEqual(edit.edit_payload['topology_operation_audit_summary'][0]['operation_type'], 'combine_feeders')
        self.assertNotIn('sld_payload', edit.edit_payload)
        self.assertLess(len(json.dumps(edit.edit_payload)), original_size)

    def test_retention_keeps_latest_full_history_and_compacts_older_rows(self):
        project_id = self._project_id('hret')
        project = make_project_record(proj_id=project_id)
        oldest = self._create_history_row(project, 1)
        middle = self._create_history_row(project, 2)
        newest = self._create_history_row(project, 3)
        active = self._create_history_row(project, 4, status='applied')

        dry_run = compact_sld_topology_history(project_id=project_id, keep_full=1, keep_reset=0, dry_run=True)
        for edit in (oldest, middle, newest, active):
            edit.refresh_from_db()
            self.assertNotIn('history_payload_compacted', edit.edit_payload)
        self.assertEqual(dry_run['candidate_count'], 2)

        summary = compact_sld_topology_history(project_id=project_id, keep_full=1, keep_reset=0)

        self.assertEqual(summary['compacted_count'], 2)
        oldest.refresh_from_db()
        middle.refresh_from_db()
        newest.refresh_from_db()
        active.refresh_from_db()
        self.assertTrue(oldest.edit_payload['history_payload_compacted'])
        self.assertTrue(middle.edit_payload['history_payload_compacted'])
        self.assertNotIn('history_payload_compacted', newest.edit_payload)
        self.assertNotIn('history_payload_compacted', active.edit_payload)

    def test_management_command_is_dry_run_until_execute_flag_is_used(self):
        project_id = self._project_id('hcmd')
        project = make_project_record(proj_id=project_id)
        oldest = self._create_history_row(project, 1)
        newest = self._create_history_row(project, 2)
        output = StringIO()

        call_command(
            'compact_sld_topology_history',
            '--project-id',
            project_id,
            '--keep-full',
            '1',
            stdout=output,
        )
        oldest.refresh_from_db()
        newest.refresh_from_db()
        self.assertNotIn('history_payload_compacted', oldest.edit_payload)
        self.assertNotIn('history_payload_compacted', newest.edit_payload)

        call_command(
            'compact_sld_topology_history',
            '--project-id',
            project_id,
            '--keep-full',
            '1',
            '--execute',
        )

        oldest.refresh_from_db()
        newest.refresh_from_db()
        self.assertTrue(oldest.edit_payload['history_payload_compacted'])
        self.assertNotIn('history_payload_compacted', newest.edit_payload)


class ProjectDataFormTests(TestCase):
    def test_form_lists_only_assigned_projects_for_logged_in_user(self):
        user = User.objects.create_user(username='planner', password='password123')
        make_managed_project(proj_id='P-001', description='Assigned', users=[user])
        make_managed_project(proj_id='P-002', description='Not Assigned')
        make_managed_project(proj_id='DEFAULT_PROJECT', description='Default project', users=[user])

        form = ProjectDataForm(user=user)
        choices = {value for value, _label in form.fields['proj_id'].choices}

        self.assertIn('', choices)
        self.assertIn('P-001', choices)
        self.assertNotIn('P-002', choices)
        self.assertNotIn('DEFAULT_PROJECT', choices)

    def test_default_project_id_helper_is_case_insensitive(self):
        self.assertTrue(is_default_project_id('DEFAULT_PROJECT'))
        self.assertTrue(is_default_project_id('default_project'))
        self.assertFalse(is_default_project_id('project_default'))

    def test_form_saves_setup_for_registered_project_and_hides_tracer_family(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        form = ProjectDataForm(data=make_project_form_payload())

        self.assertTrue(form.is_valid(), form.errors)
        project = form.save()
        self.assertEqual(project.proj_id, 'PLANT_A_001')
        self.assertEqual(project.eht_db_fault_rating_ka, Decimal('15'))
        self.assertAlmostEqual(project.eht_db_source_impedance_ohm, 230.0 / 15000.0, places=8)
        self.assertEqual(project.req_local_isolator, 'required')
        self.assertNotIn('tracer_family', form.fields)
        self.assertNotIn('req_local_isolator', form.fields)
        self.assertEqual(form.fields['proj_id'].help_text, 'Projects are managed in Django admin.')

    def test_form_accepts_custom_eht_db_fault_rating(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        form = ProjectDataForm(data=make_project_form_payload(
            eht_db_fault_rating_ka_preset='OTHER',
            eht_db_fault_rating_ka_custom='18.50',
        ))

        self.assertTrue(form.is_valid(), form.errors)
        project = form.save()
        self.assertEqual(project.eht_db_fault_rating_ka, Decimal('18.50'))
        self.assertAlmostEqual(project.eht_db_source_impedance_ohm, 230.0 / 18500.0, places=8)

    def test_form_rejects_missing_or_too_small_custom_eht_db_fault_rating(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        missing = ProjectDataForm(data=make_project_form_payload(
            eht_db_fault_rating_ka_preset='OTHER',
            eht_db_fault_rating_ka_custom='',
        ))
        too_small = ProjectDataForm(data=make_project_form_payload(
            eht_db_fault_rating_ka_preset='OTHER',
            eht_db_fault_rating_ka_custom='0.50',
        ))

        self.assertFalse(missing.is_valid())
        self.assertIn('eht_db_fault_rating_ka_custom', missing.errors)
        self.assertFalse(too_small.is_valid())
        self.assertIn('eht_db_fault_rating_ka_custom', too_small.errors)

    def test_form_accepts_small_standard_breaker_ratings(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        self.assertEqual([rating for rating, _label in MAX_CB_SIZE[:3]], [2, 4, 6])

        for rating in ['2', '4', '6']:
            form = ProjectDataForm(data=make_project_form_payload(max_cb_size=rating))
            self.assertTrue(form.is_valid(), form.errors)

    def test_form_rejects_unregistered_project_id(self):
        make_managed_project(proj_id='PLANT_B_001', description='Plant B')
        form = ProjectDataForm(data=make_project_form_payload())

        self.assertFalse(form.is_valid())
        self.assertIn('proj_id', form.errors)

    def test_existing_project_form_keeps_current_project_selected(self):
        project = make_project_record(proj_id='p1')
        form = ProjectDataForm(instance=project)

        self.assertFalse(form.fields['proj_id'].disabled)
        self.assertEqual(form.initial['proj_id'], 'p1')

    def test_project_save_syncs_local_isolator_requirement_from_location(self):
        project = make_project_record(proj_id='p-sync', isolator_location='noIsolator', req_local_isolator='required')

        project.refresh_from_db()
        self.assertEqual(project.req_local_isolator, 'not_required')

    def test_form_rejects_project_setup_values_that_break_domain_rules(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        form = ProjectDataForm(data=make_project_form_payload(
            min_amb_t='45.00',
            max_amb_t='20.00',
            restrict_cb_current='125.00',
            heat_loss_sf='0.80',
        ))

        self.assertFalse(form.is_valid())
        self.assertIn('min_amb_t', form.errors)
        self.assertIn('restrict_cb_current', form.errors)
        self.assertIn('heat_loss_sf', form.errors)

    def test_project_data_direct_save_runs_domain_validation(self):
        with self.assertRaises(ValidationError):
            make_project_record(proj_id='p-invalid', voltage=0)

        with self.assertRaises(ValidationError):
            make_project_record(proj_id='p-bad-fr', eht_db_fault_rating_ka=0.5)


class ProjectDataViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username='project-admin',
            email='project-admin@example.com',
            password='password123',
        )
        self.client.force_login(self.user)

    def test_create_project_data_view_persists_registered_project_id(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        response = self.client.post(reverse('create_project_data'), data=make_project_form_payload())

        self.assertEqual(response.status_code, 200)
        self.assertTrue(ProjectData.objects.filter(proj_id='PLANT_A_001').exists())

    def test_create_project_data_view_updates_existing_project_vendor(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        make_project_record(proj_id='PLANT_A_001', vendor='THR')

        response = self.client.post(
            reverse('create_project_data'),
            data=make_project_form_payload(proj_id='PLANT_A_001', vendor='CHR'),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(ProjectData.objects.get(proj_id='PLANT_A_001').vendor, 'CHR')

    def test_base_workspace_form_posts_to_project_data_save_route(self):
        response = self.client.get(reverse('base'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'action="/create-project-data/"')

    def test_base_workspace_uses_versioned_sld_script(self):
        response = self.client.get(reverse('base'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'js/sld_workspace.js?v=sld-r3-hit-targets')

    def test_base_workspace_renders_cold_cable_project_settings(self):
        response = self.client.get(reverse('base'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cold cable standard')
        self.assertContains(response, 'EHT DB fault rating')
        self.assertContains(response, '15 kA')
        self.assertContains(response, 'Other EHT DB fault rating')
        self.assertContains(response, 'Cold cable conductor material')
        self.assertContains(response, 'Cold cable installation method')
        self.assertContains(response, 'Cable grouping derating factor')
        self.assertContains(response, 'Minimum cold cable size')
        self.assertContains(response, 'MCB characteristic curve')
        self.assertContains(response, 'RCD / earth fault protection provided')
        self.assertContains(response, 'E - Multi-core on open cable tray or ladder')
        self.assertContains(response, 'D2 - Direct buried in ground (coming soon)')
        self.assertContains(response, 'value="D2" disabled')
        self.assertContains(response, 'Method E is active for this phase')
        self.assertNotContains(response, 'Cold cable catalogue readiness')

    def test_update_project_data_workspace_form_posts_to_project_update_route(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        make_project_record(proj_id='PLANT_A_001')

        response = self.client.get(reverse('update_project_data', args=['PLANT_A_001']))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'action="/edit-project-data/PLANT_A_001/"')

    def test_update_project_data_view_renders_blank_form_for_registered_project_without_setup(self):
        make_managed_project(proj_id='PLANT_B_001', description='Plant B')

        response = self.client.get(
            reverse('update_project_data', args=['PLANT_B_001']),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ProjectData.objects.filter(proj_id='PLANT_B_001').exists())
        self.assertIn('PLANT_B_001', response.json()['form_html'])

    def test_update_project_data_partial_renders_cold_cable_project_settings(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        make_project_record(proj_id='PLANT_A_001')

        response = self.client.get(
            reverse('update_project_data', args=['PLANT_A_001']),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        form_html = response.json()['form_html']
        self.assertIn('Cold cable standard', form_html)
        self.assertIn('EHT DB fault rating', form_html)
        self.assertIn('15 kA', form_html)
        self.assertIn('Other EHT DB fault rating', form_html)
        self.assertIn('Cold cable conductor material', form_html)
        self.assertIn('Cold cable installation method', form_html)
        self.assertIn('Cable grouping derating factor', form_html)
        self.assertIn('Minimum cold cable size', form_html)
        self.assertIn('MCB characteristic curve', form_html)
        self.assertIn('RCD / earth fault protection provided', form_html)
        self.assertIn('D2 - Direct buried in ground (coming soon)', form_html)
        self.assertIn('value="D2" disabled', form_html)
        self.assertIn('Method E is active for this phase', form_html)
        self.assertNotIn('Cold cable catalogue readiness', form_html)

    def test_default_project_button_copies_template_values_into_selected_project(self):
        make_managed_project(proj_id='PLANT_A_001', description='Plant A')
        make_default_project_record(proj_id='DEFAULT_PROJECT')

        response = self.client.post(
            reverse('create_project_data'),
            data={'proj_id': 'PLANT_A_001', 'action': 'load_defaults'},
        )

        self.assertEqual(response.status_code, 200)
        project = ProjectData.objects.get(proj_id='PLANT_A_001')
        self.assertEqual(project.vendor, 'THR')
        self.assertEqual(float(project.startup_t), 5.0)
        self.assertEqual(project.isolator_location, 'incomingOnly')
        self.assertEqual(project.req_local_isolator, 'required')

    def test_default_project_button_does_not_auto_create_template_project(self):
        make_managed_project(proj_id='PLANT_C_001', description='Plant C')

        response = self.client.post(
            reverse('create_project_data'),
            data={'proj_id': 'PLANT_C_001', 'action': 'load_defaults'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ProjectData.objects.filter(proj_id='PLANT_C_001').exists())
        self.assertFalse(ProjectData.objects.filter(proj_id__iexact=DEFAULT_PROJECT_ID).exists())


class ResultAndBoqViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username=f'results-{uuid4().hex[:8]}',
            email='results@example.com',
            password='password123',
        )
        self.client.force_login(self.user)

    def _make_variable_3c_cold_cable_project(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p-variable-3c-report', ['LINE-001'])
        project = ProjectData.objects.get(proj_id='p-variable-3c-report')
        project.allowablevdrop = 5.0
        project.save(update_fields=['allowablevdrop'])
        branch = PowerDistributionBranch.objects.select_related(
            'distribution',
            'distribution__line',
            'distribution__line__process_line_calculation',
        ).get(distribution__line__proj_id='p-variable-3c-report')
        branch.cable_length_db_to_jb = 30.0
        tagged_components = branch.tagged_components
        tagged_components['component_details']['MCB']['metadata']['operating_current'] = 8.0
        tagged_components['component_details']['MCB']['metadata']['per_circuit_operating_current'] = 8.0
        tagged_components['component_details']['MCB']['metadata']['breaker_size'] = 10.0
        branch.tagged_components = tagged_components
        branch.save(update_fields=['cable_length_db_to_jb', 'tagged_components'])
        cable3c_details = [
            downstream['component_details']['Cable3C']
            for downstream in branch.tagged_components['Downstream']
        ]
        for detail, length in zip(cable3c_details, [15.0, 150.0]):
            CableScheduleOverride.objects.create(
                project=project,
                component_id=detail['component_id'],
                component_uid=detail['component_uid'],
                display_tag=detail['display_tag'],
                component_type='Cable3C',
                line_id='LINE-001',
                line_uid=str(branch.distribution.line.uid),
                branch_index=branch.branch_index,
                circuit_index=detail['circuit_index'],
                generated_length_m=12.0,
                manual_length_m=length,
            )
        result = size_cold_cable_for_branch(branch, project)
        return project, result

    def test_import_input_view_renders_uploaded_line_list(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(reverse('import_input_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Input Line List')
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'data-toggle="table"')

    def test_input_data_export_returns_uploaded_line_sheet(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(reverse('input_data_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('p1_input_data.xlsx', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Input Data'])

        input_rows = list(workbook['Input Data'].iter_rows(values_only=True))
        self.assertIn(
            (
                'Project ID', 'Excel Row', 'Line ID', 'PID No', 'Area', 'Train',
                'Service Type', 'Line Size', 'Line Length', 'Valve Qty', 'Flange Qty',
                'Support Qty', 'Pipe Material Class', 'Insulation Material',
                'Insulation Thickness', 'Maintenance Temp', 'Operating Temp',
                'Design Temp', 'Emergency Supply', 'Discipline', 'Remarks', 'Status',
            ),
            input_rows,
        )
        self.assertTrue(any(row[2] == line.line_id for row in input_rows[1:]))

    def test_result_view_prompts_for_project_selection_when_missing(self):
        response = self.client.get(reverse('result_view'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Select a project in the Project Data form')

    def test_result_view_renders_stored_project_results(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Calculation Results')
        self.assertContains(response, 'data-toggle="table"')
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'V-001')
        self.assertContains(response, 'V-ALT-001')
        self.assertContains(response, 'MCB_001')

    def test_result_view_surfaces_cold_cable_sizing_results(self):
        call_command('populate_cold_cable_catalogue')
        make_calculated_project_snapshot()
        size_cold_cables_for_project('p1')

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cold Cable Sizing')
        self.assertContains(response, 'Cold cable engineering')
        self.assertContains(response, 'Critical Branch 2.5 mm²')
        self.assertContains(response, 'MT')

    def test_result_view_surfaces_panel_load_summary(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        size_cold_cables_for_project('p1')

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Panel / Load Summary')
        self.assertContains(response, 'Project p1 main distribution')
        self.assertContains(response, 'Grouped by panel/source metadata when present')
        self.assertContains(response, 'Selected 0')
        self.assertContains(response, 'Review 2')
        self.assertContains(response, 'Unsizeable 0')

    def test_panel_summary_uses_branch_current_before_line_total_fallback(self):
        branch_template = {
            'distribution': {'line': {'line_id': 'LINE-SHARED'}},
            'circuit_count': 1,
            'tagged_components': {
                'MCB': 'MCB-SHARED',
                'component_details': {
                    'MCB': {
                        'metadata': {
                            'operating_current': 10.0,
                            'line_operating_current': 20.0,
                            'breaker_size': 16.0,
                        },
                    },
                },
            },
        }
        branches = []
        for branch_index in [0, 1]:
            branch = deepcopy(branch_template)
            branch['branch_index'] = branch_index
            branch['cold_cable_result'] = SimpleNamespace(
                per_circuit_operating_current_a=10.0,
                line_operating_current_a=20.0,
                circuit_count=1,
                breaker_size_a=16.0,
                sizing_status='selected',
            )
            branches.append(branch)

        _rows, totals = _build_panel_load_summary('p1', branches, SimpleNamespace(voltage=230.0))

        self.assertEqual(totals['branch_count'], 2)
        self.assertEqual(totals['mcb_count'], 1)
        self.assertAlmostEqual(totals['load_current_a'], 20.0)
        self.assertAlmostEqual(totals['connected_load_kw'], 4.6)
        self.assertAlmostEqual(totals['breaker_capacity_a'], 16.0)

    def test_panel_summary_uses_line_current_when_per_circuit_current_is_zero(self):
        branch = {
            'distribution': {'line': {'line_id': 'LINE-FALLBACK'}},
            'branch_index': 1,
            'circuit_count': 2,
            'tagged_components': {'MCB': 'MCB-FALLBACK'},
            'cold_cable_result': SimpleNamespace(
                per_circuit_operating_current_a=0.0,
                line_operating_current_a=18.0,
                circuit_count=2,
                breaker_size_a=20.0,
                sizing_status='selected',
            ),
        }

        rows, totals = _build_panel_load_summary('p1', [branch], SimpleNamespace(voltage=230.0))

        self.assertAlmostEqual(totals['load_current_a'], 18.0)
        self.assertEqual(rows[0]['load_basis_display'], 'cold_cable_result.line_operating_current_a')

    def test_panel_summary_multiplies_per_circuit_current_by_circuit_count(self):
        branch = {
            'distribution': {'line': {'line_id': 'LINE-MULTI'}},
            'branch_index': 1,
            'circuit_count': 2,
            'tagged_components': {'MCB': 'MCB-MULTI'},
            'cold_cable_result': SimpleNamespace(
                per_circuit_operating_current_a=10.0,
                line_operating_current_a=99.0,
                circuit_count=2,
                breaker_size_a=25.0,
                sizing_status='selected',
            ),
        }

        rows, totals = _build_panel_load_summary('p1', [branch], SimpleNamespace(voltage=230.0))

        self.assertAlmostEqual(totals['load_current_a'], 20.0)
        self.assertAlmostEqual(totals['connected_load_kw'], 4.6)
        self.assertEqual(rows[0]['load_basis_display'], 'cold_cable_result.per_circuit_operating_current_a x circuit_count')

    def test_verification_report_lists_only_available_working_projects(self):
        user = User.objects.create_user(username='planner', password='password123')
        make_managed_project(proj_id='P-001', description='Assigned', users=[user])
        make_project_record(proj_id='P-001')
        make_managed_project(proj_id='P-002', description='Not assigned')
        make_project_record(proj_id='P-002')
        make_managed_project(proj_id=DEFAULT_PROJECT_ID, description='Default project', users=[user])
        make_default_project_record()
        self.client.force_login(user)

        response = self.client.get(reverse('verification_report_view'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'P-001')
        self.assertNotContains(response, 'P-002')
        self.assertNotContains(response, DEFAULT_PROJECT_ID)

    def test_verification_report_renders_cold_cable_evidence(self):
        call_command('populate_cold_cable_catalogue')
        line = make_calculated_project_snapshot()
        size_cold_cables_for_project('p1')

        response = self.client.get(
            reverse('verification_report_view'),
            {'project_id': 'p1', 'line_uid': str(line.uid)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Calculation Verification Report')
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'Cold Cable Sizing')
        self.assertContains(response, 'VD_feeder = 2')
        self.assertContains(response, 'L-PE Fault Loop Check')

    def test_verification_report_does_not_truncate_cold_cable_branches(self):
        call_command('populate_cold_cable_catalogue')
        line = make_calculated_project_snapshot()
        size_cold_cables_for_project('p1')
        base_result = ColdCableResult.objects.get(project_id='p1', branch_index=1)
        for branch_index in [2, 3, 4]:
            clone = ColdCableResult.objects.get(pk=base_result.pk)
            clone.pk = None
            clone.branch = None
            clone.branch_index = branch_index
            clone.save()

        response = self.client.get(
            reverse('verification_report_view'),
            {'project_id': 'p1', 'line_uid': str(line.uid)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Branch 4')

    def test_verification_report_rejects_unavailable_project(self):
        user = User.objects.create_user(username='planner', password='password123')
        make_managed_project(proj_id='P-001', description='Assigned', users=[user])
        make_project_record(proj_id='P-001')
        make_managed_project(proj_id='P-002', description='Not assigned')
        make_project_record(proj_id='P-002')
        self.client.force_login(user)

        response = self.client.get(reverse('verification_report_view'), {'project_id': 'P-002'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Project not found or unavailable.')
        self.assertNotContains(response, 'No confirmed lines found in this project')

    def test_result_view_shows_cable_override_summary(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        cable_node = next(node for node in payload['nodes'] if node['component_type'] == 'Cable3C')
        CableScheduleOverride.objects.create(
            project_id='p1',
            component_id=cable_node['component_id'],
            component_uid=cable_node['component_uid'],
            display_tag=cable_node['display_tag'],
            component_type=cable_node['component_type'],
            line_id=cable_node['line_id'],
            line_uid=cable_node['line_uid'],
            branch_index=cable_node['branch_index'],
            circuit_index=cable_node['circuit_index'],
            generated_length_m=(cable_node['metadata'] or {}).get('length_m'),
            manual_length_m=88.5,
        )

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cable Overrides')
        self.assertContains(response, cable_node['display_tag'])
        self.assertContains(response, '88.50 m')

    def test_result_view_marks_sld_tracer_override_as_review_only(self):
        line = make_calculated_project_snapshot()
        TracerSelectionOverride.objects.create(
            project_id='p1',
            line=line,
            selected_v_uid='V-ALT-001',
            selected_option_rank=1,
            remarks='Use alternate tracer after field review.',
        )

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'tracer selection override')
        self.assertContains(response, 'V-ALT-001')
        self.assertContains(response, 'Review-only')

    def test_result_view_uses_applied_topology_summary_override(self):
        make_calculated_project_snapshot()
        generated_payload = build_project_sld_payload('p1')
        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={
                'downstream_summaries': {'result': {'branch_count': 99}},
                'cable_schedule_rows': [{
                    'distribution': {'line': {'line_id': 'EDITED-LINE'}},
                    'branch_index': 1,
                    'branch_type': 'edited',
                    'connected_to': 'manual',
                    'circuit_count': 2,
                    'cable_length_db_to_jb': 12.5,
                    'cable_length_jb_to_jb': None,
                    'tagged_components': {'MCB': 'MCB_001-M', 'JB3PH': 'JB3PH_001-M', 'Downstream': [1, 2]},
                }],
            },
            validation_summary={'status': 'passed'},
        )

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '99 branch rows')
        self.assertContains(response, 'EDITED-LINE')
        self.assertContains(response, 'MCB_001-M')

    def test_cable_schedule_view_prompts_for_project_selection_when_missing(self):
        response = self.client.get(reverse('cable_schedule_view'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Select a project in the Project Data form')

    def test_cable_schedule_view_renders_active_schedule_table(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(reverse('cable_schedule_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cable Schedule')
        self.assertContains(response, 'data-toggle="table"')
        self.assertContains(response, 'Cable Tag')
        self.assertContains(response, 'Cable Specification')
        self.assertContains(response, 'Connected From')
        self.assertContains(response, 'Manual Edit')
        self.assertContains(response, 'data-field="manual-edit"')
        self.assertContains(response, 'data-visible="false"')
        self.assertContains(response, 'LINE-001')
        self.assertContains(response, 'CCAB')
        self.assertContains(response, 'MCB_001')

    def test_cable_schedule_view_surfaces_calculated_cold_cable_size(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        size_cold_cables_for_project('p1')

        response = self.client.get(reverse('cable_schedule_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Calculated Cold Cable')
        self.assertContains(response, 'Feeder Cable')
        self.assertContains(response, '3C x 2.5 mm2')
        self.assertContains(response, 'Total Cable Copper Tonnage')

    def test_cable_schedule_view_surfaces_each_3c_segment_and_critical_status(self):
        self._make_variable_3c_cold_cable_project()

        response = self.client.get(reverse('cable_schedule_view'), {'project_id': 'p-variable-3c-report'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Branch Cable')
        self.assertContains(response, 'critical Branch')
        self.assertContains(response, '3C x 2.5 mm2')
        self.assertContains(response, '3C x 6 mm2')

    def test_cable_schedule_summary_totals_branch_conductor_mass_once(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        size_cold_cables_for_project('p1')

        schedule_data = build_cable_schedule_workspace_data('p1')
        expected_mass = sum(result.conductor_mass_total_mt or 0 for result in ColdCableResult.objects.filter(project_id='p1'))

        self.assertGreater(schedule_data['summary']['total_conductor_mass_mt'], 0)
        self.assertAlmostEqual(schedule_data['summary']['total_conductor_mass_mt'], expected_mass, places=9)

    def test_cable_schedule_view_shows_manual_override_status_and_remarks(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        generated_payload = build_project_sld_payload('p1')
        cable_node = next(node for node in generated_payload['nodes'] if node['component_type'] == 'Cable3C')
        CableScheduleOverride.objects.create(
            project_id='p1',
            component_id=cable_node['component_id'],
            component_uid=cable_node['component_uid'],
            display_tag=cable_node['display_tag'],
            component_type=cable_node['component_type'],
            line_id=cable_node['line_id'],
            line_uid=cable_node['line_uid'],
            branch_index=cable_node['branch_index'],
            circuit_index=cable_node['circuit_index'],
            generated_length_m=(cable_node['metadata'] or {}).get('length_m'),
            manual_length_m=55,
            generated_cable_size=(cable_node['metadata'] or {}).get('cable_size'),
            manual_cable_size='3C x 6',
            remarks='Measured along pipe rack A.',
        )

        response = self.client.get(reverse('cable_schedule_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Manual')
        self.assertContains(response, '<td>Yes</td>', html=True)
        self.assertContains(response, 'Measured along pipe rack A.')

    def test_cable_schedule_view_warns_when_manual_size_is_below_calculated_size(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        size_cold_cables_for_project('p1')
        generated_payload = build_project_sld_payload('p1')
        cable_node = next(node for node in generated_payload['nodes'] if node['component_type'] == 'Cable3C')
        CableScheduleOverride.objects.create(
            project_id='p1',
            component_id=cable_node['component_id'],
            component_uid=cable_node['component_uid'],
            display_tag=cable_node['display_tag'],
            component_type=cable_node['component_type'],
            line_id=cable_node['line_id'],
            line_uid=cable_node['line_uid'],
            branch_index=cable_node['branch_index'],
            circuit_index=cable_node['circuit_index'],
            generated_length_m=(cable_node['metadata'] or {}).get('length_m'),
            manual_cable_size='3C x 1.5',
        )

        response = self.client.get(reverse('cable_schedule_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Manual size below calculated value.')
        self.assertContains(response, '1 size review')

    def test_cable_schedule_export_includes_manual_size_review_columns(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        size_cold_cables_for_project('p1')
        generated_payload = build_project_sld_payload('p1')
        cable_node = next(node for node in generated_payload['nodes'] if node['component_type'] == 'Cable3C')
        CableScheduleOverride.objects.create(
            project_id='p1',
            component_id=cable_node['component_id'],
            component_uid=cable_node['component_uid'],
            display_tag=cable_node['display_tag'],
            component_type=cable_node['component_type'],
            line_id=cable_node['line_id'],
            line_uid=cable_node['line_uid'],
            branch_index=cable_node['branch_index'],
            circuit_index=cable_node['circuit_index'],
            generated_length_m=(cable_node['metadata'] or {}).get('length_m'),
            manual_cable_size='3C x 1.5',
        )

        response = self.client.get(reverse('cable_schedule_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Cable Schedule'].iter_rows(values_only=True))
        header = rows[0]
        review_status_index = header.index('Manual Size Review')
        review_note_index = header.index('Manual Size Review Note')
        export_row = next(row for row in rows[1:] if row[1] == cable_node['display_tag'])
        self.assertEqual(export_row[review_status_index], 'undersized')
        self.assertIn('below the calculated cold cable size', export_row[review_note_index])

    def test_cable_schedule_view_uses_applied_topology_rows(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        generated_payload = build_project_sld_payload('p1')
        mcb_component_ids = [
            node['component_id']
            for node in generated_payload['nodes']
            if node['component_type'] == 'MCB'
        ]
        apply_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': mcb_component_ids,
                'trunk_length_m': 25,
                'cable_size': '4C x 10',
            }),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)

        response = self.client.get(reverse('cable_schedule_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Manual SLD topology')
        self.assertContains(response, 'MCB_001-M')
        self.assertContains(response, '4C x 10')
        self.assertContains(response, '25.00')
        self.assertContains(response, 'MCB to Distribution JB')

    def test_cable_schedule_export_requires_project_id(self):
        response = self.client.get(reverse('cable_schedule_export_view'))

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Project ID is required to export cable schedule.')

    def test_cable_schedule_export_returns_schedule_sheet(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(reverse('cable_schedule_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('p1_cable_schedule.xlsx', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Cable Schedule'])

        rows = list(workbook['Cable Schedule'].iter_rows(values_only=True))
        self.assertEqual(rows[0], (
            'Sr. No',
            'Cable Tag',
            'Cable Specification',
            'Calculated Cold Cable Size',
            'Cold Cable Segment Role',
            'Cold Cable Circuit Index',
            'Cold Cable Status',
            'Voltage Drop Status',
            'Fault Protection Status',
            'Total Path VD (%)',
            'Load-End Voltage (V)',
            'Fault Current (A)',
            'Length Basis',
            'Critical Branch Segment',
            'Conductor Mass (MT)',
            'Cable Length (m)',
            'Connected From',
            'Connected To',
            'Line IDs',
            'Purpose',
            'Cable Drum Tag',
            'Cable Route Details',
            'Remarks',
            'Manual Size Review',
            'Manual Size Review Note',
            'Rev. No.',
        ))
        self.assertTrue(any(row[1] and str(row[1]).startswith('CCAB') for row in rows[1:]))
        self.assertTrue(any(row[18] == 'LINE-001' for row in rows[1:]))

    def test_cable_schedule_export_includes_per_3c_segment_evidence(self):
        self._make_variable_3c_cold_cable_project()

        response = self.client.get(reverse('cable_schedule_export_view'), {'project_id': 'p-variable-3c-report'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Cable Schedule'].iter_rows(values_only=True))
        header = rows[0]
        size_index = header.index('Calculated Cold Cable Size')
        role_index = header.index('Cold Cable Segment Role')
        length_index = header.index('Cable Length (m)')
        critical_index = header.index('Critical Branch Segment')
        vd_index = header.index('Total Path VD (%)')
        load_end_index = header.index('Load-End Voltage (V)')

        outgoing_rows = [row for row in rows[1:] if row[role_index] == 'Branch Cable']
        self.assertTrue(any(row[size_index] == '3C x 2.5 mm2' and row[length_index] == 15 for row in outgoing_rows))
        critical_rows = [row for row in outgoing_rows if row[critical_index] == 'Yes']
        self.assertEqual(len(critical_rows), 1)
        self.assertEqual(critical_rows[0][size_index], '3C x 6 mm2')
        self.assertGreater(critical_rows[0][vd_index], 0)
        self.assertGreater(critical_rows[0][load_end_index], 0)

    def test_cable_schedule_export_uses_applied_topology_rows(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        generated_payload = build_project_sld_payload('p1')
        mcb_component_ids = [
            node['component_id']
            for node in generated_payload['nodes']
            if node['component_type'] == 'MCB'
        ]
        apply_response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_ids': mcb_component_ids,
                'trunk_length_m': 25,
                'cable_size': '4C x 10',
            }),
            content_type='application/json',
        )
        self.assertEqual(apply_response.status_code, 200)

        response = self.client.get(reverse('cable_schedule_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Cable Schedule'].iter_rows(values_only=True))
        header = rows[0]
        tag_index = header.index('Cable Tag')
        spec_index = header.index('Cable Specification')
        length_index = header.index('Cable Length (m)')
        from_index = header.index('Connected From')
        self.assertTrue(any(
            row[tag_index] and '-M' in row[tag_index] and row[spec_index] == '4C x 10' and row[length_index] == 25
            for row in rows[1:]
        ))
        self.assertTrue(any(row[from_index] == 'MCB_001-M' for row in rows[1:]))

    def test_result_export_returns_line_branch_and_alternate_sheets(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('p1_results.xlsx', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, [
            'Line Results',
            'Selection Diagnostics',
            'Power Distribution',
            'Panel Load Summary',
            'Cold Cable Sizing',
            'Cold Cable Branch Segments',
            'Alternate Tracers',
            'MI Selection',
        ])

        line_rows = list(workbook['Line Results'].iter_rows(values_only=True))
        branch_rows = list(workbook['Power Distribution'].iter_rows(values_only=True))
        alternate_rows = list(workbook['Alternate Tracers'].iter_rows(values_only=True))

        self.assertTrue(any(row[1] == line.line_id for row in line_rows[1:]))
        self.assertTrue(any(row[1] == line.line_id for row in branch_rows[1:]))
        self.assertTrue(any(row[1] == line.line_id for row in alternate_rows[1:]))

    def test_result_export_includes_panel_load_summary_sheet(self):
        call_command('populate_cold_cable_catalogue')
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        size_cold_cables_for_project('p1')

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Panel Load Summary'].iter_rows(values_only=True))
        header = rows[0]
        data = rows[1]
        self.assertEqual(data[header.index('Source Label')], 'Project p1 main distribution')
        self.assertEqual(data[header.index('MCB Count')], 2)
        self.assertEqual(data[header.index('Cold Cable Selected')], 0)
        self.assertEqual(data[header.index('Cold Cable Review Required')], 2)
        self.assertEqual(data[header.index('Cold Cable Unsizeable')], 0)
        self.assertGreater(data[header.index('Connected Load (kW)')], 0)

    def test_result_export_includes_cold_cable_sizing_sheet(self):
        call_command('populate_cold_cable_catalogue')
        line = make_calculated_project_snapshot()
        size_cold_cables_for_project('p1')

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Cold Cable Sizing'].iter_rows(values_only=True))
        header = rows[0]
        phase_basis_index = header.index('Phase Balance Basis')
        phase_l1_index = header.index('L1 Current (A)')
        phase_imbalance_index = header.index('Phase Imbalance (A)')
        self.assertTrue(any(row[1] == line.line_id and row[15] == 2.5 for row in rows[1:]))
        self.assertTrue(any(row[39] and 'project default' in row[39] for row in rows[1:]))
        self.assertTrue(any(row[phase_basis_index] and 'round-robin' in row[phase_basis_index] for row in rows[1:]))
        self.assertTrue(any(row[phase_l1_index] and row[phase_l1_index] > 0 for row in rows[1:]))
        self.assertTrue(any(row[phase_imbalance_index] is not None for row in rows[1:]))

    def test_result_export_includes_per_3c_segment_sheet(self):
        self._make_variable_3c_cold_cable_project()

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p-variable-3c-report'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Cold Cable Branch Segments'].iter_rows(values_only=True))
        header = rows[0]
        critical_size_index = header.index('Branch Critical Size (mm2)')
        segment_size_index = header.index('Branch Segment Size (mm2)')
        critical_index = header.index('Critical Segment')
        length_index = header.index('Segment Length (m)')
        tag_index = header.index('Segment Display Tag')
        phase_label_index = header.index('Phase Label')
        phase_basis_index = header.index('Phase Basis')

        self.assertEqual(len(rows), 3)
        self.assertTrue(any(row[segment_size_index] == 2.5 and row[length_index] == 15 for row in rows[1:]))
        self.assertEqual([row[phase_label_index] for row in rows[1:]], ['L1', 'L2'])
        self.assertTrue(all(row[phase_basis_index] == 'round_robin_by_outgoing_circuit_index' for row in rows[1:]))
        critical_rows = [row for row in rows[1:] if row[critical_index] == 'Yes']
        self.assertEqual(len(critical_rows), 1)
        self.assertEqual(critical_rows[0][critical_size_index], 6)
        self.assertEqual(critical_rows[0][segment_size_index], 6)
        self.assertTrue(critical_rows[0][tag_index])

    def test_result_view_surfaces_3c_segments_and_critical_size(self):
        self._make_variable_3c_cold_cable_project()

        response = self.client.get(reverse('result_view'), {'project_id': 'p-variable-3c-report'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Critical Branch 6 mm²')
        self.assertContains(response, 'Branch Cable segments')
        self.assertContains(response, 'Branch 2.5 mm²')
        self.assertContains(response, 'Branch 6 mm²')
        self.assertContains(response, '(critical)')
        self.assertContains(response, 'L1 8.0 A')
        self.assertContains(response, 'L2 8.0 A')
        self.assertContains(response, 'L3 0.0 A')
        self.assertContains(response, 'Imbalance 8.0 A')

    def test_result_view_and_export_surface_mi_selection_records(self):
        line = make_calculated_project_snapshot()
        mi_result = make_mi_selection_result(line)

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'MI Selection Records')
        self.assertContains(response, mi_result.heater.part_number)
        self.assertContains(response, 'available_alternative')

        export_response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertIn('MI Selection', workbook.sheetnames)
        rows = list(workbook['MI Selection'].iter_rows(values_only=True))
        header = rows[0]
        data = rows[1]
        self.assertEqual(data[header.index('Line ID')], line.line_id)
        self.assertEqual(data[header.index('MI Selection Status')], 'available_alternative')
        self.assertEqual(data[header.index('Heater Part Number')], mi_result.heater.part_number)

    def test_result_view_includes_mi_fallback_line_without_sr_calculation(self):
        make_project_record(proj_id='p1')
        line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-MI-ONLY',
            service_type='EP',
            line_size=3.0,
            line_length=25.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=320.0,
            oper_temp=330.0,
            design_temp=360.0,
            status='confirmed',
        )
        make_mi_selection_result(line, status='selected')

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'LINE-MI-ONLY')
        self.assertContains(response, 'MI fallback selected')
        self.assertContains(response, 'Per heater set')
        self.assertContains(response, 'MI MVP design basis and review notes')
        self.assertContains(response, 'Each MI heater set is treated as an independently protected branch')
        self.assertContains(response, 'MI T-class status remains design-review evidence')

    def test_result_view_summarizes_mi_fallback_as_selected_output_not_unresolved_sr_failure(self):
        line = make_mi_calculated_project_snapshot()

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'SR 0 /')
        self.assertContains(response, 'MI 1')
        self.assertContains(response, 'Auto fallback')
        self.assertContains(response, 'Cold lead 2.00 m')
        self.assertContains(response, 'Physical JB terminal capacity')
        self.assertNotContains(response, 'did not receive a selectable SR tracer')

    def test_result_view_explains_rejected_mi_fallback_without_zero_design_values(self):
        line = make_mi_rejected_project_snapshot()

        response = self.client.get(reverse('result_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'triggered MI fallback but did not receive a selected MI heater')
        self.assertContains(response, 'NO_VALIDATED_MI_CATALOGUE_DATA')
        self.assertContains(response, '3 MI family row(s) exist for THR, but none are marked as validated')
        self.assertContains(response, 'Validate reviewed MI catalogue rows for the selected vendor')
        self.assertNotContains(response, '0.00 W/m')
        self.assertNotContains(response, '0.00 W total')

    def test_result_export_identifies_mi_fallback_heater_type_and_lengths(self):
        line = make_mi_calculated_project_snapshot()

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Line Results'].iter_rows(values_only=True))
        header = rows[0]
        data = next(row for row in rows[1:] if row[header.index('Line ID')] == line.line_id)
        self.assertEqual(data[header.index('Heating Cable Type')], 'MI')
        self.assertEqual(data[header.index('Selected Tracer')], f'MIQ-R{line.uid}')
        self.assertEqual(data[header.index('MI Heated Length excl. Cold Leads (m)')], 18.0)
        self.assertEqual(data[header.index('MI Cold Lead Length (m)')], 2.0)
        self.assertIn('MI heated length excludes cold leads', data[header.index('Heating Cable Length Basis')])
        self.assertIn('Independent breaker per MI heater set', data[header.index('MI Design Basis Notes')])
        self.assertIn('T-class requires design review', data[header.index('MI Design Basis Notes')])

    def test_result_export_explains_rejected_mi_fallback_next_action(self):
        line = make_mi_rejected_project_snapshot()

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['MI Selection'].iter_rows(values_only=True))
        header = rows[0]
        data = next(row for row in rows[1:] if row[header.index('Line ID')] == line.line_id)
        self.assertEqual(data[header.index('MI Selection Status')], 'rejected')
        self.assertEqual(data[header.index('Rejection Code')], 'NO_VALIDATED_MI_CATALOGUE_DATA')
        self.assertIn('No validated MI catalogue rows', data[header.index('Rejection Message')])
        self.assertIn('none are marked as validated', data[header.index('Diagnostic Evidence')])
        self.assertIn('Validate reviewed MI catalogue rows', data[header.index('Next Action')])

    def test_result_export_marks_sld_tracer_override_review_only(self):
        line = make_calculated_project_snapshot()
        TracerSelectionOverride.objects.create(
            project_id='p1',
            line=line,
            selected_v_uid='V-ALT-001',
            selected_option_rank=1,
        )

        response = self.client.get(reverse('result_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Line Results'].iter_rows(values_only=True))
        header = rows[0]
        data = rows[1]
        self.assertEqual(data[header.index('SLD Tracer Override')], 'V-ALT-001')
        self.assertIn('Review-only', data[header.index('SLD Override Review Status')])

    def test_boq_view_renders_consolidated_and_line_items(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(reverse('boq_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Bill of Quantities')
        self.assertContains(response, 'data-toggle="table"')
        self.assertContains(response, 'TRACER')
        self.assertContains(response, 'MCB')
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'Miniature Circuit Breaker')

    def test_boq_view_uses_applied_topology_summary_override(self):
        make_calculated_project_snapshot()
        generated_payload = build_project_sld_payload('p1', apply_topology=False)
        SLDTopologyEdit.objects.create(
            project_id='p1',
            edit_type='combine_feeders',
            status='applied',
            baseline_fingerprint=payload_fingerprint(generated_payload),
            edit_payload={'downstream_summaries': {'boq': {'mcb_total': 7, 'junction_box_total': 11}}},
            validation_summary={'status': 'passed'},
        )

        response = self.client.get(reverse('boq_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '7 / 11')

    def test_boq_line_detail_view_renders_inline_detail_partial(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(
            reverse('boq_line_detail_view'),
            {'project_id': 'p1', 'line_id': line.line_id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'Miniature Circuit Breaker')

    def test_boq_export_returns_summary_and_per_line_sheets(self):
        line = make_calculated_project_snapshot()

        response = self.client.get(reverse('boq_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('p1_boq.xlsx', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['BOQ Summary', 'BOQ Per Line'])

        summary_rows = list(workbook['BOQ Summary'].iter_rows(values_only=True))
        detail_rows = list(workbook['BOQ Per Line'].iter_rows(values_only=True))

        self.assertIn(('Project ID', 'Item Code', 'Description', 'Quantity', 'Unit'), summary_rows)
        self.assertIn(('Project ID', 'Line ID', 'Service Type', 'Item Code', 'Description', 'Quantity', 'Unit'), detail_rows)
        self.assertTrue(any(row[1] == 'TRACER' for row in summary_rows[1:]))
        self.assertTrue(any(row[1] == line.line_id for row in detail_rows[1:]))

    def test_sld_workspace_view_renders_project_backed_summary(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]

        response = self.client.get(reverse('sld_workspace_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Single Line Diagram')
        self.assertContains(response, 'Graph Ready')
        self.assertContains(response, 'SLD Validation')
        self.assertContains(response, 'Validation Passed')
        self.assertContains(response, 'without changing the calculated topology')
        self.assertContains(response, 'LINE-001')
        self.assertContains(response, reverse('sld_payload_view'))
        self.assertContains(response, reverse('sld_layout_view'))
        self.assertContains(response, reverse('sld_layout_reset_view'))
        self.assertContains(response, reverse('sld_topology_combine_preview_view'))
        self.assertContains(response, reverse('sld_topology_combine_apply_view'))
        self.assertContains(response, reverse('sld_topology_split_preview_view'))
        self.assertContains(response, reverse('sld_topology_split_apply_view'))
        self.assertContains(response, reverse('sld_topology_downstream_jb_preview_view'))
        self.assertContains(response, reverse('sld_topology_downstream_jb_apply_view'))
        self.assertContains(response, reverse('sld_topology_attach_jb_preview_view'))
        self.assertContains(response, reverse('sld_topology_attach_jb_apply_view'))
        self.assertContains(response, reverse('sld_topology_reset_view'))
        self.assertContains(response, reverse('sld_validation_view'))
        self.assertContains(response, line.line_id)
        self.assertContains(response, 'id="sld-diagram-shell"')
        self.assertContains(response, 'data-project-id="p1"')
        self.assertContains(response, 'Generated Topology')
        self.assertContains(response, 'Save Layout')
        self.assertContains(response, 'Reset Layout')
        self.assertContains(response, 'Apply Edit')
        self.assertContains(response, reverse('sld_cable_override_save_view'))
        self.assertContains(response, 'id="sld-cable-editor"')
        self.assertNotContains(response, 'Preview Edit')

    def test_sld_workspace_view_supports_line_focused_rendering(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        response = self.client.get(
            reverse('sld_workspace_view'),
            {'project_id': 'p1', 'line_id': 'line-002'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Focused view for line')
        self.assertContains(response, 'LINE-002')
        self.assertContains(response, 'data-selected-line-id="LINE-002"')
        self.assertContains(response, 'Property Inspector')
        self.assertContains(response, 'id="sld-line-filter-form"')
        self.assertContains(response, f'action="{reverse("base")}#sld-tab-pane"')
        self.assertContains(response, f'data-url="{reverse("sld_workspace_view")}"')
        self.assertContains(response, 'data-scroll-to=".sld-panel"')
        self.assertContains(response, 'id="sld-fit-view"')
        self.assertContains(response, 'id="sld-fit-selected-line"')
        self.assertContains(response, 'id="sld-export-pdf"')
        self.assertContains(response, 'Export PDF')
        self.assertContains(response, reverse('sld_pdf_export_view'))
        self.assertContains(response, 'id="sld-line-filter-reset"')
        self.assertNotContains(response, 'href="?project_id=p1"')

    def test_sld_workspace_view_reports_unknown_line_focus_request(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(
            reverse('sld_workspace_view'),
            {'project_id': 'p1', 'line_id': 'LINE-999'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'No SLD line group was found for line ID')
        self.assertContains(response, 'LINE-999')

    def test_sld_workspace_view_renders_inspector_and_navigation_tools(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(
            reverse('sld_workspace_view'),
            {'project_id': 'p1'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Property Inspector')
        self.assertContains(response, 'Select a component in the diagram')
        self.assertContains(response, 'id="sld-zoom-in"')
        self.assertContains(response, 'id="sld-zoom-out"')
        self.assertContains(response, 'id="sld-fit-view"')
        self.assertContains(response, 'id="sld-fit-selected-line"')
        self.assertContains(response, 'id="sld-export-pdf"')
        self.assertNotContains(response, 'Export SVG')

    def test_sld_pdf_export_view_returns_full_project_pdf(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002', 'LINE-003'])

        response = self.client.get(reverse('sld_pdf_export_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('p1_sld.pdf', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertGreater(len(response.content), 1000)

    def test_sld_pdf_cable_label_uses_compact_cold_cable_size_and_vd(self):
        from eht.sld_pdf import _cable_label

        label = _cable_label({
            'display_tag': 'CCAB4C_001',
            'metadata': {
                'cold_cable': {
                    'calculated_size': '4C x 10 mm2',
                    'vd_total_pct': 5.54,
                },
            },
        })

        self.assertEqual(label, '4Cx10mm² (Vd 5.54%)')

    def test_sld_pdf_renderer_draws_multicircuit_branch_links(self):
        from eht.sld_pdf import _pdf_rows

        make_calculated_project_snapshot()
        payload = build_project_sld_payload('p1')

        pdf_bytes = build_sld_pdf('p1', payload)
        rows = _pdf_rows(payload)

        self.assertTrue(pdf_bytes.startswith(b'%PDF'))
        self.assertGreaterEqual(len(rows), 2)
        self.assertTrue(all(row['path'][0]['component_type'] == 'MCB' for row in rows))

    def test_sld_pdf_renderer_routes_3ph_jb_outgoings_from_distinct_slots(self):
        from eht.sld_pdf import _edge_source_offsets, _pdf_trees, _tree_layout

        make_calculated_project_snapshot()
        payload = build_project_sld_payload('p1')
        trees = _pdf_trees(payload)
        layouts = [_tree_layout(tree) for tree in trees]
        layout = next(layout for layout in layouts if any(
            item['node'].get('component_type') == 'JB3PH'
            for item in layout['positions'].values()
        ))

        source_offsets = _edge_source_offsets(layout['edges'], layout['positions'])
        jb3ph_ids = {
            node_id
            for node_id, item in layout['positions'].items()
            if item['node'].get('component_type') == 'JB3PH'
        }
        jb3ph_outgoing_offsets = [
            offset
            for (source_id, _target_id), offset in source_offsets.items()
            if source_id in jb3ph_ids
        ]

        self.assertGreaterEqual(len(jb3ph_outgoing_offsets), 2)
        self.assertGreaterEqual(len(set(jb3ph_outgoing_offsets)), 2)
        self.assertIn(10, jb3ph_outgoing_offsets)
        self.assertIn(0, jb3ph_outgoing_offsets)

    def test_sld_pdf_renderer_keeps_shared_manual_components_on_combined_lines(self):
        from eht.sld_pdf import _line_rows, _pdf_trees

        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        generated_payload = build_project_sld_payload('p1')
        mcb_component_ids = [
            node['component_id']
            for node in generated_payload['nodes']
            if node['component_type'] == 'MCB'
        ]
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': mcb_component_ids}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        payload = build_project_sld_payload('p1')
        line_two_group = next(group for group in payload['line_groups'] if group['line_id'] == 'LINE-002')
        line_two_rows = _line_rows(payload, line_two_group)
        trees = _pdf_trees(payload)
        line_two_tags = {
            node['display_tag']
            for row in line_two_rows
            for node in [*row['upstream'], *row['downstream']]
        }

        self.assertIn('MCB_001-M', line_two_tags)
        self.assertTrue(any(tag.startswith('JB3PH_001') for tag in line_two_tags))
        self.assertEqual(len(trees), 1)
        self.assertEqual(len(trees[0]['paths']), 4)

    def test_sld_pdf_renderer_marks_review_required_topology_exports(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        payload['meta']['topology_edit_review_required'] = True
        payload['meta']['manual_topology_warning'] = 'Manual topology edit requires review before issue.'

        with patch('eht.sld_pdf._start_page') as start_page:
            pdf_bytes = build_sld_pdf('p1', payload)

        self.assertTrue(pdf_bytes.startswith(b'%PDF'))
        self.assertIn('Manual topology edit requires review before issue.', start_page.call_args.args)

    def test_sld_cable_override_save_and_reset_views_update_payload(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])
        payload = build_project_sld_payload('p1')
        cable_node = next(node for node in payload['nodes'] if node['component_type'] == 'Cable3C')

        save_response = self.client.post(
            reverse('sld_cable_override_save_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_id': cable_node['component_id'],
                'manual_length_m': 77.25,
                'manual_cable_size': '3C x 4',
                'remarks': 'Field measured route.',
            }),
            content_type='application/json',
        )

        self.assertEqual(save_response.status_code, 200)
        self.assertTrue(CableScheduleOverride.objects.filter(project_id='p1', component_id=cable_node['component_id'], is_active=True).exists())
        adjusted_payload = build_project_sld_payload('p1')
        adjusted_node = next(node for node in adjusted_payload['nodes'] if node['component_id'] == cable_node['component_id'])
        self.assertEqual(adjusted_node['metadata']['length_m'], 77.25)
        self.assertEqual(adjusted_node['metadata']['cable_size'], '3C x 4')

        reset_response = self.client.post(
            reverse('sld_cable_override_reset_view'),
            data=json.dumps({'project_id': 'p1', 'component_id': cable_node['component_id']}),
            content_type='application/json',
        )

        self.assertEqual(reset_response.status_code, 200)
        self.assertFalse(CableScheduleOverride.objects.get(project_id='p1', component_id=cable_node['component_id']).is_active)

    def test_sld_tracer_override_save_and_reset_views_update_payload(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]
        AlternateTracer.objects.create(
            line=line,
            option_rank=1,
            v_uid='V-ALT-001',
            a_coeff=0.0,
            b_coeff=0.0,
            c_coeff=80.0,
            power_at_startup_t=10.0,
            ohm_per_km=1.0,
            res_corrFactor_mica=1.0,
            tracer_family='SR-ALT',
            voltage_float=230.0,
            voltage_correction_factor=1.0,
            power_output=24.5,
            spiral_factor=1.15,
            tracer_length=21.25,
            tracer_with_margin=23.38,
        )
        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')

        save_response = self.client.post(
            reverse('sld_tracer_override_save_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_id': tracer_node['component_id'],
                'selected_v_uid': 'V-ALT-001',
                'remarks': 'Reviewed against alternate tracer.',
            }),
            content_type='application/json',
        )

        self.assertEqual(save_response.status_code, 200)
        self.assertTrue(TracerSelectionOverride.objects.filter(project_id='p1', line=line, is_active=True).exists())
        adjusted_payload = build_project_sld_payload('p1')
        adjusted_node = next(node for node in adjusted_payload['nodes'] if node['component_id'] == tracer_node['component_id'])
        self.assertEqual(adjusted_node['metadata']['tracer_selection']['selected']['v_uid'], 'V-ALT-001')
        self.assertTrue(adjusted_node['metadata']['tracer_selection']['override_active'])

        reset_response = self.client.post(
            reverse('sld_tracer_override_reset_view'),
            data=json.dumps({'project_id': 'p1', 'line_uid': str(line.uid)}),
            content_type='application/json',
        )

        self.assertEqual(reset_response.status_code, 200)
        self.assertFalse(TracerSelectionOverride.objects.get(project_id='p1', line=line).is_active)

    def test_sld_tracer_override_save_accepts_mi_available_alternative(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]
        mi_result = make_mi_selection_result(line)
        expected_uid = f'MI:{mi_result.heater.part_number}:{mi_result.cold_lead_option_code}'
        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')

        save_response = self.client.post(
            reverse('sld_tracer_override_save_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_id': tracer_node['component_id'],
                'selected_v_uid': expected_uid,
                'remarks': 'Use MI for this line after engineering review.',
            }),
            content_type='application/json',
        )

        self.assertEqual(save_response.status_code, 200)
        override = TracerSelectionOverride.objects.get(project_id='p1', line=line, is_active=True)
        self.assertEqual(override.selected_v_uid, expected_uid)
        adjusted_payload = build_project_sld_payload('p1')
        adjusted_node = next(node for node in adjusted_payload['nodes'] if node['component_id'] == tracer_node['component_id'])
        tracer_selection = adjusted_node['metadata']['tracer_selection']
        self.assertEqual(tracer_selection['selected']['v_uid'], expected_uid)
        self.assertEqual(tracer_selection['selected']['tracer_family'], 'MI')
        self.assertTrue(tracer_selection['override_active'])

    def test_sld_mi_tracer_override_survives_mi_result_recalculation(self):
        line = make_rich_sld_project_snapshot('p1', ['LINE-001'])[0]
        mi_result = make_mi_selection_result(line)
        expected_uid = f'MI:{mi_result.heater.part_number}:{mi_result.cold_lead_option_code}'
        payload = build_project_sld_payload('p1')
        tracer_node = next(node for node in payload['nodes'] if node['component_type'] == 'Tracer')
        self.client.post(
            reverse('sld_tracer_override_save_view'),
            data=json.dumps({
                'project_id': 'p1',
                'component_id': tracer_node['component_id'],
                'selected_v_uid': expected_uid,
            }),
            content_type='application/json',
        )
        old_result_id = mi_result.id
        heater = mi_result.heater
        cold_lead = mi_result.cold_lead_option
        mi_result.delete()
        refreshed_result = SelectedMIHeater.objects.create(
            line=line,
            heater=heater,
            cold_lead_option=cold_lead,
            selection_status='available_alternative',
            cold_lead_option_code=cold_lead.option_code,
            cold_lead_length_m=cold_lead.length_m,
            heated_length_m=106.0,
            heater_resistance_ohms=10.6,
            cold_lead_resistance_total_ohms=0.04,
            power_nominal_w=5235.8,
            power_density_w_m=49.39,
            current_nominal_a=21.54,
            current_cold_start_a=21.54,
            max_sheath_temp_published_c=180.0,
            project_t_class_limit_c=200.0,
            t_class_verdict='pass',
            selection_basis={'selection_mode': 'available_alternative'},
        )

        adjusted_payload = build_project_sld_payload('p1')
        adjusted_node = next(node for node in adjusted_payload['nodes'] if node['component_id'] == tracer_node['component_id'])
        tracer_selection = adjusted_node['metadata']['tracer_selection']

        self.assertNotEqual(refreshed_result.id, old_result_id)
        self.assertEqual(tracer_selection['selected']['v_uid'], expected_uid)
        self.assertTrue(tracer_selection['override_active'])
        self.assertEqual(tracer_selection['selected']['power_output'], 49.39)

    def test_sld_payload_view_returns_json_graph_payload(self):
        make_calculated_project_snapshot()

        response = self.client.get(reverse('sld_payload_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['project_id'], 'p1')
        self.assertEqual(payload['schema_version'], SLD_GRAPH_SCHEMA_VERSION)
        self.assertEqual(payload['meta']['branch_count'], 1)
        self.assertEqual(payload['meta']['node_count'], 4)
        self.assertEqual(payload['meta']['edge_count'], 3)
        self.assertIn('nodes', payload)
        self.assertIn('edges', payload)
        self.assertTrue(any(node['display_tag'] == 'MCB_001' for node in payload['nodes']))

    def test_sld_payload_view_supports_line_filtering(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        full_payload = build_project_sld_payload('p1')
        line_two_component_ids = {
            node['component_id']
            for node in full_payload['nodes']
            if 'LINE-002' in node.get('line_ids', [])
        }

        response = self.client.get(
            reverse('sld_payload_view'),
            {'project_id': 'p1', 'line_id': 'line-002'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['schema_version'], SLD_GRAPH_SCHEMA_VERSION)
        self.assertEqual([group['line_id'] for group in payload['line_groups']], ['LINE-002'])
        self.assertEqual(payload['meta']['branch_count'], 1)
        self.assertEqual(payload['meta']['node_count'], len(line_two_component_ids))
        self.assertEqual(set(node['component_id'] for node in payload['nodes']), line_two_component_ids)
        self.assertTrue(all('LINE-002' in node.get('line_ids', []) for node in payload['nodes']))
        self.assertTrue(all('LINE-002' in edge.get('line_ids', []) for edge in payload['edges']))

    def test_sld_payload_view_keeps_shared_manual_components_in_focused_line(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])
        generated_payload = build_project_sld_payload('p1')
        mcb_component_ids = [
            node['component_id']
            for node in generated_payload['nodes']
            if node['component_type'] == 'MCB'
        ]
        response = self.client.post(
            reverse('sld_topology_combine_apply_view'),
            data=json.dumps({'project_id': 'p1', 'component_ids': mcb_component_ids}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)

        focused_response = self.client.get(
            reverse('sld_payload_view'),
            {'project_id': 'p1', 'line_id': 'line-002'},
        )

        self.assertEqual(focused_response.status_code, 200)
        payload = focused_response.json()
        tags = {node['display_tag'] for node in payload['nodes']}
        self.assertIn('MCB_001-M', tags)
        self.assertTrue(any(tag.startswith('JB3PH_001') for tag in tags))

    def test_sld_payload_view_supports_partial_line_filtering(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001', 'LINE-002'])

        response = self.client.get(
            reverse('sld_payload_view'),
            {'project_id': 'p1', 'line_id': '002'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([group['line_id'] for group in payload['line_groups']], ['LINE-002'])
        self.assertTrue(payload['nodes'])

    def test_sld_payload_view_line_filter_keeps_duplicate_line_ids_distinct(self):
        lines = make_rich_sld_project_snapshot('p1', ['LINE-DUP', 'LINE-DUP'])
        expected_line_uids = {str(line.uid) for line in lines}

        response = self.client.get(
            reverse('sld_payload_view'),
            {'project_id': 'p1', 'line_id': 'line-dup'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['meta']['branch_count'], 2)
        self.assertEqual([group['line_id'] for group in payload['line_groups']], ['LINE-DUP', 'LINE-DUP'])
        self.assertEqual({group['line_uid'] for group in payload['line_groups']}, expected_line_uids)
        self.assertEqual({node['line_uid'] for node in payload['nodes']}, expected_line_uids)
        self.assertEqual({edge['line_uid'] for edge in payload['edges']}, expected_line_uids)
        self.assertEqual(
            len([node['component_id'] for node in payload['nodes']]),
            len({node['component_id'] for node in payload['nodes']}),
        )

    def test_sld_payload_view_reports_unknown_line_filter(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(
            reverse('sld_payload_view'),
            {'project_id': 'p1', 'line_id': 'LINE-999'},
        )

        self.assertEqual(response.status_code, 404)
        self.assertIn('LINE-999', response.json()['error'])

    def test_sld_payload_view_does_not_build_layout_or_validation(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        with (
            patch('eht.views.get_project_sld_layout') as layout_mock,
            patch('eht.views.validate_project_sld_payload') as validation_mock,
        ):
            response = self.client.get(reverse('sld_payload_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        layout_mock.assert_not_called()
        validation_mock.assert_not_called()

    def test_sld_validation_view_returns_project_report(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        response = self.client.get(reverse('sld_validation_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        report = response.json()
        self.assertEqual(report['project_id'], 'p1')
        self.assertEqual(report['status'], 'passed')
        self.assertIn('summary', report)
        self.assertIn('checks', report)
        self.assertIn('branch_checks', report)

    def test_sld_validation_view_does_not_load_layout(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        with patch('eht.views.get_project_sld_layout') as layout_mock:
            response = self.client.get(reverse('sld_validation_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        layout_mock.assert_not_called()

    def test_sld_layout_view_does_not_run_validation(self):
        make_rich_sld_project_snapshot('p1', ['LINE-001'])

        with patch('eht.views.validate_project_sld_payload') as validation_mock:
            response = self.client.get(reverse('sld_layout_view'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        validation_mock.assert_not_called()


class ProjectWorkspaceCleanupTests(TestCase):
    def test_clear_project_workspace_data_removes_input_and_derived_outputs(self):
        line = make_calculated_project_snapshot()

        cleanup_summary = clear_project_workspace_data('p1')

        self.assertEqual(cleanup_summary['project_id'], 'p1')
        self.assertGreaterEqual(cleanup_summary['input_lines'], 1)
        self.assertFalse(HeatTracingInput.objects.filter(proj_id='p1').exists())
        self.assertFalse(HeatLoss.objects.filter(uid=str(line.uid)).exists())
        self.assertFalse(SelectedTracer.objects.filter(line_id=line.uid).exists())
        self.assertFalse(AlternateTracer.objects.filter(line_id=line.uid).exists())
        self.assertFalse(PowerDistribution.objects.filter(uid=str(line.uid)).exists())
        self.assertFalse(ProcessLineCalculation.objects.filter(uid=str(line.uid)).exists())
        self.assertFalse(PowerDistributionBranch.objects.exists())
        self.assertFalse(BOQ.objects.filter(project_id='p1').exists())


class ConfirmValidDataViewTests(TransactionTestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='tester', password='password123')
        self.client.force_login(self.user)
        make_project_record()
        seed_reference_data()

    def create_pending_line(self, **overrides):
        defaults = {
            'proj_id': 'p1',
            'line_id': 'LINE-001',
            'service_type': 'EP',
            'line_size': 2.0,
            'line_length': 10.0,
            'ins_mat_type': 'Mineral Wool',
            'insul_thick': 50.0,
            'maint_temp': 100.0,
            'oper_temp': 80.0,
            'design_temp': 120.0,
            'valve_qty': 0,
            'flange_qty': 0,
            'support_qty': 0,
            'status': 'pending',
        }
        defaults.update(overrides)
        return HeatTracingInput.objects.create(**defaults)

    def test_confirm_valid_data_processes_pending_rows_and_stores_results(self):
        line = self.create_pending_line()

        response = self.client.post(reverse('confirm_valid_data'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['project_id'], 'p1')
        self.assertEqual(payload['confirmed_rows'], 1)
        self.assertEqual(payload['result_counts']['heat_loss'], 1)
        self.assertEqual(payload['result_counts']['selected_tracers'], 1)
        self.assertEqual(payload['result_counts']['boq_lines'], 1)

        line.refresh_from_db()
        self.assertEqual(line.status, 'confirmed')
        self.assertTrue(HeatLoss.objects.filter(line=line).exists())

    def test_confirm_valid_data_runs_calculations_outside_confirmation_transaction(self):
        line = self.create_pending_line()

        def fake_run_project_calculations(project_id):
            self.assertFalse(connection.in_atomic_block)
            return (
                {'heat_loss': []},
                {
                    'heat_loss': 0,
                    'selected_tracers': 0,
                    'alternative_tracers': 0,
                    'power_distribution': 0,
                    'boq_lines': 0,
                    'consolidated_boq_items': 0,
                    'tracer_power_param': 0,
                },
            )

        with patch('eht.views.run_project_calculations', side_effect=fake_run_project_calculations):
            response = self.client.post(reverse('confirm_valid_data'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 200)
        line.refresh_from_db()
        self.assertEqual(line.status, 'confirmed')


class CalculateViewHardeningTests(TransactionTestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='uploadtester', password='password123')
        self.client.force_login(self.user)
        make_project_record()

    def create_pending_line(self, **overrides):
        defaults = {
            'proj_id': 'p1',
            'line_id': 'LINE-001',
            'service_type': 'EP',
            'line_size': 2.0,
            'line_length': 10.0,
            'ins_mat_type': 'Mineral Wool',
            'insul_thick': 50.0,
            'maint_temp': 100.0,
            'oper_temp': 80.0,
            'design_temp': 120.0,
            'valve_qty': 0,
            'flange_qty': 0,
            'support_qty': 0,
            'status': 'pending',
        }
        defaults.update(overrides)
        return HeatTracingInput.objects.create(**defaults)

    def test_calculate_view_does_not_clear_existing_workspace_when_upload_has_no_valid_rows(self):
        existing_line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-EXISTING',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )
        upload = SimpleUploadedFile(
            'input.xlsx',
            b'fake-xlsx-content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        with patch('eht.views.sanitize_file', return_value=([], [{'row_number': 2, 'errors': ['bad row']}], '/tmp/error.xlsx')):
            response = self.client.post(
                reverse('calculate_view'),
                {'project_id': 'p1', 'file': upload},
            )

        self.assertEqual(response.status_code, 400)
        self.assertTrue(HeatTracingInput.objects.filter(pk=existing_line.pk).exists())
        self.assertEqual(HeatTracingInput.objects.filter(proj_id='p1').count(), 1)

    def test_calculate_view_replaces_workspace_only_after_valid_rows_are_ready(self):
        stale_line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-OLD',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )
        HeatLoss.objects.create(uid=str(stale_line.uid), line=stale_line, heat_loss=12.5, tracer_adder=1.2)

        valid_rows = [{
            'XLID': 1,
            'Line_ID': 'LINE-NEW',
            'Service_Type': 'EP',
            'Line_Size': 2.0,
            'Line_Length': 11.0,
            'Ins_Mat_Type': 'Mineral Wool',
            'Insul_Thick': 50.0,
            'Maint_T': 120.0,
            'Oper_T': 100.0,
            'Design_T': 140.0,
            'IsDeleted': False,
            'PID_No': '',
            'Area': '',
            'Train': '',
            'Valve_Qty': 0,
            'Flange_Qty': 0,
            'Support_Qty': 0,
            'Pipe_Mat_Class': '',
            'Emergency_Supply': False,
            'Discipline': '',
            'Remarks': '',
        }]
        upload = SimpleUploadedFile(
            'input.xlsx',
            b'fake-xlsx-content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        with patch('eht.views.sanitize_file', return_value=(valid_rows, [{'row_number': 3, 'errors': ['bad row']}], '/tmp/error.xlsx')):
            response = self.client.post(
                reverse('calculate_view'),
                {'project_id': 'p1', 'file': upload},
            )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(HeatTracingInput.objects.filter(pk=stale_line.pk).exists())
        self.assertFalse(HeatLoss.objects.filter(uid=str(stale_line.uid)).exists())
        replacement_line = HeatTracingInput.objects.get(proj_id='p1', line_id='LINE-NEW')
        self.assertEqual(replacement_line.status, 'pending')
        self.assertEqual(HeatTracingInput.objects.filter(proj_id='p1').count(), 1)

    def test_calculate_view_runs_calculations_outside_upload_transaction(self):
        valid_rows = [{
            'XLID': 1,
            'Line_ID': 'LINE-NEW',
            'Service_Type': 'EP',
            'Line_Size': 2.0,
            'Line_Length': 11.0,
            'Ins_Mat_Type': 'Mineral Wool',
            'Insul_Thick': 50.0,
            'Maint_T': 120.0,
            'Oper_T': 100.0,
            'Design_T': 140.0,
            'IsDeleted': False,
            'PID_No': '',
            'Area': '',
            'Train': '',
            'Valve_Qty': 0,
            'Flange_Qty': 0,
            'Support_Qty': 0,
            'Pipe_Mat_Class': '',
            'Emergency_Supply': False,
            'Discipline': '',
            'Remarks': '',
        }]
        upload = SimpleUploadedFile(
            'input.xlsx',
            b'fake-xlsx-content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        def fake_run_project_calculations(project_id):
            self.assertFalse(connection.in_atomic_block)
            return ({'heat_loss': []}, {'heat_loss': 0, 'selected_tracers': 0, 'alternative_tracers': 0, 'power_distribution': 0, 'boq_lines': 0, 'consolidated_boq_items': 0, 'tracer_power_param': 0})

        with patch('eht.views.sanitize_file', return_value=(valid_rows, [], '')), patch(
            'eht.views.run_project_calculations',
            side_effect=fake_run_project_calculations,
        ):
            response = self.client.post(
                reverse('calculate_view'),
                {'project_id': 'p1', 'file': upload},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(HeatTracingInput.objects.get(proj_id='p1').status, 'confirmed')

    def test_calculate_view_saves_visible_project_setup_before_running_calculation(self):
        valid_rows = [{
            'XLID': 1,
            'Line_ID': 'LINE-NEW',
            'Service_Type': 'EP',
            'Line_Size': 2.0,
            'Line_Length': 11.0,
            'Ins_Mat_Type': 'Mineral Wool',
            'Insul_Thick': 50.0,
            'Maint_T': 120.0,
            'Oper_T': 100.0,
            'Design_T': 140.0,
            'IsDeleted': False,
            'PID_No': '',
            'Area': '',
            'Train': '',
            'Valve_Qty': 0,
            'Flange_Qty': 0,
            'Support_Qty': 0,
            'Pipe_Mat_Class': '',
            'Emergency_Supply': False,
            'Discipline': '',
            'Remarks': '',
        }]
        upload = SimpleUploadedFile(
            'input.xlsx',
            b'fake-xlsx-content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        ManagedProject.objects.get(proj_id='p1').assigned_users.add(self.user)
        payload = make_project_form_payload(proj_id='p1', vendor='SST', voltage='240.00')
        payload.update({'project_id': 'p1', 'file': upload})

        def fake_run_project_calculations(project_id):
            self.assertEqual(ProjectData.objects.get(proj_id=project_id).vendor, 'SST')
            return ({'heat_loss': []}, {'heat_loss': 0, 'selected_tracers': 0, 'alternative_tracers': 0, 'power_distribution': 0, 'boq_lines': 0, 'consolidated_boq_items': 0, 'tracer_power_param': 0})

        with patch('eht.views.sanitize_file', return_value=(valid_rows, [], '')), patch(
            'eht.views.run_project_calculations',
            side_effect=fake_run_project_calculations,
        ):
            response = self.client.post(reverse('calculate_view'), payload)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(ProjectData.objects.get(proj_id='p1').vendor, 'SST')

    def test_calculate_view_backfills_missing_cold_cable_fields_from_stale_project_form(self):
        valid_rows = [{
            'XLID': 1,
            'Line_ID': 'LINE-NEW',
            'Service_Type': 'EP',
            'Line_Size': 2.0,
            'Line_Length': 11.0,
            'Ins_Mat_Type': 'Mineral Wool',
            'Insul_Thick': 50.0,
            'Maint_T': 120.0,
            'Oper_T': 100.0,
            'Design_T': 140.0,
            'IsDeleted': False,
            'PID_No': '',
            'Area': '',
            'Train': '',
            'Valve_Qty': 0,
            'Flange_Qty': 0,
            'Support_Qty': 0,
            'Pipe_Mat_Class': '',
            'Emergency_Supply': False,
            'Discipline': '',
            'Remarks': '',
        }]
        upload = SimpleUploadedFile(
            'input.xlsx',
            b'fake-xlsx-content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        ManagedProject.objects.get(proj_id='p1').assigned_users.add(self.user)
        payload = make_project_form_payload(proj_id='p1', vendor='SST', voltage='240.00')
        for field_name in PROJECT_FORM_COLD_CABLE_DEFAULTS:
            payload.pop(field_name, None)
        payload.pop('rcd_provided', None)
        payload.update({'project_id': 'p1', 'file': upload})

        def fake_run_project_calculations(project_id):
            project = ProjectData.objects.get(proj_id=project_id)
            self.assertEqual(project.cable_standard, 'IEC_60502_1')
            self.assertEqual(project.cable_conductor_material, 'Cu')
            self.assertEqual(project.cable_insulation_type, 'XLPE')
            self.assertEqual(project.cable_install_method, 'E')
            self.assertEqual(float(project.cable_grouping_derating), 1.0)
            self.assertEqual(project.min_cold_cable_size_mm2, 'CALCULATED')
            self.assertEqual(project.mcb_curve, 'C')
            self.assertTrue(project.rcd_provided)
            return ({'heat_loss': []}, {'heat_loss': 0, 'selected_tracers': 0, 'alternative_tracers': 0, 'power_distribution': 0, 'boq_lines': 0, 'consolidated_boq_items': 0, 'tracer_power_param': 0})

        with patch('eht.views.sanitize_file', return_value=(valid_rows, [], '')), patch(
            'eht.views.run_project_calculations',
            side_effect=fake_run_project_calculations,
        ):
            response = self.client.post(reverse('calculate_view'), payload)

        self.assertEqual(response.status_code, 200)

    def test_calculate_view_rejects_mismatched_visible_project_setup(self):
        existing_line = HeatTracingInput.objects.create(
            proj_id='p1',
            line_id='LINE-EXISTING',
            service_type='EP',
            line_size=2.0,
            line_length=10.0,
            ins_mat_type='Mineral Wool',
            insul_thick=50.0,
            maint_temp=120.0,
            oper_temp=100.0,
            design_temp=140.0,
            status='confirmed',
        )
        upload = SimpleUploadedFile(
            'input.xlsx',
            b'fake-xlsx-content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        payload = make_project_form_payload(proj_id='p2', vendor='SST')
        payload.update({'project_id': 'p1', 'file': upload})

        response = self.client.post(reverse('calculate_view'), payload)

        self.assertEqual(response.status_code, 400)
        self.assertIn('does not match', response.json()['error'])
        self.assertTrue(HeatTracingInput.objects.filter(pk=existing_line.pk).exists())
        self.assertEqual(ProjectData.objects.get(proj_id='p1').vendor, 'CHR')

    def test_confirm_valid_data_rejects_requests_when_no_rows_are_pending(self):
        line = self.create_pending_line(status='confirmed')

        response = self.client.post(reverse('confirm_valid_data'), {'project_id': 'p1'})

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'No valid uploaded data is pending confirmation.')
        line.refresh_from_db()
        self.assertEqual(line.status, 'confirmed')
        self.assertFalse(HeatLoss.objects.filter(line=line).exists())
