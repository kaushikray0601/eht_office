from openpyxl.utils import get_column_letter


def polish_openpyxl_workbook(writer, *, freeze_panes='A2', max_width=48):
    """Apply small usability polish to all sheets written by pandas/openpyxl."""
    workbook = getattr(writer, 'book', None)
    if workbook is None:
        return

    for worksheet in workbook.worksheets:
        if worksheet.max_row and worksheet.max_column:
            worksheet.freeze_panes = freeze_panes
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            width = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                width = max(width, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max(width + 2, 10), max_width)
